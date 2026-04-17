from __future__ import annotations

import re
from calendar import month_name, monthrange
from collections import Counter
from collections.abc import Sequence
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from binge_schedule.models import Segment, WeekDef


def _safe_excel_sheet_title(name: str) -> str:
    """Excel tab name: max 31 chars, no []:*?/\\."""
    s = "".join(ch for ch in str(name).strip() if ch not in "[]:*?/\\")[:31]
    return s or "Week"


def ensure_grids_workbooks_for_weeks(weeks: Sequence[WeekDef]) -> list[str]:
    """Create any missing **grids** workbooks referenced by ``weeks`` (blank program area).

    Each workbook gets one sheet per ``sheet_name`` in the order given. At export, ``seed_grids_from_prior_month``
    can copy April’s show titles into a new month before building the BINGE list.
    """
    files_order: list[str] = []
    by_file: dict[str, list] = {}
    for w in weeks:
        key = str(Path(w.grids_file).expanduser().resolve())
        if key not in by_file:
            by_file[key] = []
            files_order.append(key)
        by_file[key].append(w)

    created: list[str] = []
    for key in files_order:
        path = Path(key)
        if path.is_file():
            continue
        path.parent.mkdir(parents=True, exist_ok=True)
        wlist = by_file[key]
        wb = Workbook()
        ws = wb.active
        seen_yaml_sheet: set[str] = set()
        tab_titles: list[str] = []
        used_safe: set[str] = set()
        for w in wlist:
            if w.sheet_name in seen_yaml_sheet:
                continue
            seen_yaml_sheet.add(w.sheet_name)
            safe = _safe_excel_sheet_title(w.sheet_name)
            if safe in used_safe:
                raise ValueError(
                    f"Grids workbook {path}: two different ``sheet_name`` values map to the same "
                    f"Excel tab {safe!r} after sanitization/truncation. Rename tabs in YAML so each "
                    "tab is unique within the first 31 characters."
                )
            used_safe.add(safe)
            tab_titles.append(safe)
        if not tab_titles:
            tab_titles = ["Week"]
        ws.title = tab_titles[0]
        for sn in tab_titles[1:]:
            wb.create_sheet(sn)
        # Pandas drops an all-empty used range; two NBSP corners force a 52×8 frame. ``strip()`` removes them so
        # the Mon–Sun program block (rows 5–52, cols B–H) reads as blank.
        _pad = "\u00a0"
        for sheet in wb.worksheets:
            sheet.cell(row=1, column=1, value=_pad)
            sheet.cell(row=52, column=8, value=_pad)
        wb.save(path)
        created.append(str(path))
    return created


def load_grid_sheet(path: str, sheet_name: str) -> list[list[Optional[str]]]:
    """Return 48 rows × 7 columns of stripped strings or None."""
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    block = df.iloc[4:52, 1:8]
    rows: list[list[Optional[str]]] = []
    for r in range(48):
        row: list[Optional[str]] = []
        for c in range(7):
            v = block.iloc[r, c]
            if pd.isna(v):
                row.append(None)
            else:
                s = str(v).strip()
                row.append(s if s else None)
        rows.append(row)
    return rows


def _is_empty(v: Optional[str]) -> bool:
    return v is None or not str(v).strip()


def segments_for_day(col: list[Optional[str]]) -> list[Segment]:
    """Split one weekday column into segments.

    - A **block** is a filled cell followed only by empty cells until the next fill.
    - A **single** is a filled cell whose immediate next row is also filled (typical strip).
    """
    if len(col) != 48:
        raise ValueError(f"Expected 48 slots, got {len(col)}")
    out: list[Segment] = []
    i = 0
    while i < 48:
        if _is_empty(col[i]):
            i += 1
            continue
        title = col[i]  # type: ignore
        if i + 1 < 48 and _is_empty(col[i + 1]):
            j = i + 1
            while j < 48 and _is_empty(col[j]):
                j += 1
            out.append(Segment(i, j, title))
            i = j
        else:
            out.append(Segment(i, i + 1, title))
            i += 1
    return out


def day_dates(week_monday: date) -> list[date]:
    return [week_monday + timedelta(days=d) for d in range(7)]


def slot_clock_to_time(slot: int) -> time:
    if not (0 <= slot < 48):
        raise ValueError(slot)
    base = datetime(2000, 1, 1, 0, 0) + timedelta(minutes=30 * slot)
    return base.time()


def slot_label(slot: int) -> str:
    """Half-hour slot as plain text (e.g. 0:00, 23:30) — no Excel time serial."""
    if not (0 <= slot < 48):
        raise ValueError(slot)
    minutes = 30 * slot
    h, m = divmod(minutes, 60)
    return f"{h}:{m:02d}"


def combine_date_time(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute, t.second)


def parse_monday(s: str) -> date:
    return date.fromisoformat(s.strip())


def week_overlaps_calendar_month(monday: date, year: int, month: int) -> bool:
    """True if this Mon–Sun week shares any calendar day with ``year``/``month``."""
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    week_end = monday + timedelta(days=6)
    return week_end >= start and monday <= end


def grid_program_all_empty(grid: list[list[Optional[str]]]) -> bool:
    return all(_is_empty(v) for row in grid for v in row)


def _month_weeks_sorted(all_weeks: Sequence[WeekDef], year: int, month: int) -> list[WeekDef]:
    out: list[WeekDef] = []
    for w in all_weeks:
        d = parse_monday(w.monday)
        if d.year == year and d.month == month:
            out.append(w)
    out.sort(key=lambda w: w.monday)
    return out


def weeks_with_monday_in_calendar_month(
    weeks: Sequence[WeekDef], year: int, month: int
) -> list[WeekDef]:
    """Weeks whose anchor Monday lies in ``year``/``month``.

    Excludes straddle ISO weeks whose Monday is still in the previous month (e.g. Mon 2026-04-27 when building
    **May**), so monthly exports start with the **first Monday of that month** when combined with ``weeks:`` entries.
    """
    return _month_weeks_sorted(weeks, year, month)


def _reference_week_prior_month(wk: WeekDef, all_weeks: Sequence[WeekDef]) -> Optional[WeekDef]:
    d = parse_monday(wk.monday)
    if d.month == 1:
        py, pm = d.year - 1, 12
    else:
        py, pm = d.year, d.month - 1
    prev = _month_weeks_sorted(all_weeks, py, pm)
    if not prev:
        return None
    curr = _month_weeks_sorted(all_weeks, d.year, d.month)
    idx = next((i for i, w in enumerate(curr) if w.monday == wk.monday), -1)
    if idx < 0:
        return None
    return prev[idx] if idx < len(prev) else prev[-1]


def _find_grid_worksheet(wb, sheet_name: str):
    want = _safe_excel_sheet_title(sheet_name)
    for ws in wb.worksheets:
        if ws.title == sheet_name or _safe_excel_sheet_title(ws.title) == want:
            return ws
    raise KeyError(
        f"No worksheet matching sheet_name {sheet_name!r} (sanitized tab title {want!r}). "
        f"Existing tabs: {[ws.title for ws in wb.worksheets]!r}"
    )


def _write_program_cells_on_sheet(ws, grid: list[list[Optional[str]]]) -> None:
    for slot in range(48):
        for day in range(7):
            v = grid[slot][day]
            ws.cell(row=5 + slot, column=2 + day, value=v)


def seed_grids_from_prior_month(
    week_list: Sequence[WeekDef], all_weeks: Sequence[WeekDef]
) -> list[str]:
    """Copy Mon–Sun program text from the **previous calendar month** when the target week’s grid is still blank.

    Week *k* within a month (sorted by Monday) maps to week *k* in the prior month (same index; if there are fewer
    weeks in the prior month, the last week is reused). Writes in place to each ``grids_file`` so the Nikki-driven
    export sees the same strip pattern as last month on new dates.
    """
    messages: list[str] = []
    pending: dict[str, list[tuple[WeekDef, WeekDef, list[list[Optional[str]]]]]] = {}
    for wk in week_list:
        path = Path(wk.grids_file)
        if not path.is_file():
            continue
        try:
            cur = load_grid_sheet(str(path), wk.sheet_name)
        except Exception as exc:  # noqa: BLE001
            messages.append(f"{wk.sheet_name}: could not read grid ({exc}); skipping seed.")
            continue
        if not grid_program_all_empty(cur):
            continue
        ref = _reference_week_prior_month(wk, all_weeks)
        if ref is None:
            messages.append(
                f"{wk.monday} ({wk.sheet_name!r}): no ``weeks:`` entries for the previous calendar month "
                "in your setup file - add that month or paste the strip in Excel."
            )
            continue
        ref_path = Path(ref.grids_file)
        if not ref_path.is_file():
            messages.append(
                f"{wk.monday}: would copy from {ref.monday}, but that month's grids file is missing: {ref.grids_file}"
            )
            continue
        try:
            src = load_grid_sheet(str(ref_path), ref.sheet_name)
        except Exception as exc:  # noqa: BLE001
            messages.append(f"{wk.monday}: cannot load reference grid {ref.sheet_name!r}: {exc}")
            continue
        if grid_program_all_empty(src):
            messages.append(
                f"{wk.monday}: reference week {ref.monday} tab {ref.sheet_name!r} has no program text; not seeding."
            )
            continue
        key = str(path.expanduser().resolve())
        pending.setdefault(key, []).append((wk, ref, src))

    for path_str, triples in pending.items():
        path = Path(path_str)
        wb = load_workbook(path)
        for wk, ref, src in triples:
            ws = _find_grid_worksheet(wb, wk.sheet_name)
            _write_program_cells_on_sheet(ws, src)
            messages.append(
                f"Copied program grid from **{ref.monday}** (`{Path(ref.grids_file).name}` / `{ref.sheet_name}`) "
                f"into **{wk.monday}** (`{path.name}` / `{wk.sheet_name}`)."
            )
        wb.save(path)
    return messages


def _infer_primary_calendar_month(week_list: Sequence[WeekDef]) -> Optional[tuple[int, int]]:
    """Calendar month that most selected weeks start in (e.g. May build picks (2026, 5) even with one Apr Monday)."""
    if not week_list:
        return None
    c = Counter()
    for w in week_list:
        d = parse_monday(w.monday)
        c[(d.year, d.month)] += 1
    (y, m), _ = c.most_common(1)[0]
    return (y, m)


def sync_straddle_weeks_to_canonical_grids_file(week_list: Sequence[WeekDef]) -> list[str]:
    """Put straddle weeks on the same grids workbook as the rest of the target month.

    Example: building **May** selects Mon 2026-04-27 (Apr 27–May 3). That week usually lives in the April grids file
    in YAML; copy its program block into the May grids workbook so May 1–3 sit beside the other May tabs on disk.
    """
    messages: list[str] = []
    primary = _infer_primary_calendar_month(week_list)
    if not primary:
        return messages
    py, pm = primary
    in_month = [
        w
        for w in week_list
        if parse_monday(w.monday).year == py and parse_monday(w.monday).month == pm
    ]
    if not in_month:
        return messages
    canonical = Path(in_month[0].grids_file).expanduser().resolve()

    for wk in week_list:
        mon = parse_monday(wk.monday)
        if not week_overlaps_calendar_month(mon, py, pm):
            continue
        src_path = Path(wk.grids_file).expanduser().resolve()
        if src_path == canonical:
            continue
        try:
            src_grid = load_grid_sheet(str(src_path), wk.sheet_name)
        except Exception as exc:  # noqa: BLE001
            messages.append(f"{wk.sheet_name}: could not load straddle source ({exc}); skip copy to {canonical.name}.")
            continue
        if grid_program_all_empty(src_grid):
            continue
        if not canonical.is_file():
            messages.append(f"Cannot copy {wk.sheet_name} — canonical grids file missing: {canonical}")
            continue
        wb = load_workbook(canonical)
        try:
            ws = _find_grid_worksheet(wb, wk.sheet_name)
        except KeyError:
            safe = _safe_excel_sheet_title(wk.sheet_name)
            ws = wb.create_sheet(safe)
            _pad = "\u00a0"
            ws.cell(row=1, column=1, value=_pad)
            ws.cell(row=52, column=8, value=_pad)
        _write_program_cells_on_sheet(ws, src_grid)
        wb.save(canonical)
        mn = month_name[pm]
        messages.append(
            f"Copied straddle week {wk.monday} from `{src_path.name}` into `{canonical.name}` (tab `{wk.sheet_name}`) "
            f"so {mn} days (e.g. early-month days on Mon Sun week) appear in that month's grids file."
        )
    return messages


def parse_sheet_tab_monday(title: str) -> Optional[date]:
    """Parse a grid worksheet tab into that week's calendar anchor date (Monday in ISO configs).

    Accepts ``YYYY-MM-DD`` or legacy labels like ``4-6-2026`` (month-day-year).
    Returns ``None`` if the title does not match a known pattern.
    """
    title = title.strip()
    if not title:
        return None
    try:
        return date.fromisoformat(title)
    except ValueError:
        pass
    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})-(\d{4})", title)
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None
