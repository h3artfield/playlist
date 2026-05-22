from __future__ import annotations

from datetime import datetime
from datetime import date, timedelta
from io import BytesIO
import json
from pathlib import Path
import os
import shutil
import sys
import threading
from typing import Any, Optional
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from pydantic import BaseModel, Field
import yaml

from binge_schedule.config_io import load_build_config
from binge_schedule.content_catalog import canonical_rows_from_config
from binge_schedule.rule_analyzer import analyze_schedule_rules
from binge_schedule.schedule_blocks import blocks_to_week_grid, empty_slots_for_blocks, grid_to_blocks


DEFAULT_CONFIG = Path(os.environ.get("SCHEDULE_BUILDER_DEFAULT_CONFIG", "config/april_2026.yaml"))


class BlocksPayload(BaseModel):
    blocks: list[dict[str, Any]] = Field(default_factory=list)
    catalog_rows: list[dict[str, Any]] = Field(default_factory=list)


class BlocksToGridPayload(BaseModel):
    week_monday: date
    blocks: list[dict[str, Any]] = Field(default_factory=list)
    require_complete: bool = False


class GridToBlocksPayload(BaseModel):
    week_monday: date
    grid: list[list[Optional[str]]]


class SaveBaseSchedulePayload(BaseModel):
    station_id: str
    week_monday: date
    week_count: int = 1
    blocks: list[dict[str, Any]] = Field(default_factory=list)
    suggested_rules: list[dict[str, Any]] = Field(default_factory=list)
    notes: list[dict[str, Any]] = Field(default_factory=list)
    saved_directory: str = ""


class DownloadSchedulePayload(BaseModel):
    station_id: str = ""
    week_monday: date
    week_count: int = 1
    blocks: list[dict[str, Any]] = Field(default_factory=list)
    notes: list[dict[str, Any]] = Field(default_factory=list)


class AutoGeneratePayload(BaseModel):
    base_path: str = ""
    week_count: int = 1


class BaseSchedulePathPayload(BaseModel):
    path: str = ""


class AppSettingsPayload(BaseModel):
    theme: str = "dark"
    accent_primary: str = "#2563eb"
    accent_secondary: str = "#7c3aed"
    primary_save_directory: str = ""
    backup_save_directory: str = ""
    backup_enabled: bool = True
    desktop_window_mode: str = "windowed"


class PickDirectoryPayload(BaseModel):
    kind: str = "primary"


class ImportContentRowPayload(BaseModel):
    content_type: str = "series"
    show_name: str = ""
    episode_number: str = ""
    episode_title: str = ""
    runtime_minutes: Optional[int] = None
    slot_minutes: Optional[int] = None
    genre: str = ""


class ImportContentRowsPayload(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list)


class UpdateShowRowsPayload(BaseModel):
    display_name: str
    rows: list[dict[str, Any]] = Field(default_factory=list)


class RenameShowPayload(BaseModel):
    display_name: str = ""
    new_display_name: str = ""


class ShowDisplayNamePayload(BaseModel):
    display_name: str = ""


class ImportPreviewSheetConfig(BaseModel):
    sheet_name: str
    include: bool = True
    header_row: int = Field(default=1, ge=0, le=50)
    row_kind: str = "auto"
    default_series_title: str = ""
    mapping: dict[str, str] = Field(default_factory=dict)
    layout: str = "header"
    data_start_row: int = Field(default=1, ge=1)
    inferred_column_names: list[str] = Field(default_factory=list)


class ImportPreviewPayload(BaseModel):
    session_id: str
    sheets: list[ImportPreviewSheetConfig] = Field(default_factory=list)


class ImportCommitPayload(BaseModel):
    session_id: str
    sheets: list[ImportPreviewSheetConfig] = Field(default_factory=list)


class ImportSheetAnalyzePayload(BaseModel):
    session_id: str
    sheet_name: str
    header_row: int = Field(default=1, ge=0, le=50)


class ImportSampleRowsPayload(BaseModel):
    session_id: str
    sheet: ImportPreviewSheetConfig


def create_app() -> FastAPI:
    app = FastAPI(title="Playlist Schedule Builder API", version="0.1.0")
    app.add_middleware(
        CORSMiddleware,
        allow_origins=[
            "http://127.0.0.1:5173",
            "http://localhost:5173",
            "http://127.0.0.1:4173",
            "http://localhost:4173",
            "http://127.0.0.1:8765",
            "http://localhost:8765",
        ],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.get("/api/health")
    def health() -> dict[str, Any]:
        from binge_schedule.app_settings import primary_saved_schedules_root
        from binge_schedule.runtime_paths import content_import_wizard_available, is_desktop_runtime

        return {
            "status": "ok",
            "features": {
                "auto_generate_weeks": True,
                "auto_generate_date_shift": True,
                "content_import_wizard": content_import_wizard_available(),
                "desktop_runtime": is_desktop_runtime(),
                "app_settings": True,
                "export_to_downloads": True,
            },
            "primary_save_directory": primary_saved_schedules_root().as_posix(),
        }

    @app.get("/api/desktop-download")
    def desktop_download_info() -> dict[str, Any]:
        from binge_schedule.legal import desktop_download_meta
        from binge_schedule.runtime_paths import is_desktop_runtime

        meta = desktop_download_meta()
        return {
            **meta,
            "desktop_runtime": is_desktop_runtime(),
            "show_download_cta": not is_desktop_runtime(),
        }

    @app.get("/api/settings")
    def get_settings() -> dict[str, Any]:
        from binge_schedule.app_settings import settings_for_api

        return settings_for_api()

    def _save_settings(payload: AppSettingsPayload) -> dict[str, Any]:
        from binge_schedule.app_settings import save_settings, settings_for_api

        try:
            save_settings(payload.model_dump())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        saved = settings_for_api()
        if saved.get("desktop_runtime"):
            from binge_schedule.desktop_window import apply_desktop_window_mode as apply_mode

            mode = str(saved.get("desktop_window_mode") or "windowed")
            saved["desktop_window_applied"] = apply_mode(mode)
        return saved

    @app.put("/api/settings")
    def put_settings(payload: AppSettingsPayload) -> dict[str, Any]:
        return _save_settings(payload)

    @app.post("/api/settings")
    def post_settings(payload: AppSettingsPayload) -> dict[str, Any]:
        return _save_settings(payload)

    @app.post("/api/settings/pick-directory")
    def pick_settings_directory(payload: PickDirectoryPayload) -> dict[str, str]:
        from binge_schedule.app_settings import pick_directory_dialog

        kind = (payload.kind or "primary").strip().lower()
        title = "Primary schedule save folder" if kind == "primary" else "Backup schedule save folder"
        try:
            selected = pick_directory_dialog(title=title)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {"path": selected}

    @app.post("/api/desktop/shutdown")
    def desktop_shutdown() -> dict[str, bool]:
        """Close the packaged desktop process when the UI browser tab exits."""
        is_desktop = os.environ.get("SCHEDULE_BUILDER_DESKTOP_RUNTIME") == "1"
        if is_desktop:
            threading.Timer(0.25, lambda: os._exit(0)).start()
        return {"desktop_runtime": is_desktop, "shutdown_requested": is_desktop}

    @app.get("/api/content-catalog")
    def content_catalog(config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        rows = canonical_rows_from_config(cfg)
        static_payload = _static_catalog_payload()
        if (
            static_payload is not None
            and not rows
            and static_payload.get("row_count", 0) > 0
            and Path(config) == DEFAULT_CONFIG
        ):
            return static_payload
        return {
            "schema_version": 1,
            "row_count": len(rows),
            "rows": rows,
        }

    @app.post("/api/content/import")
    def import_content_row(payload: ImportContentRowPayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import import build_manual_row, import_content_rows

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        try:
            row = build_manual_row(
                content_type=payload.content_type,
                show_name=payload.show_name,
                episode_number=payload.episode_number,
                episode_title=payload.episode_title,
                runtime_minutes=payload.runtime_minutes,
                slot_minutes=payload.slot_minutes,
                genre=payload.genre,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return import_content_rows(cfg, [row])

    @app.post("/api/content/import/batch")
    def import_content_batch(payload: ImportContentRowsPayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import import import_content_rows

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        if not payload.rows:
            raise HTTPException(status_code=400, detail="No content rows provided")
        return import_content_rows(cfg, payload.rows)

    @app.put("/api/content/show-rows")
    def update_show_rows(payload: UpdateShowRowsPayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import import replace_show_catalog_rows

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        if not payload.display_name.strip():
            raise HTTPException(status_code=400, detail="display_name is required")
        try:
            return replace_show_catalog_rows(cfg, payload.display_name.strip(), payload.rows)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/content/show/rename")
    def rename_show(payload: RenameShowPayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import import rename_show_catalog

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        try:
            return rename_show_catalog(cfg, payload.display_name, payload.new_display_name)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/content/show/delete")
    def delete_show(payload: ShowDisplayNamePayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import import delete_show_catalog

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        try:
            return delete_show_catalog(cfg, payload.display_name)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/content/import/upload")
    async def import_content_upload(
        file: UploadFile = File(...),
        config: str = str(DEFAULT_CONFIG),
    ) -> dict[str, Any]:
        from binge_schedule.content_import import import_content_rows, parse_upload_file

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        payload = await file.read()
        if not payload:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")
        try:
            rows = parse_upload_file(file.filename or "upload.csv", payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not parse upload: {exc}") from exc
        if not rows:
            raise HTTPException(status_code=400, detail="No content rows found in the uploaded file")
        return import_content_rows(cfg, rows)

    @app.post("/api/content/import/parse")
    async def import_content_parse(file: UploadFile = File(...)) -> dict[str, Any]:
        from binge_schedule.content_import_wizard import create_import_session, parse_session_response

        payload = await file.read()
        if not payload:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")
        try:
            session_id = create_import_session(file.filename or "upload.csv", payload)
            return parse_session_response(session_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not parse upload: {exc}") from exc

    @app.post("/api/content/import/sheet")
    def import_content_sheet(payload: ImportSheetAnalyzePayload) -> dict[str, Any]:
        from binge_schedule.content_import_wizard import analyze_sheet_in_session

        try:
            return analyze_sheet_in_session(payload.session_id, payload.sheet_name, payload.header_row)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not analyze sheet: {exc}") from exc

    @app.post("/api/content/import/sample-rows")
    def import_content_sample_rows(payload: ImportSampleRowsPayload) -> dict[str, Any]:
        from binge_schedule.content_import_wizard import sample_rows_for_config

        if not payload.session_id:
            raise HTTPException(status_code=400, detail="session_id is required")
        try:
            rows = sample_rows_for_config(payload.session_id, payload.sheet.model_dump())
            return {"sample_rows": rows}
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not build sample rows: {exc}") from exc

    @app.post("/api/content/import/preview")
    def import_content_preview(
        payload: ImportPreviewPayload,
        config: str = str(DEFAULT_CONFIG),
    ) -> dict[str, Any]:
        from binge_schedule.content_import_wizard import preview_import

        if not payload.session_id:
            raise HTTPException(status_code=400, detail="session_id is required")
        try:
            cfg_path = _safe_config_path(config)
            cfg = load_build_config(cfg_path)
            sheet_payloads = [sheet.model_dump() for sheet in payload.sheets]
            return preview_import(payload.session_id, sheet_payloads, cfg=cfg)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Preview failed: {exc}") from exc

    @app.post("/api/content/import/commit")
    def import_content_commit(payload: ImportCommitPayload, config: str = str(DEFAULT_CONFIG)) -> dict[str, Any]:
        from binge_schedule.content_import_wizard import commit_import

        cfg_path = _safe_config_path(config)
        cfg = load_build_config(cfg_path)
        if not payload.session_id:
            raise HTTPException(status_code=400, detail="session_id is required")
        try:
            sheet_payloads = [sheet.model_dump() for sheet in payload.sheets]
            return commit_import(cfg, payload.session_id, sheet_payloads)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Import failed: {exc}") from exc

    @app.get("/api/base-schedules")
    def base_schedules() -> dict[str, Any]:
        schedules = _builder_base_schedules()
        ready = sorted(
            [item for item in schedules if item["ready_to_generate"]],
            key=_base_schedule_sort_key,
            reverse=True,
        )
        return {
            "count": len(schedules),
            "ready_count": len(ready),
            "schedules": schedules,
            "active": ready[0] if ready else None,
        }

    def _base_schedule_detail_response(path: str) -> dict[str, Any]:
        if not path.strip():
            raise HTTPException(status_code=400, detail="path is required")
        base = _load_builder_base_schedule(path.strip())
        marker = base["marker"]
        raw = base["raw"]
        blocks = marker.get("draft_blocks") if isinstance(marker.get("draft_blocks"), list) else []
        weeks = raw.get("weeks") if isinstance(raw.get("weeks"), list) else []
        return {
            "path": base["path"].as_posix(),
            "label": _base_schedule_label(base["path"], str(marker.get("station_id") or "")),
            "station_id": str(marker.get("station_id") or ""),
            "week_monday": str(marker.get("week_monday") or ""),
            "week_count": int(marker.get("week_count") or len(weeks) or 1),
            "created_at": str(marker.get("created_at") or ""),
            "draft_block_count": len(blocks),
            "blocks": blocks,
        }

    @app.get("/api/base-schedules/detail")
    def base_schedule_detail(path: str = "") -> dict[str, Any]:
        return _base_schedule_detail_response(path)

    @app.post("/api/base-schedules/view")
    def base_schedule_view(payload: BaseSchedulePathPayload) -> dict[str, Any]:
        return _base_schedule_detail_response(payload.path)

    @app.post("/api/base-schedules/delete")
    def delete_base_schedule(payload: BaseSchedulePathPayload) -> dict[str, Any]:
        deleted_path = _delete_builder_base_schedule(payload.path)
        return {"deleted": True, "path": deleted_path}

    @app.post("/api/schedule/analyze-rules")
    def analyze_rules(payload: BlocksPayload) -> dict[str, Any]:
        rules = analyze_schedule_rules(payload.blocks, catalog_rows=payload.catalog_rows)
        return {
            "rule_count": len(rules),
            "rules": [rule.to_dict() for rule in rules],
        }

    @app.post("/api/schedule/blocks-to-grid")
    def blocks_to_grid(payload: BlocksToGridPayload) -> dict[str, Any]:
        try:
            grid = blocks_to_week_grid(
                payload.blocks,
                week_monday=payload.week_monday,
                require_complete=payload.require_complete,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        missing = empty_slots_for_blocks(payload.blocks, week_monday=payload.week_monday)
        return {
            "week_monday": payload.week_monday.isoformat(),
            "missing_slot_count": len(missing),
            "missing_slots": [{"slot": slot, "day_index": day_index} for slot, day_index in missing[:200]],
            "grid": grid,
        }

    @app.post("/api/schedule/grid-to-blocks")
    def grid_to_calendar_blocks(payload: GridToBlocksPayload) -> dict[str, Any]:
        try:
            blocks = grid_to_blocks(payload.grid, week_monday=payload.week_monday)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {
            "week_monday": payload.week_monday.isoformat(),
            "block_count": len(blocks),
            "blocks": [block.to_dict() for block in blocks],
        }

    @app.post("/api/schedule/download/{report_kind}")
    def download_schedule_report(report_kind: str, payload: DownloadSchedulePayload) -> StreamingResponse:
        week_count = _bounded_week_count(payload.week_count)
        station_label = _station_label(payload.station_id)
        if report_kind == "binge":
            data = _build_binge_preview_workbook(payload.blocks, station_id=payload.station_id)
            filename = f"{station_label}.xlsx"
        elif report_kind == "grids":
            data = _build_grids_preview_workbook(
                payload.blocks,
                station_id=payload.station_id,
                week_monday=payload.week_monday,
                week_count=week_count,
            )
            filename = f"{station_label} GRIDS.xlsx"
        else:
            raise HTTPException(status_code=404, detail="Unknown report kind")
        return StreamingResponse(
            BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.post("/api/schedule/export-to-downloads")
    def export_schedule_to_downloads(payload: DownloadSchedulePayload) -> dict[str, Any]:
        week_count = _bounded_week_count(payload.week_count)
        station_label = _station_label(payload.station_id)
        folder_name = _downloads_export_folder_name(payload.station_id, payload.week_monday)
        export_dir = _user_downloads_directory() / folder_name
        export_dir.mkdir(parents=True, exist_ok=True)
        binge_path = export_dir / f"{station_label}.xlsx"
        grids_path = export_dir / f"{station_label} GRIDS.xlsx"
        notes_path = export_dir / "Warnings and Notes.csv"
        binge_path.write_bytes(_build_binge_preview_workbook(payload.blocks, station_id=payload.station_id))
        grids_path.write_bytes(
            _build_grids_preview_workbook(
                payload.blocks,
                station_id=payload.station_id,
                week_monday=payload.week_monday,
                week_count=week_count,
            )
        )
        notes_path.write_text(_notes_csv(payload.notes, station_id=payload.station_id), encoding="utf-8")
        return {
            "saved": True,
            "directory": export_dir.as_posix(),
            "folder_name": folder_name,
            "files": [binge_path.as_posix(), grids_path.as_posix(), notes_path.as_posix()],
        }

    @app.post("/api/schedule/auto-generate")
    def auto_generate_schedule(payload: AutoGeneratePayload) -> dict[str, Any]:
        base = _load_builder_base_schedule(payload.base_path)
        marker = base["marker"]
        generate_week_count = _bounded_week_count(payload.week_count)
        blocks, week_monday, week_count = _auto_generate_blocks(
            base["path"],
            base["raw"],
            generate_week_count=generate_week_count,
        )
        return {
            "station_id": str(marker.get("station_id") or ""),
            "week_monday": week_monday.isoformat(),
            "week_count": week_count,
            "blocks": blocks,
        }

    @app.post("/api/base-schedules/save")
    def save_base_schedule(payload: SaveBaseSchedulePayload) -> dict[str, Any]:
        station_id = payload.station_id.strip()
        if not station_id:
            raise HTTPException(status_code=400, detail="Station ID is required")
        if not payload.blocks:
            raise HTTPException(status_code=400, detail="Schedule has no blocks")
        week_count = _bounded_week_count(payload.week_count)
        missing = []
        for week_index in range(week_count):
            week_monday = payload.week_monday + timedelta(days=week_index * 7)
            missing.extend(empty_slots_for_blocks(payload.blocks, week_monday=week_monday))
        if missing:
            raise HTTPException(status_code=400, detail=f"Schedule has {len(missing)} empty half-hour slots")
        path = _save_builder_base_schedule(
            station_id=station_id,
            week_monday=payload.week_monday,
            week_count=week_count,
            blocks=payload.blocks,
            suggested_rules=payload.suggested_rules,
            notes=payload.notes,
            saved_directory=payload.saved_directory,
        )
        return {
            "saved": True,
            "path": path.as_posix(),
            "label": _base_schedule_label(path, station_id),
            "station_id": station_id,
        }

    _register_splash_routes(app)
    ui_dist = _ui_dist_path()
    if ui_dist is not None:
        app.mount("/", StaticFiles(directory=ui_dist, html=True), name="scheduler-ui")

    return app


def _register_splash_routes(app: FastAPI) -> None:
    from fastapi.responses import FileResponse

    dist = _ui_dist_path()
    if dist is None:
        return
    splash_html = dist / "splash.html"
    splash_mp4 = dist / "splash.mp4"
    if splash_html.is_file():

        @app.get("/splash.html")
        def splash_page() -> FileResponse:
            return FileResponse(splash_html, media_type="text/html")

    if splash_mp4.is_file():

        @app.get("/splash.mp4")
        def splash_video() -> FileResponse:
            return FileResponse(splash_mp4, media_type="video/mp4")


def _static_catalog_payload() -> dict[str, Any] | None:
    path = _static_catalog_path()
    if path is None:
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return payload
    return None


def _static_catalog_path() -> Path | None:
    from binge_schedule.runtime_paths import catalog_publish_targets

    for candidate in catalog_publish_targets():
        if candidate.is_file():
            return candidate
    return None


def _ui_dist_path() -> Path | None:
    from binge_schedule.runtime_paths import react_dist_path

    return react_dist_path()


def _safe_config_path(raw: str) -> Path:
    from binge_schedule.runtime_paths import default_config_path, resolve_config_path

    try:
        p = resolve_config_path(raw or str(default_config_path()))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"Config not found: {raw}") from exc
    if p.suffix.lower() not in {".yaml", ".yml"}:
        raise HTTPException(status_code=400, detail="Config path must be a YAML file")
    return p


def _save_builder_base_schedule(
    *,
    station_id: str,
    week_monday: date,
    week_count: int,
    blocks: list[dict[str, Any]],
    suggested_rules: list[dict[str, Any]],
    notes: list[dict[str, Any]],
    saved_directory: str = "",
) -> Path:
    safe_station = _safe_station_id(station_id)
    created_at = datetime.now()
    overwrite_raw = str(saved_directory or "").strip()
    if overwrite_raw:
        save_dir = Path(overwrite_raw).expanduser().resolve()
        root = _saved_schedules_root().resolve()
        if save_dir != root and root not in save_dir.parents:
            raise HTTPException(
                status_code=400,
                detail="Save directory must be inside the configured schedules folder.",
            )
        save_dir.mkdir(parents=True, exist_ok=True)
    else:
        save_dir = _saved_schedule_dir(safe_station, created_at)
        save_dir.mkdir(parents=True, exist_ok=True)
    path = save_dir / "base_schedule.yaml"
    source = _base_schedule_source_config()
    base = {
        "gracenote_binge_id": int(source.get("gracenote_binge_id", 0) or 0),
        "nikki_workbook": source.get("nikki_workbook") or "../data/2024 Nikki Spreadsheets.xlsx",
        "timezone_note": source.get("timezone_note") or "local",
        "wrap_episodes": bool(source.get("wrap_episodes", True)),
        "cursor_state_file": f"episode_cursors_{safe_station}.json",
        "schedule_builder": {
            "managed": True,
            "kind": "base_schedule",
            "source": "react_schedule_builder",
            "station_id": station_id,
            "week_monday": week_monday.isoformat(),
            "week_count": week_count,
            "created_at": created_at.isoformat(timespec="seconds"),
            "saved_directory": save_dir.as_posix(),
            "draft_block_count": len(blocks),
            "draft_blocks": blocks,
            "suggested_rules": suggested_rules,
            "notes": notes,
        },
        "shows": source.get("shows") if isinstance(source.get("shows"), dict) else {},
        "weeks": [
            {
                "monday": (week_monday + timedelta(days=week_index * 7)).isoformat(),
                "grids_file": f"../data/base_schedules/{safe_station}/base_schedule_grids.xlsx",
                "sheet_name": _sheet_name_for_week(week_monday + timedelta(days=week_index * 7)),
            }
            for week_index in range(week_count)
        ],
    }
    path.write_text(yaml.safe_dump(base, sort_keys=False, allow_unicode=False), encoding="utf-8")
    station_label = _station_label(station_id)
    (save_dir / f"{station_label}.xlsx").write_bytes(_build_binge_preview_workbook(blocks, station_id=station_id))
    (save_dir / f"{station_label} GRIDS.xlsx").write_bytes(
        _build_grids_preview_workbook(blocks, station_id=station_id, week_monday=week_monday, week_count=week_count)
    )
    (save_dir / "Warnings and Notes.csv").write_text(_notes_csv(notes, station_id=station_id), encoding="utf-8")
    _save_auto_cursors(save_dir / f"episode_cursors_{safe_station}.json", _seed_cursors_from_template(blocks, _catalog_episodes_by_show()))
    from binge_schedule.app_settings import mirror_saved_schedule_dir

    mirror_saved_schedule_dir(save_dir)
    return path


def _build_binge_preview_workbook(blocks: list[dict[str, Any]], *, station_id: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _station_label(station_id)[:31]
    headers = ["Station ID", "Date", "Start", "End", "Show", "Episode", "Slot", "Runtime", "Avails"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    for block in sorted(blocks, key=lambda item: str(item.get("start") or "")):
        start = _parse_iso_datetime(block.get("start"))
        end = _parse_iso_datetime(block.get("end"))
        slot_minutes = _minutes_between(start, end)
        runtime = _float_or_none(block.get("runtimeMinutes") or block.get("runtime_minutes"))
        episode = _episode_label(block)
        ws.append(
            [
                station_id,
                start.date().isoformat() if start else "",
                _clock_label(start),
                _clock_label(end),
                str(block.get("show") or ""),
                episode,
                f"{slot_minutes} min" if slot_minutes is not None else "",
                _duration_label(runtime),
                _duration_label(max(0, slot_minutes - runtime)) if slot_minutes is not None and runtime is not None else "",
            ]
        )
    _autosize_columns(ws)
    return _workbook_bytes(wb)


def _build_grids_preview_workbook(
    blocks: list[dict[str, Any]],
    *,
    station_id: str = "",
    week_monday: date,
    week_count: int,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    station_label = _station_label(station_id)
    for week_index in range(week_count):
        monday = week_monday + timedelta(days=week_index * 7)
        grid = blocks_to_week_grid(blocks, week_monday=monday, require_complete=False)
        ws = wb.create_sheet(_sheet_name_for_week(monday))
        ws.append(["", "", "", "", station_label, "", "", "", ""])
        ws.append(["", "", "", "", f"Week of {monday.isoformat()}", "", "", "", ""])
        ws.append(["", *[(monday + timedelta(days=i)).isoformat() for i in range(7)], ""])
        ws.append([station_label, "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday", station_label])
        for slot in range(48):
            time_label = _slot_label(slot)
            ws.append([time_label, *[_grid_show_only(grid[slot][day]) for day in range(7)], time_label])
        _merge_grids_blocks(ws, blocks, monday)
        _apply_grid_borders(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1] + ws[2] + ws[3] + ws[4]:
            cell.font = Font(bold=True)
        _autosize_columns(ws, max_width=26)
    return _workbook_bytes(wb)


def _user_downloads_directory() -> Path:
    downloads = Path.home() / "Downloads"
    downloads.mkdir(parents=True, exist_ok=True)
    return downloads


def _downloads_export_folder_name(station_id: str, week_monday: date) -> str:
    station_label = _station_label(station_id)
    date_label = f"{week_monday.month}-{week_monday.day}-{week_monday.year}"
    raw = f"{station_label} {date_label}"
    cleaned = "".join(ch for ch in raw if ch not in '<>:"/\\|?*')
    return cleaned.strip() or "Schedule Export"


def _notes_csv(notes: list[dict[str, Any]], *, station_id: str = "") -> str:
    rows = [["Station ID", "Type", "Show", "Message"]]
    for note in notes:
        rows.append(
            [
                station_id,
                str(note.get("kind") or ""),
                str(note.get("show") or ""),
                str(note.get("message") or ""),
            ]
        )
    return "\n".join(",".join(_csv_escape(cell) for cell in row) for row in rows) + "\n"


def _csv_escape(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _merge_grids_blocks(ws, blocks: list[dict[str, Any]], week_monday: date) -> None:
    """Match the GRIDS report look: long programs are one merged vertical cell."""
    for block in blocks:
        start = _parse_iso_datetime(block.get("start"))
        end = _parse_iso_datetime(block.get("end"))
        if start is None or end is None:
            continue
        day_index = (start.date() - week_monday).days
        if day_index < 0 or day_index > 6:
            continue
        start_slot = _datetime_to_slot_start(start)
        end_slot = _datetime_to_end_slot(start, end)
        if end_slot <= start_slot:
            continue
        row = 5 + start_slot
        col = 2 + day_index
        ws.cell(row=row, column=col).value = str(block.get("show") or ws.cell(row=row, column=col).value or "")
        if end_slot - start_slot > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=5 + end_slot - 1, end_column=col)


def _apply_grid_borders(ws) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=4, max_row=52, min_col=1, max_col=9):
        for cell in row:
            cell.border = border


def _parse_iso_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _minutes_between(start: datetime | None, end: datetime | None) -> int | None:
    if start is None or end is None:
        return None
    return max(0, int(round((end - start).total_seconds() / 60)))


def _datetime_to_slot_start(value: datetime) -> int:
    minutes = value.hour * 60 + value.minute
    return max(0, min(47, minutes // 30))


def _datetime_to_end_slot(start: datetime, end: datetime) -> int:
    if end.date() > start.date():
        return 48
    minutes = end.hour * 60 + end.minute
    return max(0, min(48, (minutes + 29) // 30))


def _clock_label(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%I:%M %p").lstrip("0")


def _slot_label(slot: int) -> str:
    total = slot * 30
    hour = total // 60
    minute = total % 60
    return datetime(2026, 1, 1, hour, minute).strftime("%I:%M %p").lstrip("0")


def _float_or_none(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _duration_label(minutes: float | None) -> str:
    if minutes is None:
        return ""
    total_seconds = int(round(minutes * 60))
    mins = total_seconds // 60
    seconds = total_seconds % 60
    return f"{mins}:{seconds:02d}" if seconds else f"{mins} min"


def _episode_label(block: dict[str, Any]) -> str:
    code = str(block.get("episodeCode") or block.get("episode_code") or "").strip()
    title = str(block.get("episodeTitle") or block.get("episode_title") or "").strip()
    return " - ".join(part for part in (code, title) if part)


def _grid_show_only(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    first = text.splitlines()[0].strip()
    if " - (" in first:
        first = first.split(" - (", 1)[0].strip()
    return first


def _autosize_columns(ws, *, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        longest = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max_width, max(10, longest + 2))


def _workbook_bytes(wb: Workbook) -> bytes:
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _bounded_week_count(raw: int) -> int:
    try:
        val = int(raw)
    except Exception:
        return 1
    return max(1, min(4, val))


def _sheet_name_for_week(week_monday: date) -> str:
    return f"{week_monday.month}-{week_monday.day}-{week_monday.year}"


def _base_schedule_source_config() -> dict[str, Any]:
    from binge_schedule.runtime_paths import default_config_path, resolve_config_path

    for raw in ("config/blank_schedule.yaml", str(default_config_path())):
        try:
            path = resolve_config_path(raw)
            loaded = yaml.safe_load(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if isinstance(loaded, dict):
            return loaded
    return {}


def _normalize_schedule_path_key(path: Path) -> str:
    try:
        return path.resolve().as_posix().casefold()
    except OSError:
        return path.as_posix().casefold()


def _delete_builder_base_schedule(raw_path: str) -> str:
    base = _load_builder_base_schedule(raw_path)
    schedule_file = base["path"].resolve()
    save_dir = schedule_file.parent
    root = _saved_schedules_root().resolve()
    if save_dir != root and root not in save_dir.parents:
        raise HTTPException(
            status_code=400,
            detail="Only schedules inside the configured save folder can be deleted.",
        )
    if not save_dir.is_dir():
        raise HTTPException(status_code=404, detail="Saved schedule folder not found")
    shutil.rmtree(save_dir)
    return save_dir.as_posix()


def _load_builder_base_schedule(raw_path: str = "") -> dict[str, Any]:
    schedules = _builder_base_schedules()
    selected: Path | None = None
    if raw_path.strip():
        requested = Path(raw_path.strip())
        requested_key = _normalize_schedule_path_key(requested)
        for item in schedules:
            path = Path(str(item.get("path") or ""))
            if _normalize_schedule_path_key(path) == requested_key:
                selected = path
                break
            if path.name == "base_schedule.yaml" and str(path.as_posix()).casefold().endswith(
                requested.as_posix().casefold()
            ):
                selected = path
                break
        if selected is None and requested.is_file() and requested.name == "base_schedule.yaml":
            selected = requested
        if selected is None:
            cwd_candidate = (Path.cwd() / requested).resolve()
            if cwd_candidate.is_file():
                selected = cwd_candidate
    elif schedules:
        ready = [item for item in schedules if item.get("ready_to_generate")]
        if ready:
            selected = Path(str(ready[0]["path"]))
    if selected is None or not selected.is_file():
        raise HTTPException(status_code=404, detail="No saved base schedule found")
    try:
        raw = yaml.safe_load(selected.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read base schedule: {selected}") from exc
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="Saved base schedule is not valid")
    marker = raw.get("schedule_builder")
    if not isinstance(marker, dict) or marker.get("managed") is not True:
        raise HTTPException(status_code=400, detail="Saved schedule is not builder-managed")
    return {"path": selected, "raw": raw, "marker": marker}


def _parse_week_monday(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Saved base schedule has an invalid week_monday") from exc


def _remint_block_id(block: dict[str, Any], start: datetime) -> str:
    raw_id = str(block.get("id") or "block")
    stem = raw_id.rsplit("-", 1)[0] if "-" in raw_id else raw_id
    return f"{stem}-{int(start.timestamp() * 1000)}"


def _auto_generate_blocks(
    base_path: Path,
    raw: dict[str, Any],
    *,
    generate_week_count: int,
) -> tuple[list[dict[str, Any]], date, int]:
    marker = raw.get("schedule_builder") if isinstance(raw.get("schedule_builder"), dict) else {}
    template_blocks = marker.get("draft_blocks") if isinstance(marker.get("draft_blocks"), list) else []
    if not template_blocks:
        raise HTTPException(status_code=400, detail="Saved base schedule has no template blocks")
    template_week_count = _bounded_week_count(int(marker.get("week_count") or 1))
    generate_week_count = _bounded_week_count(generate_week_count)
    base_monday = _parse_week_monday(marker.get("week_monday"))
    next_monday = base_monday + timedelta(days=template_week_count * 7)

    station_id = str(marker.get("station_id") or "")
    safe_station = _safe_station_id(station_id)
    cursor_path = base_path.parent / f"episode_cursors_{safe_station}.json"
    episodes_by_show = _catalog_episodes_by_show()
    cursors = _load_auto_cursors(cursor_path)
    seeded = _seed_cursors_from_template(template_blocks, episodes_by_show)
    if not cursors:
        cursors = seeded
    else:
        for show, index in seeded.items():
            cursors[show] = max(int(cursors.get(show, 0) or 0), int(index or 0))

    generated: list[dict[str, Any]] = []
    for gen_week_index in range(generate_week_count):
        template_week_index = gen_week_index % template_week_count
        target_week_monday = next_monday + timedelta(days=gen_week_index * 7)
        template_week_start = base_monday + timedelta(days=template_week_index * 7)

        for block in sorted(template_blocks, key=lambda item: str(item.get("start") or "")):
            start = _parse_iso_datetime(block.get("start"))
            end = _parse_iso_datetime(block.get("end"))
            if start is None:
                continue
            block_template_week = (start.date() - base_monday).days // 7
            if block_template_week != template_week_index:
                continue

            day_offset = start.date() - template_week_start
            new_start = datetime.combine(target_week_monday + day_offset, start.time())
            if end is not None:
                new_end = new_start + (end - start)
            else:
                new_end = new_start + timedelta(minutes=30)

            next_block = dict(block)
            next_block["start"] = new_start.isoformat(timespec="seconds")
            next_block["end"] = new_end.isoformat(timespec="seconds")
            next_block["id"] = _remint_block_id(block, new_start)

            show = str(next_block.get("show") or "").strip()
            if _auto_generates_episode(next_block) and show in episodes_by_show:
                episodes = episodes_by_show[show]
                if episodes:
                    index = int(cursors.get(show, 0) or 0) % len(episodes)
                    episode = episodes[index]
                    cursors[show] = (index + 1) % len(episodes)
                    next_block.update(
                        {
                            "episodeId": episode["id"],
                            "episode_id": episode["id"],
                            "episodeCode": episode["code"],
                            "episodeTitle": episode["title"],
                            "title": " ".join(part for part in (episode["code"], episode["title"]) if part).strip(),
                            "runtimeMinutes": episode["runtime_minutes"],
                            "content_type": episode["content_type"],
                            "contentType": episode["content_type"],
                        }
                    )
            generated.append(next_block)

    generated.sort(key=lambda item: str(item.get("start") or ""))
    _save_auto_cursors(cursor_path, cursors)
    return generated, next_monday, generate_week_count


def _catalog_episodes_by_show() -> dict[str, list[dict[str, Any]]]:
    from binge_schedule.content_catalog import canonical_rows_from_config
    from binge_schedule.runtime_paths import default_config_path

    cfg = load_build_config(default_config_path())
    rows = canonical_rows_from_config(cfg)
    if not rows:
        static_payload = _static_catalog_payload()
        if static_payload is not None and isinstance(static_payload.get("rows"), list):
            rows = static_payload["rows"]
    by_show: dict[str, list[dict[str, Any]]] = {}
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        if row.get("playable") is False:
            continue
        if row.get("availability_status") and row.get("availability_status") not in {"available", "metadata_only"}:
            continue
        content_type = _frontend_content_type(str(row.get("content_type") or ""))
        if content_type != "Series / show":
            continue
        show = str(row.get("display_name") or "").strip()
        if not show:
            continue
        scheduled = _float_or_none(row.get("binge_row_minutes") or row.get("runtime_minutes")) or 30
        runtime = _float_or_none(row.get("runtime_minutes")) or scheduled
        episode = {
            "id": str(row.get("episode_key") or f"{row.get('series_key') or show}-{index}"),
            "show": show,
            "code": str(row.get("episode_code") or row.get("episode_number") or "EP").strip(),
            "title": str(row.get("episode_title") or show).strip(),
            "runtime_minutes": runtime,
            "content_type": content_type,
        }
        by_show.setdefault(show, []).append(episode)
    return by_show


def _frontend_content_type(value: str) -> str:
    normalized = value.lower().replace(" ", "_")
    if normalized in {"movie", "movies", "special", "specials", "film", "feature"}:
        return "Movie / special"
    if normalized in {"paid", "paid_programming", "infomercial", "ministry"}:
        return "Paid programming"
    return "Series / show"


def _auto_generates_episode(block: dict[str, Any]) -> bool:
    content_type = str(block.get("contentType") or block.get("content_type") or "").casefold()
    code = str(block.get("episodeCode") or block.get("episode_code") or "").upper()
    if "paid" in content_type or code in {"PAID", "LIT"}:
        return False
    if "movie" in content_type:
        return False
    return True


def _load_auto_cursors(path: Path) -> dict[str, int]:
    if not path.is_file():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    raw = payload.get("cursors") if isinstance(payload, dict) else {}
    if not isinstance(raw, dict):
        return {}
    out: dict[str, int] = {}
    for show, value in raw.items():
        try:
            out[str(show)] = max(0, int(value))
        except (TypeError, ValueError):
            continue
    return out


def _save_auto_cursors(path: Path, cursors: dict[str, int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"cursors": cursors}, indent=2), encoding="utf-8")


def _seed_cursors_from_template(
    blocks: list[dict[str, Any]],
    episodes_by_show: dict[str, list[dict[str, Any]]],
) -> dict[str, int]:
    cursors: dict[str, int] = {}
    for block in blocks:
        if not _auto_generates_episode(block):
            continue
        show = str(block.get("show") or "").strip()
        episodes = episodes_by_show.get(show) or []
        if not episodes:
            continue
        pos = _episode_position(block, episodes)
        if pos is not None:
            cursors[show] = max(cursors.get(show, 0), pos + 1)
    for show, episodes in episodes_by_show.items():
        if show in cursors and episodes:
            cursors[show] = cursors[show] % len(episodes)
    return cursors


def _episode_position(block: dict[str, Any], episodes: list[dict[str, Any]]) -> int | None:
    tokens = {
        str(block.get("episodeId") or block.get("episode_id") or "").strip().casefold(),
        str(block.get("episodeCode") or block.get("episode_code") or "").strip().casefold(),
        str(block.get("episodeTitle") or block.get("episode_title") or "").strip().casefold(),
    }
    tokens.discard("")
    for index, episode in enumerate(episodes):
        if {
            str(episode.get("id") or "").strip().casefold(),
            str(episode.get("code") or "").strip().casefold(),
            str(episode.get("title") or "").strip().casefold(),
        } & tokens:
            return index
    return None


def _saved_schedules_root() -> Path:
    from binge_schedule.runtime_paths import saved_schedules_root

    return saved_schedules_root()


def _saved_schedule_dir(safe_station: str, created_at: datetime) -> Path:
    return _saved_schedules_root() / safe_station / created_at.strftime("%Y-%m-%d_%H-%M-%S")


def _safe_station_id(station_id: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in station_id.strip())
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return cleaned or "station"


def _station_label(station_id: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {" ", "-", "_"} else "_" for ch in station_id.strip()).strip() or "Station"


def _base_schedule_sort_key(item: dict[str, Any]) -> str:
    created = str(item.get("created_at") or "").strip()
    if created:
        return created
    return str(item.get("path") or "")


def _base_schedule_label(path: Path, station_id: str | None = None) -> str:
    sid = (station_id or "").strip()
    if sid:
        return f"Station {sid}"
    return path.stem.replace("_", " ").title()


def _builder_base_schedules() -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    candidates: list[Path] = []
    config_dir = Path("config")
    if config_dir.is_dir():
        candidates.extend(sorted(config_dir.glob("*.yaml")))
    saved_root = _saved_schedules_root()
    if saved_root.is_dir():
        candidates.extend(sorted(saved_root.glob("*/base_schedule.yaml")))
        candidates.extend(sorted(saved_root.glob("*/*/base_schedule.yaml")))
    for path in candidates:
        try:
            raw = yaml.safe_load(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(raw, dict):
            continue
        marker = raw.get("schedule_builder")
        if not isinstance(marker, dict) or marker.get("managed") is not True:
            continue
        station_id = str(marker.get("station_id") or "").strip()
        weeks = raw.get("weeks") if isinstance(raw.get("weeks"), list) else []
        shows = raw.get("shows") if isinstance(raw.get("shows"), dict) else {}
        out.append(
            {
                "path": path.as_posix(),
                "label": _base_schedule_label(path, station_id),
                "station_id": station_id,
                "kind": marker.get("kind") or "base_schedule",
                "source": marker.get("source") or "",
                "week_count": int(marker.get("week_count") or len(weeks) or 1),
                "template_week_count": int(marker.get("week_count") or len(weeks) or 1),
                "week_monday": str(marker.get("week_monday") or ""),
                "created_at": str(marker.get("created_at") or ""),
                "show_count": len(shows),
                "draft_block_count": int(marker.get("draft_block_count") or 0),
                "ready_to_generate": int(marker.get("draft_block_count") or 0) > 0,
            }
        )
    return out


app = create_app()
