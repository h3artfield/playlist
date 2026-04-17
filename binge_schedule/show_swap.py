"""Apply playlist show swaps to grids workbooks and setup YAML (and cursors when adding a new show)."""

from __future__ import annotations

import json
import re
from datetime import date, time, timedelta
from pathlib import Path
from typing import Any, Optional

import yaml
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from binge_schedule.binge_to_grid import (
    parse_binge_date_cell,
    parse_binge_time_cell,
    wall_time_to_slot_start,
)
from binge_schedule.config_io import load_build_config
from binge_schedule.cursor_state import resolved_cursor_state_path
from binge_schedule.grid import load_grid_sheet, parse_monday, segments_for_day
from binge_schedule.models import BuildConfig, ShowDef, WeekDef
from binge_schedule.workbook_discover import parse_workbook_tab_option, synthetic_series_for_tab


def _is_noop_swap(old_labels: list[str], new_display: str) -> bool:
    """True when every old label matches the replacement display name (after strip, case-insensitive)."""
    nd = new_display.strip().casefold()
    if not old_labels or not nd:
        return False
    for o in old_labels:
        if o.strip().casefold() != nd:
            return False
    return True


def replace_cell_show_text(text: str, old_labels: list[str], new_display: str) -> str:
    """Match grid cell text to ``resolve_show``-style rules: exact name, then longest-prefix among ``old_labels``."""
    olds = sorted({str(o).strip() for o in old_labels if o and str(o).strip()}, key=len, reverse=True)
    if not olds:
        return text
    s = str(text)
    for old in olds:
        if s == old:
            return new_display
    for old in olds:
        if s.startswith(old):
            return new_display + s[len(old) :]
    return text


def _unique_show_key(base: str, existing: set[str]) -> str:
    k = base
    n = 0
    while k in existing:
        n += 1
        k = f"{base}_{n}"
    return k


def _showdef_to_yaml_dict(sd: ShowDef) -> dict[str, Any]:
    d: dict[str, Any] = {
        "display_name": sd.display_name,
        "kind": sd.kind,
        "nikki_sheet": sd.nikki_sheet,
        "prefix": sd.prefix,
        "start_episode_index": sd.start_episode_index,
    }
    if sd.nikki_style:
        d["nikki_style"] = sd.nikki_style
    if sd.nikki_row_filter:
        d["nikki_row_filter"] = sd.nikki_row_filter
    if sd.overnight_repeat_after:
        d["overnight_repeat_after"] = sd.overnight_repeat_after
    if sd.binge_row_minutes != 30:
        d["binge_row_minutes"] = sd.binge_row_minutes
    return d


def _week_def_for_date(cfg: BuildConfig, d: date) -> Optional[WeekDef]:
    for w in cfg.weeks:
        m = parse_monday(w.monday)
        if m <= d < m + timedelta(days=7):
            return w
    return None


def parse_schedule_anchor(raw: Any) -> Optional[tuple[date, time]]:
    if raw is None or not isinstance(raw, dict):
        return None
    dv = raw.get("date")
    sv = raw.get("start")
    if dv is None or sv is None:
        return None
    try:
        d = parse_binge_date_cell(dv)
        t = parse_binge_time_cell(sv)
        return d, t
    except (ValueError, TypeError):
        return None


def _write_grid_program_cell(ws, excel_row: int, excel_col: int, new_val: str) -> bool:
    cell = ws.cell(row=excel_row, column=excel_col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= excel_row <= rng.max_row and rng.min_col <= excel_col <= rng.max_col:
                ws.cell(rng.min_row, rng.min_col).value = new_val
                return True
        return False
    cell.value = new_val
    return True


def _rewrite_grids_target_segment(
    gpath: Path,
    sheet_name: str,
    day_index: int,
    slot: int,
    old_labels: list[str],
    new_display: str,
) -> tuple[int, list[str]]:
    """Replace only the program block that contains ``slot`` on ``day_index`` (one BINGE row / airing)."""
    warnings: list[str] = []
    if not gpath.is_file():
        return 0, [f"Grids workbook missing: {gpath}"]
    try:
        grid = load_grid_sheet(str(gpath), sheet_name)
    except Exception as e:
        return 0, [f"Could not load grid `{sheet_name}` in `{gpath.name}`: {e}"]
    if not (0 <= day_index <= 6):
        return 0, [f"Invalid day index {day_index}."]
    if not (0 <= slot < 48):
        return 0, [f"Invalid time slot {slot}."]

    col = [grid[r][day_index] for r in range(48)]
    try:
        segs = segments_for_day(col)
    except ValueError as e:
        return 0, [str(e)]

    target = None
    for seg in segs:
        if seg.start_slot <= slot < seg.end_slot:
            target = seg
            break
    if target is None:
        return 0, [
            "No program block contains that date/time in the grids (check **DATE** / **START TIME** on the BINGE row)."
        ]

    wb = None
    changed = 0
    excel_col = 2 + day_index
    try:
        wb = load_workbook(gpath, read_only=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            return 0, [f"Sheet `{sheet_name}` not in `{gpath.name}`."]
        ws = wb[sheet_name]
        for r in range(target.start_slot, target.end_slot):
            raw = col[r]
            if raw is None or not str(raw).strip():
                continue
            s = str(raw).strip()
            new_s = replace_cell_show_text(s, old_labels, new_display)
            if new_s != s:
                if _write_grid_program_cell(ws, 5 + r, excel_col, new_s):
                    changed += 1
        if changed == 0:
            new_s = replace_cell_show_text(target.cell_text, old_labels, new_display)
            if new_s != target.cell_text.strip():
                if _write_grid_program_cell(ws, 5 + target.start_slot, excel_col, new_s):
                    changed = 1
        wb.save(gpath)
    except OSError as e:
        return changed, [f"Could not open/save {gpath}: {e}"]
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return changed, warnings


def _rewrite_grids_file(path: Path, sheet_names: set[str], old_labels: list[str], new_display: str) -> tuple[int, list[str]]:
    """Update program cells (rows 5–52, cols B–H). Returns (cells_changed, warnings)."""
    warnings: list[str] = []
    changed = 0
    if not path.is_file():
        warnings.append(f"Grids workbook missing: {path}")
        return 0, warnings
    wb = None
    try:
        wb = load_workbook(path, read_only=False, data_only=False)
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                warnings.append(f"Sheet `{sn}` not in {path.name} (skipped).")
                continue
            ws = wb[sn]
            for r in range(5, 53):
                for c in range(2, 9):
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    if v is None:
                        continue
                    s = str(v)
                    new_s = replace_cell_show_text(s, old_labels, new_display)
                    if new_s != s:
                        cell.value = new_s
                        changed += 1
        wb.save(path)
    except OSError as e:
        warnings.append(f"Could not open/save {path}: {e}")
        return changed, warnings
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return changed, warnings


def _ensure_cursor_entry(cfg: BuildConfig, show_key: str) -> list[str]:
    out: list[str] = []
    p = resolved_cursor_state_path(cfg)
    if p is None:
        return out
    data: dict[str, Any] = {}
    if p.is_file():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            data = {}
    cursors = data.get("cursors") or {}
    if not isinstance(cursors, dict):
        cursors = {}
    if show_key not in cursors:
        cursors[show_key] = 0
        data["cursors"] = cursors
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(json.dumps(data, indent=2), encoding="utf-8")
            out.append(f"Initialized cursor for `{show_key}` in `{p.name}`.")
        except OSError as e:
            out.append(f"Could not write cursors ({e}).")
    return out


def apply_show_swap(
    cfg_path: Path,
    old_show_labels: list[str],
    archive_pick: str,
    *,
    schedule_anchor: Any = None,
) -> tuple[bool, list[str]]:
    """
    Persist a show swap.

    With **schedule_anchor** ``{\"date\", \"start\"}`` from the BINGE row, only the grid **block for that
    date/time** is updated (the correct April vs May workbook is chosen by date).

    Without an anchor, every matching cell in all weeks is updated (legacy bulk behavior).

    If **archive_pick** is a workbook tab not yet on the playlist, append a **shows:** entry to the YAML
    and seed the cursor file for that key.

    Does not edit the Nikki ``.xlsx`` binary; new series use existing tabs via ``nikki_sheet`` in YAML.

    Returns ``(ok, messages)``. ``ok`` is True if grids were updated, YAML was added, or the swap was a
    **no-op** (replacement **display_name** already matches the old label(s)). ``ok`` is False on I/O errors
    or when grids exist but no cells matched and the swap was not a no-op.
    """
    messages: list[str] = []
    olds = [str(x).strip() for x in old_show_labels if x and str(x).strip()]
    if not olds:
        return False, ["No show labels to replace."]

    cfg_path = cfg_path.resolve()
    if not cfg_path.is_file():
        return False, [f"Setup file not found: {cfg_path}"]

    cfg = load_build_config(cfg_path)
    tab = parse_workbook_tab_option(archive_pick)

    if tab is not None:
        sd = synthetic_series_for_tab(tab)
        slug = re.sub(r"[^a-zA-Z0-9]+", "_", tab).strip("_").lower()[:50] or "sheet"
        new_key = _unique_show_key(f"tab_{slug}", set(cfg.shows.keys()))
        new_display = sd.display_name.strip()
        added_new_show = True
    else:
        if archive_pick not in cfg.shows:
            return False, [f"Unknown show key `{archive_pick}`."]
        new_key = archive_pick
        new_display = cfg.shows[archive_pick].display_name.strip()
        added_new_show = False

    if not added_new_show and _is_noop_swap(olds, new_display):
        return True, [
            "You chose the **same** show as the replacement — **no grid change** (that slot already has that program title).",
            "Pick a **different** show in the archive if you meant to replace the row.",
            f"**{new_display}**",
        ]

    if added_new_show:
        try:
            raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
        except (OSError, yaml.YAMLError) as e:
            return False, [f"Could not read YAML: {e}"]
        if not isinstance(raw, dict):
            return False, ["Invalid YAML root."]
        shows_block = raw.get("shows")
        if not isinstance(shows_block, dict):
            shows_block = {}
            raw["shows"] = shows_block
        if new_key in shows_block:
            messages.append(f"Show key `{new_key}` already in YAML; updating grids only.")
        else:
            shows_block[new_key] = _showdef_to_yaml_dict(
                ShowDef(
                    key=new_key,
                    display_name=sd.display_name,
                    kind=sd.kind,
                    nikki_sheet=sd.nikki_sheet,
                    prefix=sd.prefix,
                    start_episode_index=sd.start_episode_index,
                )
            )
            messages.append(f"Added show `{new_key}` to `{cfg_path.name}` (Nikki tab `{sd.nikki_sheet}`).")
        yaml_text = yaml.dump(
            raw,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
            width=120,
        )
        try:
            cfg_path.write_text(yaml_text, encoding="utf-8")
        except OSError as e:
            return False, [f"Could not write YAML: {e}"]
        messages.append(
            "YAML was saved (formatting/comments may change — use **git diff** if you rely on top-of-file notes)."
        )
        cfg = load_build_config(cfg_path)
        messages.extend(_ensure_cursor_entry(cfg, new_key))

    by_file: dict[Path, set[str]] = {}
    for w in cfg.weeks:
        p = Path(w.grids_file).resolve()
        by_file.setdefault(p, set()).add(w.sheet_name)

    anchor_t = parse_schedule_anchor(schedule_anchor)
    total_cells = 0

    if anchor_t is not None:
        d, tm = anchor_t
        wd = _week_def_for_date(cfg, d)
        if wd is None:
            return False, [f"No **weeks:** entry covers calendar date **{d.isoformat()}**."]
        monday = parse_monday(wd.monday)
        day_index = (d - monday).days
        if not (0 <= day_index <= 6):
            return False, [f"Date **{d}** does not fall in the ISO week starting **{wd.monday}**."]
        slot = wall_time_to_slot_start(tm)
        gpath = Path(wd.grids_file).resolve()
        n, warns = _rewrite_grids_target_segment(gpath, wd.sheet_name, day_index, slot, olds, new_display)
        messages.extend(warns)
        total_cells = n
        if n:
            messages.append(
                f"Updated **{n}** cell(s) in `{gpath.name}` · `{wd.sheet_name}` — **only** the airing "
                f"**{d.isoformat()}** **{tm.strftime('%H:%M')}** (other dates/weeks left as-is)."
            )
    else:
        messages.append(
            "**No DATE/START TIME anchor** — updated **every** matching program cell in **all** configured grid weeks."
        )
        for gpath, sheets in by_file.items():
            n, warns = _rewrite_grids_file(gpath, sheets, olds, new_display)
            total_cells += n
            messages.extend(warns)
            if n:
                messages.append(f"Updated **{n}** grid cell(s) in `{gpath.name}`.")

    if not by_file:
        messages.append("No **weeks:** in config — grids were not changed.")

    if total_cells == 0 and by_file:
        messages.append(
            "No grid cells matched the old label(s) in the target block. Check that **SHOW** matches **display_name** text in that grid cell."
        )

    messages.append(f"Replacement **display_name** used in grids: **{new_display}**.")

    success = added_new_show or total_cells > 0
    if not success:
        return False, messages
    return True, messages
