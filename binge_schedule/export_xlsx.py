from __future__ import annotations

import re
import shutil
from collections.abc import Sequence
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from binge_schedule.binge_overrides import BingeRowOverride, apply_binge_row_overrides
from binge_schedule.binge_to_grid import (
    binge_dataframe_to_grid,
    normalize_binge_df_columns,
    split_binge_df_by_monday,
)
from binge_schedule.binge_pattern import (
    load_reference_episode_actions,
    load_reference_week_dataframe,
    merge_literal_reference_binge_days,
    reconcile_catalog_from_binge_dataframe,
    sync_cursors_from_reference_binge_week,
)
from binge_schedule.build import (
    _short_program_title,
    build_catalog,
    build_grids_matrix,
    rows_for_week,
)
from binge_schedule.cursor_state import (
    apply_saved_cursors,
    resolved_cursor_state_path,
    save_cursors_after_export,
)
from binge_schedule.models import BingeRow, BuildConfig, WeekDef
from binge_schedule.grid import (
    ensure_grids_workbooks_for_weeks,
    load_grid_sheet,
    parse_monday,
    seed_grids_from_prior_month,
    segments_for_binge_scheduling,
    sync_straddle_weeks_to_canonical_grids_file,
)
from binge_schedule.overnight_repeat import apply_overnight_repeats_with_prev


def _resolved_save_binge_reference_copy_path(cfg: BuildConfig) -> Optional[Path]:
    raw = cfg.save_binge_reference_copy_to
    if not raw or not str(raw).strip():
        return None
    p = Path(str(raw).strip())
    if p.is_absolute():
        return p
    base = cfg.config_path.parent if cfg.config_path else Path.cwd()
    return (base / p).resolve()


def _sanitize_station_dir(label: str) -> str:
    s = "".join(ch if ch not in ':/\\*?"<>|\n\r\t' else "_" for ch in label.strip())
    s = s.strip("._") or "station"
    return s


def _export_station_labels(cfg: BuildConfig, override: Optional[Sequence[str]]) -> list[str]:
    if override is not None:
        return [x.strip() for x in override if str(x).strip()]
    if cfg.export_stations:
        return list(cfg.export_stations)
    return []


def is_verbose_seed_noise(s: str) -> bool:
    """True for legacy informational seed lines (per-show cursor sync, literal-copy stats) — omit from UI/CLI."""
    if "sync: cursor[" in s:
        return True
    if s.lstrip().lower().startswith("literal copy:"):
        return True
    if s.startswith("Archived BINGE reference copy"):
        return False
    if s.startswith("Station copy ["):
        return False
    return False


def _find_week_containing_date(all_weeks: Sequence[WeekDef], d: date) -> Optional[WeekDef]:
    for w in all_weeks:
        m = parse_monday(w.monday)
        if m <= d < m + timedelta(days=7):
            return w
    return None


def _cursor_warmup_week_if_needed(
    requested: Sequence[WeekDef], all_weeks: Sequence[WeekDef]
) -> Optional[WeekDef]:
    """Return the ISO week that contains the Sunday *before* the earliest requested Monday, if that week is not already in the request.

    Exporting **May** only (Mon May 4+) skips the Apr 27–May 3 sheet; without this, episode cursors never advance
    through Sun May 3 and Mon 0:00 starts from a stale Nikki index (wrong episode codes vs Sun night).
    """
    if not requested:
        return None
    first = min(requested, key=lambda w: parse_monday(w.monday))
    prev_sun = parse_monday(first.monday) - timedelta(days=1)
    warm = _find_week_containing_date(all_weeks, prev_sun)
    if warm is None:
        return None
    if warm.monday in {w.monday for w in requested}:
        return None
    return warm


# --- Legacy reference styling (matches APRIL 2026 BINGE / BINGE GRIDS examples) ---

_THIN_BLACK = Side(style="thin", color="FF000000")
_THIN_RED = Side(style="thin", color="FFFF0000")
_BORDER_GRID = Border(
    left=_THIN_BLACK, right=_THIN_BLACK, top=_THIN_BLACK, bottom=_THIN_BLACK
)
_BORDER_BINGE_D = Border(
    left=_THIN_RED, right=_THIN_BLACK, top=_THIN_BLACK, bottom=_THIN_BLACK
)
_BORDER_BINGE_E = Border(
    left=_THIN_RED, right=_THIN_BLACK, top=_THIN_BLACK, bottom=_THIN_BLACK
)

_FILL_DAY_A = PatternFill(fill_type="solid", fgColor="FFE2EFDA")  # light green band
_FILL_DAY_B = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
_FILL_HEADER_D = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
_FILL_GRID_HIGHLIGHT = PatternFill(fill_type="solid", fgColor="FFFFFF00")  # bright yellow

_BINGE_COL_WIDTHS: tuple[float, ...] = (
    15.28515625,
    12.140625,
    12.7109375,
    44.140625,
    55.28515625,
    44.140625,
    120.0,
)
# Grid body / weekday font (pt); smaller so uniform columns still fit long cells with wrap.
_GRID_FONT = "Calibri"
_GRID_PT_BANNER = 14
_GRID_PT_SUB = 8
_GRID_PT_DATE = 8
_GRID_PT_BODY = 7
# Program cells wrap; cap width basis so one long movie line does not force255-wide columns.
_GRID_BODY_MAX_CHARS_FOR_WIDTH = 48
_GRID_MAX_UNIFORM_WIDTH = 56.0
# Multiply computed Mon–Sun column width (before A/I = half of that).
_GRID_COL_WIDTH_SCALE = 0.5

# Bold through first "(year)" only when non-whitespace text follows (title + description).
_TITLE_THEN_YEAR_THEN_DESC = re.compile(
    r"^(?P<head>.+?\(\s*\d{4}\s*\))(?P<tail>.+)$",
    re.DOTALL,
)

_WEEKDAY_ABBR = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")


def _grids_program_rich_text(s: str, font_pt: int) -> str | CellRichText:
    """Plain show title → normal text. Title + (year) + description → bold through (year), rest normal."""
    if not s or not str(s).strip():
        return s
    m = _TITLE_THEN_YEAR_THEN_DESC.match(str(s))
    if not m:
        return s
    head, tail = m.group("head"), m.group("tail")
    if not tail.strip():
        return s
    ib = InlineFont(rFont=_GRID_FONT, sz=font_pt, b=True)
    ino = InlineFont(rFont=_GRID_FONT, sz=font_pt, b=False)
    return CellRichText(TextBlock(ib, head), TextBlock(ino, tail))


def _max_line_length_program_body(mat: list[list[Any]]) -> int:
    """Longest single line in 48×7 program block only (excludes times, banners, dates)."""
    longest = 6
    # mat[4:52] = grid body; col 1..7 = Mon..Sun
    for r in range(4, min(len(mat), 52)):
        row = mat[r]
        if len(row) < 8:
            continue
        for c in range(1, 8):
            v = row[c]
            if v is None:
                continue
            if isinstance(v, str):
                for line in v.splitlines():
                    longest = max(longest, len(line))
            elif not isinstance(v, (datetime, date, time)):
                longest = max(longest, len(str(v)))
    return longest


def _grids_uniform_width(max_line_chars: int, font_pt: int) -> float:
    capped = min(max_line_chars, _GRID_BODY_MAX_CHARS_FOR_WIDTH)
    factor = 1.0 + max(0, (11 - font_pt) * 0.035)
    w = capped / factor + 5.0
    return min(_GRID_MAX_UNIFORM_WIDTH, max(10.0, w))


def _grid_shape_ok(g: list[list[Optional[str]]]) -> bool:
    return len(g) == 48 and all(len(row) == 7 for row in g)


def _grid_cell_show_key(val: Optional[str]) -> str:
    """Stable identity for “which show” in a grid cell (first line / title before year), ignoring episode."""
    if val is None or not str(val).strip():
        return ""
    return _short_program_title(str(val)).strip().lower()


def _slot_show_equal(a: Optional[str], b: Optional[str]) -> bool:
    return _grid_cell_show_key(a) == _grid_cell_show_key(b)


def _grid_cell_plain_text(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, CellRichText):
        return "".join(str(p) for p in val)
    return str(val)


def _estimate_wrapped_lines(text: str, max_chars_per_line: int) -> int:
    """Word-aware line count for wrap_text (avoids mid-word truncation in layout)."""
    if not text or not str(text).strip():
        return 1
    m = max(8, max_chars_per_line)
    total = 0
    for para in str(text).split("\n"):
        para = para.strip()
        if not para:
            total += 1
            continue
        line_len = 0
        for word in para.split():
            wlen = len(word)
            add = wlen if line_len == 0 else 1 + wlen
            if line_len + add <= m:
                line_len += add
            else:
                total += 1
                line_len = wlen
        if line_len:
            total += 1
    return max(1, total)


def _row_height_for_wrapped_lines(n_lines: int, font_pt: float) -> float:
    return max(14.0, min(200.0, 4.0 + n_lines * font_pt * 1.22))


def _approx_chars_per_line(program_col_width: float, font_pt: float) -> int:
    return max(10, int(program_col_width * 6.2 / max(font_pt, 1.0)))


def _adjust_grids_body_row_heights(
    ws,
    program_col_width: float,
    edge_col_width: float,
    font_pt: float,
    min_row: int = 5,
    max_row: int = 52,
) -> None:
    cpl_prog = _approx_chars_per_line(program_col_width, font_pt)
    cpl_edge = _approx_chars_per_line(edge_col_width, font_pt)
    targets: dict[int, float] = {r: 14.0 for r in range(min_row, max_row + 1)}

    for m in ws.merged_cells.ranges:
        if m.min_col != m.max_col or m.min_col < 2 or m.max_col > 8:
            continue
        if m.min_row < min_row or m.max_row > max_row:
            continue
        if m.max_row <= m.min_row:
            continue
        cell = ws.cell(row=m.min_row, column=m.min_col)
        if isinstance(cell, MergedCell):
            continue
        text = _grid_cell_plain_text(cell.value)
        lines = _estimate_wrapped_lines(text, cpl_prog)
        need = _row_height_for_wrapped_lines(lines, font_pt)
        nrows = m.max_row - m.min_row + 1
        each = need / nrows
        for r in range(m.min_row, m.max_row + 1):
            targets[r] = max(targets[r], each)

    multi_row_masters: set[tuple[int, int]] = set()
    for m in ws.merged_cells.ranges:
        if (
            m.min_col == m.max_col
            and 2 <= m.min_col <= 8
            and m.max_row > m.min_row
            and m.min_row >= min_row
            and m.max_row <= max_row
        ):
            multi_row_masters.add((m.min_row, m.min_col))

    for r in range(min_row, max_row + 1):
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if (r, c) in multi_row_masters:
                continue
            text = _grid_cell_plain_text(cell.value)
            lines = _estimate_wrapped_lines(text, cpl_prog)
            targets[r] = max(targets[r], _row_height_for_wrapped_lines(lines, font_pt))

    for r in range(min_row, max_row + 1):
        for c in (1, 9):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            text = _grid_cell_plain_text(cell.value)
            lines = _estimate_wrapped_lines(text, cpl_edge)
            targets[r] = max(targets[r], _row_height_for_wrapped_lines(lines, font_pt))

    for r, h in targets.items():
        ws.row_dimensions[r].height = h


def sheet_label(monday_iso: str) -> str:
    d = parse_monday(monday_iso)
    return f"{d.month}-{d.day}-{d.year}"


def binge_rows_to_dataframe(rows: list[BingeRow]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "DATE": [r.date for r in rows],
            "START TIME": [r.start for r in rows],
            "FINISH TIME ": [r.finish for r in rows],
            "EPISODE": [r.episode for r in rows],
            "SHOW": [r.show for r in rows],
            "EPISODE #": [r.episode_num for r in rows],
            "EPISODE NAME ": [r.episode_name for r in rows],
        }
    )


def _binge_date_display(d: Any) -> str:
    if hasattr(d, "date"):
        d = d.date()
    if not isinstance(d, date):
        return str(d)
    wd = _WEEKDAY_ABBR[d.weekday()]
    return f"{wd}, {d.month:02d}/{d.day:02d}/{d.year}"


def _binge_day_fill_order(df: pd.DataFrame) -> dict[Any, int]:
    order: list[Any] = []
    for v in df["DATE"]:
        dv = v.date() if hasattr(v, "date") else v
        if dv not in order:
            order.append(dv)
    return {d: i for i, d in enumerate(order)}


def _write_binge_sheet(ws, df: pd.DataFrame) -> None:
    n_cols = len(df.columns)
    bold_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10, bold=False)
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
    body_left = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for c, name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold_font
        cell.alignment = header_align
        if c == 4:
            cell.fill = _FILL_HEADER_D
        if c == 4:
            cell.border = _BORDER_BINGE_D
        elif c == 5:
            cell.border = _BORDER_BINGE_E
        else:
            cell.border = _BORDER_GRID

    day_index = _binge_day_fill_order(df)

    for r, row in enumerate(df.itertuples(index=False), start=2):
        dval = row[0]
        dv = dval.date() if hasattr(dval, "date") else dval
        band = day_index.get(dv, 0) % 2
        row_fill = _FILL_DAY_A if band == 0 else _FILL_DAY_B

        for c in range(1, n_cols + 1):
            val = row[c - 1]
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.value = _binge_date_display(val)
            else:
                cell.value = val
            cell.font = body_font
            cell.alignment = body_left
            cell.fill = row_fill
            if c == 4:
                cell.border = _BORDER_BINGE_D
            elif c == 5:
                cell.border = _BORDER_BINGE_E
            else:
                cell.border = _BORDER_GRID

    for c in range(1, n_cols + 1):
        w = _BINGE_COL_WIDTHS[c - 1] if c <= len(_BINGE_COL_WIDTHS) else 14.0
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_binge_notes_sheet(
    ws,
    ui_notes: dict[str, str],
    override_records: Optional[list[BingeRowOverride]] = None,
) -> None:
    ws.cell(row=1, column=1, value="Item")
    ws.cell(row=1, column=2, value="Response")
    row = 2
    for k, v in ui_notes.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        row += 1
    if override_records:
        row += 1
        ws.cell(row=row, column=1, value="Manual row overrides (detail)")
        row += 1
        headers = (
            "Match date",
            "Match start",
            "New date",
            "New start",
            "New finish",
            "Episode",
            "Show",
            "Episode #",
            "Episode name",
        )
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        row += 1
        for o in override_records:
            ws.cell(row=row, column=1, value=o.match_date.isoformat())
            ws.cell(row=row, column=2, value=o.match_start)
            ws.cell(row=row, column=3, value=o.new_date.isoformat())
            ws.cell(row=row, column=4, value=o.new_start)
            ws.cell(row=row, column=5, value=o.new_finish)
            ws.cell(row=row, column=6, value=o.new_episode)
            ws.cell(row=row, column=7, value=o.new_show)
            ws.cell(row=row, column=8, value=o.new_episode_num)
            ws.cell(row=row, column=9, value=o.new_episode_name)
            row += 1


def write_binge_workbook(
    path: Path,
    sheets: dict[str, pd.DataFrame],
    *,
    ui_notes: Optional[dict[str, str]] = None,
    override_records: Optional[list[BingeRowOverride]] = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True
    for name, df in sheets.items():
        safe = name[:31]
        if first:
            ws = wb.active
            ws.title = safe
            first = False
        else:
            ws = wb.create_sheet(safe)
        _write_binge_sheet(ws, df)
    if ui_notes:
        nws = wb.create_sheet("BINGE notes", 0)
        _write_binge_notes_sheet(nws, ui_notes, override_records)
    wb.save(path)


def _apply_segment_merges(ws, grid: list[list[Optional[str]]], cfg: BuildConfig) -> None:
    excel_data_start = 5
    for day_idx in range(7):
        col = [grid[r][day_idx] for r in range(48)]
        for seg in segments_for_binge_scheduling(col, cfg):
            span = seg.end_slot - seg.start_slot
            if span > 1:
                ws.merge_cells(
                    start_row=excel_data_start + seg.start_slot,
                    end_row=excel_data_start + seg.end_slot - 1,
                    start_column=day_idx + 2,
                    end_column=day_idx + 2,
                )


def _apply_grids_program_richtext(ws, font_pt: int) -> None:
    for r in range(5, 53):
        for c in range(2, 9):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if not isinstance(v, str):
                continue
            rich = _grids_program_rich_text(v, font_pt)
            if isinstance(rich, CellRichText):
                cell.value = rich
            cell.font = Font(name=_GRID_FONT, size=font_pt, bold=False)


def _write_grids_sheet(
    ws,
    mat: list[list[Any]],
    grid: list[list[Optional[str]]],
    cfg: BuildConfig,
    prev_grid: Optional[list[list[Optional[str]]]] = None,
) -> None:
    gracenote = cfg.gracenote_binge_id
    pt_banner = _GRID_PT_BANNER
    pt_sub = _GRID_PT_SUB
    pt_date = _GRID_PT_DATE
    pt_body = _GRID_PT_BODY
    font_banner = Font(name=_GRID_FONT, size=pt_banner, bold=True)
    font_sub = Font(name=_GRID_FONT, size=pt_sub, bold=False)
    font_date = Font(name=_GRID_FONT, size=pt_date, bold=False)
    font_body = Font(name=_GRID_FONT, size=pt_body, bold=False)
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    base_uniform = _grids_uniform_width(_max_line_length_program_body(mat), pt_body)
    uniform_w = base_uniform * _GRID_COL_WIDTH_SCALE

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c1 = ws.cell(row=1, column=1, value="BINGE")
    c1.font = font_banner
    c1.alignment = align_center_wrap
    ws.row_dimensions[1].height = 22.0

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    c2 = ws.cell(row=2, column=1, value=f"Gracenote BINGE ID: {gracenote}")
    c2.font = font_sub
    c2.alignment = align_center_wrap
    ws.row_dimensions[2].height = 15.0

    # mat[2] = date row, mat[3] = weekday row, mat[4:52] = grid body
    ws.row_dimensions[3].height = 12.0
    for c in range(1, 10):
        val = mat[2][c - 1]
        cell = ws.cell(row=3, column=c, value=val)
        cell.font = font_date
        cell.alignment = align_center_wrap
        if isinstance(val, datetime):
            cell.number_format = "m/d/yyyy"
        elif isinstance(val, date):
            cell.value = datetime.combine(val, time())
            cell.number_format = "m/d/yyyy"

    for c in range(1, 10):
        val = mat[3][c - 1]
        cell = ws.cell(row=4, column=c, value=val)
        cell.font = font_body
        cell.alignment = align_center_wrap
    ws.row_dimensions[4].height = 12.0

    for r_off in range(48):
        excel_r = 5 + r_off
        row_vals = mat[4 + r_off]
        for c in range(1, 10):
            val = row_vals[c - 1]
            cell = ws.cell(row=excel_r, column=c, value=val)
            cell.font = font_body
            cell.alignment = align_center_wrap
            if isinstance(val, time):
                cell.number_format = "h:mm AM/PM;@"

    _apply_segment_merges(ws, grid, cfg)
    _apply_grids_program_richtext(ws, pt_body)

    # Col A & I = half width of Mon–Sun (B–H); all widths scaled vs. legacy formula.
    edge_w = uniform_w / 2.0
    ws.column_dimensions[get_column_letter(1)].width = edge_w
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = uniform_w
    ws.column_dimensions[get_column_letter(9)].width = edge_w

    _adjust_grids_body_row_heights(ws, uniform_w, edge_w, pt_body)

    # Thin black borders on grid area (skip merged slave cells)
    for r in range(1, 53):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = _BORDER_GRID

    # Yellow = slots whose program text changed vs. the prior week (legacy schedule diff).
    if prev_grid is not None and _grid_shape_ok(prev_grid) and _grid_shape_ok(grid):
        for slot in range(48):
            for day_idx in range(7):
                if _slot_show_equal(grid[slot][day_idx], prev_grid[slot][day_idx]):
                    continue
                cell = ws.cell(row=5 + slot, column=2 + day_idx)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = _FILL_GRID_HIGHLIGHT


def write_grids_workbook(
    path: Path,
    sheets: dict[str, tuple[list[list[Any]], list[list[Optional[str]]]]],
    cfg: BuildConfig,
    *,
    ui_notes: Optional[dict[str, str]] = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    prev_grid: Optional[list[list[Optional[str]]]] = None
    if ui_notes:
        # Week grids first so Excel opens on a real grid tab; notes at the end.
        first = True
        for name, (mat, grid) in sheets.items():
            safe = name[:31]
            if first:
                ws = wb.active
                ws.title = safe
                first = False
            else:
                ws = wb.create_sheet(safe)
            _write_grids_sheet(ws, mat, grid, cfg, prev_grid=prev_grid)
            prev_grid = grid
        nws = wb.create_sheet("BINGE notes")
        _write_binge_notes_sheet(nws, ui_notes, None)
    else:
        first = True
        for name, (mat, grid) in sheets.items():
            safe = name[:31]
            if first:
                ws = wb.active
                ws.title = safe
                first = False
            else:
                ws = wb.create_sheet(safe)
            _write_grids_sheet(ws, mat, grid, cfg, prev_grid=prev_grid)
            prev_grid = grid
    wb.save(path)


def export_both(
    cfg: BuildConfig,
    out_dir: Path,
    *,
    weeks: Optional[list[WeekDef]] = None,
    binge_row_overrides: Optional[list[BingeRowOverride]] = None,
    binge_ui_notes: Optional[dict[str, str]] = None,
    export_stations: Optional[Sequence[str]] = None,
) -> tuple[Path, Path, list[str], list[str]]:
    """Write **BINGE.xlsx** (full episode rows from Nikki) and **BINGE GRIDS.xlsx** (weekly strip layout).

    **BINGE** comes from ``rows_for_week`` (Nikki + cursors; optional ``reference_binge_file`` for advance/repeat).
    If ``reference_binge_literal_copy_before`` is set, rows before that calendar date are **replaced** with the
    reference April workbook (verbatim); later days stay generated. Cursors are reconciled from the final sheet.
    If the requested weeks skip a preceding ISO week (e.g. **May-only** without the Apr 27–May 3 sheet), the
    missing week is simulated once **only to advance cursors** so the first exported Monday follows Sunday night.
    **BINGE GRIDS** program cells match the source grids workbooks (show titles only) — not the enriched BINGE episode text.

    If ``export_stations`` (or YAML ``export_stations``) is set, also copies both workbooks into
    ``out_dir/<sanitized_label>/`` per station. If ``save_binge_reference_copy_to`` is set on the config, copies
    the generated ``BINGE.xlsx`` to that path (overwrites) for use as next month’s ``reference_binge_file``.
    """
    week_list = list(weeks if weeks is not None else cfg.weeks)
    if not week_list:
        raise ValueError("No weeks to export: pass ``weeks=`` or add entries under ``weeks:`` in config.")
    warm = _cursor_warmup_week_if_needed(week_list, cfg.weeks)
    ensure_grids_workbooks_for_weeks(([warm] if warm else []) + week_list)
    seed_messages = seed_grids_from_prior_month(week_list, cfg.weeks)
    seed_messages.extend(sync_straddle_weeks_to_canonical_grids_file(week_list))
    cat = build_catalog(cfg)
    apply_saved_cursors(cat, resolved_cursor_state_path(cfg))
    episode_actions, _, _ = load_reference_episode_actions(cfg)
    binge_sheets: dict[str, pd.DataFrame] = {}
    week_by_label: dict[str, WeekDef] = {}
    grid_raw_by_label: dict[str, list[list[Optional[str]]]] = {}

    sync_mondays = {str(x).strip() for x in (cfg.reference_binge_sync_cursor_weeks or []) if str(x).strip()}

    to_process: list[tuple[WeekDef, bool]] = []
    if warm:
        to_process.append((warm, True))
    for wk in sorted(week_list, key=lambda w: parse_monday(w.monday)):
        to_process.append((wk, False))

    prev_merged_df: Optional[pd.DataFrame] = None

    for wk, warmup_only in to_process:
        mon = parse_monday(wk.monday)
        if wk.monday in sync_mondays:
            wdf = load_reference_week_dataframe(cfg, mon)
            if wdf is not None:
                seed_messages.extend(
                    sync_cursors_from_reference_binge_week(
                        cfg,
                        cat,
                        wdf,
                        monday_label=wk.monday,
                    )
                )

        grid_raw = load_grid_sheet(wk.grids_file, wk.sheet_name)
        label = sheet_label(wk.monday)
        rows = rows_for_week(
            cfg, cat, grid_raw, wk.monday, episode_actions=episode_actions
        )
        df_norm = normalize_binge_df_columns(binge_rows_to_dataframe(rows))
        merged_df, _literal_notes = merge_literal_reference_binge_days(cfg, mon, df_norm)
        merged_df = apply_overnight_repeats_with_prev(cfg, cat, merged_df, prev_merged_df, mon)
        reconcile_catalog_from_binge_dataframe(cfg, cat, merged_df)
        prev_merged_df = merged_df
        if warmup_only:
            continue
        grid_raw_by_label[label] = grid_raw
        binge_sheets[label] = merged_df
        week_by_label[label] = wk

    override_warnings: list[str] = []
    if binge_row_overrides:
        for label, df in binge_sheets.items():
            updated, msgs = apply_binge_row_overrides(df, binge_row_overrides)
            binge_sheets[label] = updated
            for m in msgs:
                override_warnings.append(f"{label}: {m}")

    grid_sheets: dict[str, tuple[list[list[Any]], list[list[Optional[str]]]]] = {}
    for label, wk in sorted(week_by_label.items(), key=lambda x: parse_monday(x[1].monday)):
        mon = parse_monday(wk.monday)
        grid_raw = grid_raw_by_label[label]
        mat = build_grids_matrix(mon, grid_raw, cfg.gracenote_binge_id)
        grid_sheets[label] = (mat, grid_raw)

    binge_path = out_dir / "BINGE.xlsx"
    grids_path = out_dir / "BINGE GRIDS.xlsx"
    write_binge_workbook(
        binge_path,
        binge_sheets,
        ui_notes=binge_ui_notes,
        override_records=binge_row_overrides if binge_row_overrides else None,
    )
    write_grids_workbook(grids_path, grid_sheets, cfg, ui_notes=None)
    save_cursors_after_export(cat, resolved_cursor_state_path(cfg))

    archive_path = _resolved_save_binge_reference_copy_path(cfg)
    if archive_path is not None:
        archive_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(binge_path, archive_path)
        seed_messages.append(f"Archived BINGE reference copy to {archive_path} (overwrites if present).")

    station_labels = _export_station_labels(cfg, export_stations)
    for label in station_labels:
        sub = _sanitize_station_dir(label)
        dest_dir = out_dir / sub
        dest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(binge_path, dest_dir / "BINGE.xlsx")
        shutil.copy2(grids_path, dest_dir / "BINGE GRIDS.xlsx")
        seed_messages.append(f"Station copy [{label}]: {dest_dir / 'BINGE.xlsx'}")

    return binge_path, grids_path, override_warnings, seed_messages


def export_grids_from_binge_sheets(
    cfg: BuildConfig,
    physical_sheets: dict[str, pd.DataFrame],
    out_dir: Path,
    *,
    binge_row_overrides: Optional[list[BingeRowOverride]] = None,
    binge_ui_notes: Optional[dict[str, str]] = None,
) -> tuple[Path, list[str]]:
    """Build **BINGE GRIDS.xlsx** only from BINGE list workbook sheets (DataFrames)."""
    merged: dict[str, tuple[date, list[pd.DataFrame]]] = {}
    for _phys_name, df in physical_sheets.items():
        for mon, part in split_binge_df_by_monday(df).items():
            lab = sheet_label(mon.isoformat())
            if lab not in merged:
                merged[lab] = (mon, [])
            merged[lab][1].append(part)

    combined: list[tuple[str, date, pd.DataFrame]] = []
    for lab, (mon, parts) in merged.items():
        pdf = parts[0] if len(parts) == 1 else pd.concat(parts, ignore_index=True)
        combined.append((lab, mon, pdf))

    combined.sort(key=lambda x: x[1])

    grid_sheets: dict[str, tuple[list[list[Any]], list[list[Optional[str]]]]] = {}
    warnings: list[str] = []

    for lab, mon, pdf in combined:
        pdf2 = pdf.copy()
        if binge_row_overrides:
            pdf2, msgs = apply_binge_row_overrides(pdf2, binge_row_overrides)
            for m in msgs:
                warnings.append(f"{lab}: {m}")
        grid = binge_dataframe_to_grid(pdf2, mon)
        mat = build_grids_matrix(mon, grid, cfg.gracenote_binge_id)
        grid_sheets[lab] = (mat, grid)

    out_dir.mkdir(parents=True, exist_ok=True)
    grids_path = out_dir / "BINGE GRIDS.xlsx"
    write_grids_workbook(grids_path, grid_sheets, cfg, ui_notes=binge_ui_notes)
    return grids_path, warnings
