from __future__ import annotations

import re
from typing import TYPE_CHECKING, Optional

import pandas as pd

from binge_schedule.models import Episode, NikkiColumnHeaders

if TYPE_CHECKING:
    from binge_schedule.models import ShowDef


def effective_column_headers(sd: "ShowDef", *, style: str) -> NikkiColumnHeaders:
    """YAML ``nikki_columns:`` if set; else defaults for series vs ``movies`` tab."""
    if sd.nikki_columns is not None:
        return sd.nikki_columns
    if style == "movies":
        return NikkiColumnHeaders.movies_tab()
    return NikkiColumnHeaders.standard_series()


def _norm_header(s: str) -> str:
    return " ".join(str(s).replace("\xa0", " ").split()).casefold()


def _row_header_index_map(row: pd.Series) -> dict[str, int]:
    """Map normalized header text → first column index on this row."""
    m: dict[str, int] = {}
    for j in range(len(row)):
        v = row.iloc[j]
        if pd.isna(v):
            continue
        key = _norm_header(str(v))
        if key and key not in m:
            m[key] = j
    return m


def _find_header_row_and_columns(
    df: pd.DataFrame, headers: NikkiColumnHeaders
) -> tuple[Optional[int], dict[str, int]]:
    """First row that contains every required header label; return column indices by logical name."""
    required: list[tuple[str, str]] = [("episode", headers.episode)]
    if headers.season_episode:
        required.append(("season_episode", headers.season_episode))
    if headers.year:
        required.append(("year", headers.year))

    for i in range(min(35, len(df))):
        row = df.iloc[i]
        hm = _row_header_index_map(row)
        idx: dict[str, int] = {}
        ok = True
        for logical, label in required:
            lk = _norm_header(label)
            if lk not in hm:
                ok = False
                break
            idx[logical] = hm[lk]
        if not ok:
            continue
        if headers.stars:
            lk = _norm_header(headers.stars)
            if lk in hm:
                idx["stars"] = hm[lk]
        if headers.synopsis:
            lk = _norm_header(headers.synopsis)
            if lk in hm:
                idx["synopsis"] = hm[lk]
        return i, idx
    return None, {}


def _clean(s: str) -> str:
    return " ".join(str(s).replace("\xa0", " ").split()).strip()


def _parse_jim_bowie_line(cell: str) -> Optional[tuple[int, str]]:
    s = _clean(cell)
    s = re.sub(r"^:\s*", "", s)
    m = re.match(r"S(\d+)\s*E(\d+)\s*-\s*(.+)", s, re.I)
    if not m:
        m = re.match(r"(\d+)\s*E(\d+)\s*-\s*(.+)", s, re.I)
    if not m:
        return None
    season, ep, title = int(m.group(1)), int(m.group(2)), _clean(m.group(3))
    code_num = season * 100 + ep
    return code_num, title


def load_jim_bowie(df: pd.DataFrame) -> list[Episode]:
    out: list[Episode] = []
    for i in range(len(df)):
        cell = df.iloc[i, 0]
        if pd.isna(cell):
            continue
        parsed = _parse_jim_bowie_line(str(cell))
        if not parsed:
            continue
        num, title = parsed
        code = f"AJB{num}"
        out.append(Episode(raw=str(cell), title=title, code=code, episode_num=num))
    return out


def _skip_instruction_row(
    df: pd.DataFrame, i: int, header_row: int, col_idx: dict[str, int]
) -> bool:
    """Skip Carol-style marker rows (e.g. PLAY EPISODES IN GREEN)."""
    if i <= header_row:
        return True
    stars = None
    syn = None
    sc = col_idx.get("stars")
    if sc is not None and sc < len(df.columns):
        stars = df.iloc[i, sc]
    syn_c = col_idx.get("synopsis")
    if syn_c is not None and syn_c < len(df.columns):
        syn = df.iloc[i, syn_c]
    if pd.notna(stars) and str(stars).strip().upper() == "ONLY":
        return True
    if pd.notna(syn) and "PLAY EPISODES" in str(syn).upper():
        return True
    return False


def _year_int(val: object) -> Optional[int]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if hasattr(val, "year"):
        try:
            return int(val.year)  # type: ignore[attr-defined]
        except (TypeError, ValueError):
            pass
    try:
        y = int(val)
        if 1800 <= y <= 2100:
            return y
    except (TypeError, ValueError):
        pass
    return None


def load_movies(
    df: pd.DataFrame, *, prefix: str = "MOV", columns: NikkiColumnHeaders
) -> list[Episode]:
    """Movies catalog: primary column is ``Title`` (see ``NikkiColumnHeaders.movies_tab``)."""
    hr, col_idx = _find_header_row_and_columns(df, columns)
    if hr is None or "episode" not in col_idx:
        return []
    title_col = col_idx["episode"]
    year_col = col_idx.get("year")
    start = hr + 1
    out: list[Episode] = []
    for i in range(start, len(df)):
        tcell = df.iloc[i, title_col]
        if pd.isna(tcell):
            continue
        title = _clean(str(tcell))
        if not title or _norm_header(title) == _norm_header(columns.episode):
            continue
        year: Optional[int] = None
        if year_col is not None and year_col < len(df.columns):
            year = _year_int(df.iloc[i, year_col])
        display = f"{title} ({year})" if year is not None else title
        seq = len(out) + 1
        code = f"{prefix or 'MOV'}{seq}"
        out.append(
            Episode(
                raw=str(tcell),
                title=display,
                code=code,
                episode_num=seq,
                season_ep=None,
            )
        )
    return out


def _title_from_episode_cell(cell: str) -> str:
    s = _clean(cell)
    if " - " in s:
        return s.split(" - ", 1)[-1].strip()
    return s


def _code_from_hunter(cell: str) -> Optional[tuple[str, int]]:
    m = re.search(r"HUN_(\d+)", cell, re.I)
    if m:
        n = int(m.group(1))
        return f"HUN{n}", n
    return None


def _code_from_texan(cell: str) -> Optional[tuple[str, int]]:
    m = re.search(r"S(\d+)_EP(\d+)", cell, re.I)
    if m:
        s, e = int(m.group(1)), int(m.group(2))
        n = s * 100 + e
        return f"TEX{n}", n
    return None


def _code_from_renegade(cell: str) -> Optional[tuple[str, int]]:
    m = re.search(r"S(\d+)E(\d+)", cell, re.I)
    if m:
        s, e = int(m.group(1)), int(m.group(2))
        n = s * 100 + e
        return f"REN{n}", n
    return None


def _code_from_leading_digits(cell: str, prefix: str) -> Optional[tuple[str, int]]:
    m = re.match(r"(\d+)\s*-\s*", cell)
    if m:
        n = int(m.group(1))
        return f"{prefix}{n}", n
    return None


def _code_from_carol(cell: str, season_ep: str) -> tuple[str, int]:
    m = re.match(r"(\d+)\s*-\s*", cell)
    if m:
        n = int(m.group(1))
        return f"CBS{n}", n
    parts = str(season_ep).split("_")
    if len(parts) == 2:
        try:
            a, b = int(parts[0]), int(parts[1])
            n = a * 100 + b
            return f"CBS{n}", n
        except ValueError:
            pass
    h = abs(hash(cell)) % 900000 + 100000
    return f"CBS{h}", h


def _code_from_mst3k(cell: str, season_ep: str) -> tuple[str, int]:
    parts = str(season_ep).split("_")
    if len(parts) == 2:
        try:
            a, b = int(parts[0]), int(parts[1])
            n = a * 100 + b
            return f"MST{n}", n
        except ValueError:
            pass
    h = abs(hash(cell)) % 900 + 100
    return f"MST{h}", h


def _code_from_jmp(cell: str) -> Optional[tuple[str, int]]:
    m = re.search(r"JMP_(\d+)", cell, re.I)
    if m:
        n = int(m.group(1))
        return f"JMP{n}", n
    return None


def _code_from_se_underscore(cell: str, prefix: str) -> Optional[tuple[str, int]]:
    m = re.search(r"S(\d+)_E(\d+)", cell, re.I)
    if m:
        s, e = int(m.group(1)), int(m.group(2))
        n = s * 100 + e
        return f"{prefix}{n}", n
    return None


# ``ShowDef.nikki_row_filter`` — extend when new per-sheet rules appear.
ROW_FILTER_GREEN_EPISODE_CELL = "green_episode_cell"


def _rgb_string_from_openpyxl_color(color) -> Optional[str]:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return None
    if isinstance(rgb, str):
        return rgb.upper()
    return str(rgb).upper()


def _is_schedule_green_fill(fill) -> bool:
    """True when the Episode cell uses the same green fill as playable Carol rows."""
    if fill is None or fill.fill_type is None:
        return False
    ft = str(fill.fill_type).lower()
    if ft not in ("solid", "patternfill"):
        return False
    hx = _rgb_string_from_openpyxl_color(fill.start_color)
    if not hx:
        return False
    hx = hx.replace(" ", "").upper()
    # Carol 2024 content workbook: Excel theme green on playable episodes.
    if hx == "FF92D050":
        return True
    if len(hx) >= 6:
        tail = hx[-6:]
        try:
            r = int(tail[0:2], 16)
            g = int(tail[2:4], 16)
            b = int(tail[4:6], 16)
        except ValueError:
            return False
        return g >= 150 and g > r + 18 and g > b + 12
    return False


def _green_episode_row_indices(
    workbook_path: str,
    sheet_name: str,
    *,
    header_row: int,
    ep_col: int,
    n_df_rows: int,
) -> set[int]:
    """0-based DataFrame row indices whose Episode cell is filled green (playable)."""
    import openpyxl

    out: set[int] = set()
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    except OSError:
        return out
    try:
        if sheet_name not in wb.sheetnames:
            return out
        ws = wb[sheet_name]
        for i in range(header_row + 1, n_df_rows):
            c = ws.cell(row=i + 1, column=ep_col + 1)
            if _is_schedule_green_fill(c.fill):
                out.add(i)
        return out
    finally:
        wb.close()


def load_standard_sheet(
    df: pd.DataFrame,
    *,
    style: str = "generic",
    prefix: str = "",
    columns: NikkiColumnHeaders,
    workbook_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    row_filter: Optional[str] = None,
) -> list[Episode]:
    """Load rows under a header row located by ``columns`` (exact header titles)."""
    hr, col_idx = _find_header_row_and_columns(df, columns)
    if hr is None or "episode" not in col_idx:
        return []
    ep_col = col_idx["episode"]
    se_col = col_idx.get("season_episode")
    start = hr + 1
    green_rows: Optional[set[int]] = None
    if row_filter == ROW_FILTER_GREEN_EPISODE_CELL:
        if workbook_path and sheet_name:
            green_rows = _green_episode_row_indices(
                workbook_path,
                sheet_name,
                header_row=hr,
                ep_col=ep_col,
                n_df_rows=len(df),
            )
        else:
            green_rows = set()
    out: list[Episode] = []
    for i in range(start, len(df)):
        if _skip_instruction_row(df, i, hr, col_idx):
            continue
        if green_rows is not None and i not in green_rows:
            continue
        cell = df.iloc[i, ep_col]
        if pd.isna(cell):
            continue
        raw = str(cell)
        s = _clean(raw)
        if not s or _norm_header(s) == _norm_header(columns.episode):
            continue
        season_ep = ""
        if se_col is not None and se_col < len(df.columns):
            v1 = df.iloc[i, se_col]
            if pd.notna(v1) and str(v1).strip():
                season_ep = str(v1).strip()
        title = _title_from_episode_cell(s)

        code: Optional[str] = None
        epn: Optional[int] = None

        if style == "hunter":
            t = _code_from_hunter(s)
            if t:
                code, epn = t
        elif style == "texan":
            t = _code_from_texan(s)
            if t:
                code, epn = t
        elif style == "renegade":
            t = _code_from_renegade(s)
            if t:
                code, epn = t
        elif style == "real_mccoys":
            t = _code_from_leading_digits(s, prefix or "MCC")
            if t:
                code, epn = t
        elif style == "leading_episode":
            t = _code_from_leading_digits(s, prefix or "EP")
            if t:
                code, epn = t
        elif style == "carol_burnett":
            code, epn = _code_from_carol(s, season_ep)
        elif style == "mst3k":
            code, epn = _code_from_mst3k(s, season_ep)
        elif style == "jmp":
            t = _code_from_jmp(s)
            if t:
                code, epn = t
            else:
                code, epn = f"JMP{i}", i
        elif style in ("saint", "laugh_in"):
            pfx = prefix or ("SNT" if style == "saint" else "RML")
            t = _code_from_se_underscore(s, pfx)
            if t:
                code, epn = t
            else:
                code, epn = f"{pfx}{i}", i
        else:
            seq = len(out) + 1
            if prefix:
                code, epn = f"{prefix}{seq}", seq
            else:
                code, epn = f"EP{seq}", seq

        if code is None:
            seq = len(out) + 1
            code, epn = f"EP{seq}", seq

        out.append(Episode(raw=raw, title=title, code=code, episode_num=epn, season_ep=season_ep or None))
    return out


STYLE_BY_PREFIX = {
    "AJB": "jim_bowie",
    "TEX": "texan",
    "HUN": "hunter",
    "REN": "renegade",
    "MCC": "real_mccoys",
    "CBS": "carol_burnett",
    "MST": "mst3k",
}


def _is_new_shows_catalog_df(df: pd.DataFrame) -> bool:
    """True for the multi-show ``NEW SHOWS`` layout (Artist/Series + episode titles)."""
    if len(df) < 2 or len(df.columns) < 3:
        return False
    bits: list[str] = []
    for j in range(min(5, len(df.columns))):
        v = df.iloc[0, j]
        if pd.notna(v):
            bits.append(str(v).casefold())
    h = " ".join(bits)
    return "artist" in h and "series" in h and "title" in h


def load_new_shows_catalog(df: pd.DataFrame, *, prefix: str = "NS") -> list[Episode]:
    """Load the ``NEW SHOWS``-style catalog: one row per playable line with Artist/Series in column 0."""
    out: list[Episode] = []
    for i in range(1, len(df)):
        artist = _clean(str(df.iloc[i, 0])) if len(df.columns) > 0 else ""
        ti = _clean(str(df.iloc[i, 1])) if len(df.columns) > 1 else ""
        st = _clean(str(df.iloc[i, 2])) if len(df.columns) > 2 else ""
        if not ti and not st:
            continue
        if st and st == artist:
            continue
        if ti and ti == artist and (not st or st == artist):
            continue
        raw = ti or st
        if not raw or len(raw) < 2:
            continue
        display = f"{artist} — {st}" if artist else (st or ti)
        if len(display) < 4:
            continue
        seq = len(out) + 1
        code = f"{prefix}{seq}"
        out.append(Episode(raw=raw, title=display, code=code, episode_num=seq, season_ep=None))
    return out


def load_sheet(
    path: str,
    sheet_name: str,
    *,
    style: str,
    prefix: str = "",
    columns: NikkiColumnHeaders,
    row_filter: Optional[str] = None,
) -> list[Episode]:
    """Load episodes from the content workbook.

    ``row_filter`` (optional, from YAML ``nikki_row_filter``):

    - ``green_episode_cell`` — only rows whose **Episode** column cell has the sheet’s
      green fill (e.g. Carol Burnett “play only green” rule). Requires ``path`` /
      ``sheet_name`` so fills can be read with openpyxl.
    """
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)
    if _is_new_shows_catalog_df(df):
        return load_new_shows_catalog(df, prefix=prefix or "NS")
    if style == "jim_bowie":
        return load_jim_bowie(df)
    if style == "movies":
        return load_movies(df, prefix=prefix, columns=columns)
    return load_standard_sheet(
        df,
        style=style,
        prefix=prefix,
        columns=columns,
        workbook_path=path,
        sheet_name=sheet_name,
        row_filter=row_filter,
    )


def default_style_for_sheet(sheet_name: str) -> str:
    slug = sheet_name.strip().lower()
    if slug == "movies":
        return "movies"
    u = sheet_name.upper()
    if "JIM BOWIE" in u or "BOWIE" in u:
        return "jim_bowie"
    if "TEXAN" in u:
        return "texan"
    if sheet_name == "Hunter" or u == "HUNTER":
        return "hunter"
    if "RENEGADE" in u:
        return "renegade"
    if "REAL MCCOY" in u:
        return "real_mccoys"
    if "CAROL BURNETT" in u:
        return "carol_burnett"
    if "MST3K" in u or "MYSTERY SCIENCE" in u:
        return "mst3k"
    if "21 JUMP" in u:
        return "jmp"
    if "SAINT" in u and "SHERLOCK" not in u:
        return "saint"
    if "LAUGH-IN" in u or "LAUGH IN" in u:
        return "laugh_in"
    if "TIM CONWAY COMEDY" in u:
        return "leading_episode"
    return "generic"
