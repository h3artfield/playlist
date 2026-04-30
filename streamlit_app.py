"""
Schedule Builder — Streamlit UI to build BINGE exports, browse the content archive, and edit schedule sources.

Run from the project directory:
  streamlit run streamlit_app.py
"""

from __future__ import annotations

import hashlib
import io
import inspect
import json
import os
import platform
import re
import shutil
import subprocess
import tempfile
from dataclasses import asdict
from datetime import date, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd
import streamlit as st

from binge_schedule import nikki
from binge_schedule.archive_normalize import normalize_episodes_for_archive
from binge_schedule.binge_overrides import BingeRowOverride, parse_flexible_time
from binge_schedule.config_io import load_build_config
from binge_schedule.models import NikkiColumnHeaders, ShowDef
from binge_schedule.cursor_state import resolved_cursor_state_path, resolved_nikki_workbook_path
from binge_schedule.binge_to_grid import (
    normalize_binge_df_columns,
    read_binge_workbook_sheets,
    split_binge_df_by_monday,
)
from binge_schedule.export_xlsx import export_both, is_verbose_seed_noise
from binge_schedule.show_swap import apply_show_swap, parse_schedule_anchor
from binge_schedule.show_resolve import resolve_show
from binge_schedule.grid import (
    day_dates,
    ensure_grids_workbooks_for_weeks,
    load_grid_sheet,
    parse_monday,
    parse_sheet_tab_monday,
    segments_for_day,
    slot_label,
    weeks_with_monday_in_calendar_month,
)
from binge_schedule.workbook_discover import (
    parse_workbook_tab_option,
    synthetic_series_for_tab,
    workbook_tabs_not_in_yaml,
    workbook_tab_option,
)

_RAW_LITERAL_PREFIX = "__literal_text__:"
_IMPORTED_CONTENT_PREFIX = "__imported_content__:"


def _secret_or_env(key: str) -> str:
    val = str(os.environ.get(key, "") or "").strip()
    if val:
        return val
    try:
        sval = st.secrets.get(key, "")  # type: ignore[attr-defined]
    except Exception:
        return ""
    return str(sval or "").strip()


def _desktop_download_meta() -> dict[str, str]:
    url = _secret_or_env("DESKTOP_APP_DOWNLOAD_URL")
    if not url:
        repo = _secret_or_env("DESKTOP_APP_GITHUB_REPO") or "h3artfield/playlist"
        url = f"https://github.com/{repo}/releases/latest/download/ScheduleBuilderSetup.exe"
    return {
        "url": url,
        "label": _secret_or_env("DESKTOP_APP_LABEL") or "Download Desktop App (Windows)",
        "version": _secret_or_env("DESKTOP_APP_VERSION"),
        "notes_url": _secret_or_env("DESKTOP_APP_RELEASE_NOTES_URL"),
    }


def _render_desktop_download_cta() -> None:
    if _secret_or_env("SCHEDULE_BUILDER_DESKTOP_RUNTIME") == "1":
        return
    meta = _desktop_download_meta()
    if not meta:
        return
    c1, c2 = st.columns([3, 2], vertical_alignment="center")
    with c1:
        extra = f" (v{meta['version']})" if meta.get("version") else ""
        st.caption(f"Install Schedule Builder locally on Windows{extra}.")
    with c2:
        if hasattr(st, "link_button"):
            st.link_button(meta["label"], meta["url"], use_container_width=True, type="primary")
        else:
            st.markdown(f"[{meta['label']}]({meta['url']})")
    if meta.get("notes_url"):
        st.caption(f"[Release notes]({meta['notes_url']})")


def _available_base_schedule_files() -> list[str]:
    cfg_dir = Path("config")
    if not cfg_dir.is_dir():
        return []
    files: list[Path] = []
    files.extend(cfg_dir.glob("*.yaml"))
    files.extend(cfg_dir.glob("*.yml"))
    uniq: dict[str, Path] = {}
    for p in files:
        if p.is_file():
            uniq[p.as_posix()] = p
    ordered = sorted(
        uniq.values(),
        key=lambda p: (float(p.stat().st_mtime) if p.exists() else 0.0, p.name.casefold()),
        reverse=True,
    )
    return [p.as_posix() for p in ordered]


def _friendly_date(d: date) -> str:
    return f"{d.strftime('%b')} {d.day}, {d.year}"


def _base_schedule_records() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for p_str in _available_base_schedule_files():
        p = Path(p_str)
        try:
            cfg = load_build_config(p)
        except Exception:
            continue
        anchor: Optional[date] = None
        cutoff = getattr(cfg, "reference_binge_literal_copy_before", None)
        if cutoff:
            try:
                anchor = date.fromisoformat(str(cutoff)) - timedelta(days=1)
            except Exception:
                try:
                    anchor = parse_monday(str(cutoff)) - timedelta(days=1)
                except Exception:
                    anchor = None
        if anchor is None:
            week_dates = [parse_monday(w.monday) + timedelta(days=6) for w in cfg.weeks]
            anchor = max(week_dates) if week_dates else date.today()
        window_start = anchor - timedelta(days=29)
        mtime = float(p.stat().st_mtime) if p.exists() else 0.0
        records.append(
            {
                "path": p.as_posix(),
                "anchor": anchor,
                "window_start": window_start,
                "mtime": mtime,
                "label": f"Last airdate: {_friendly_date(anchor)}",
                "detail": f"Window: {_friendly_date(window_start)} -> {_friendly_date(anchor)}",
            }
        )
    records.sort(key=lambda r: (r.get("anchor"), r.get("mtime")), reverse=True)
    return records


def _baseline_window_for_cfg(cfg) -> tuple[date, date]:
    anchor: Optional[date] = None
    cutoff = getattr(cfg, "reference_binge_literal_copy_before", None)
    if cutoff:
        try:
            anchor = date.fromisoformat(str(cutoff)) - timedelta(days=1)
        except Exception:
            try:
                anchor = parse_monday(str(cutoff)) - timedelta(days=1)
            except Exception:
                anchor = None
    if anchor is None:
        week_dates = [parse_monday(w.monday) + timedelta(days=6) for w in cfg.weeks]
        anchor = max(week_dates) if week_dates else date.today()
    return anchor - timedelta(days=29), anchor


def _default_config_display() -> str:
    raw = (os.environ.get("BINGE_CONFIG_PATH") or os.environ.get("STREAMLIT_BINGE_CONFIG") or "").strip()
    if raw:
        return Path(raw).as_posix()
    records = _base_schedule_records()
    if records:
        return str(records[0]["path"])
    choices = _available_base_schedule_files()
    if choices:
        return str(choices[0])
    return Path("config/cloud.yaml").as_posix()


def _open_folder(path: Path) -> str:
    p = path.resolve()
    if not p.is_dir():
        return f"Folder not found: {p}"
    system = platform.system()
    try:
        if system == "Windows":
            subprocess.Popen(["explorer.exe", str(p)], close_fds=True)
        elif system == "Darwin":
            subprocess.Popen(["open", str(p)], close_fds=True)
        else:
            opener = shutil.which("xdg-open")
            if not opener:
                return (
                    "Opening a folder isn’t available on this system (e.g. Streamlit Cloud or a phone). "
                    "Use the **BINGE.xlsx** / **BINGE GRIDS.xlsx** download buttons."
                )
            subprocess.Popen([opener, str(p)], close_fds=True)
    except OSError:
        return (
            "Couldn’t open the folder automatically. "
            "On Streamlit Cloud or mobile, use the download buttons to save the files."
        )
    return ""


def _nikki_sheet_exists(workbook: Path, sheet: str | None) -> bool | None:
    if not workbook.is_file() or not sheet:
        return None
    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook, read_only=True)
        ok = sheet in wb.sheetnames
        wb.close()
        return ok
    except OSError:
        return None


def _add_one_month(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _months_for_build_selector(weeks: list) -> list[date]:
    """Months from the earliest ``weeks`` Monday through the month *after* the latest Monday.

    The extra month (e.g. May when only April has ``weeks:``) is the usual **next** build target;
    episode cursors already reflect April after you run April—May still needs its own ``weeks:`` lines.
    The month anchor is still used for unlock sequencing, but actual builds use Start date + week count.
    """
    dates: list[date] = []
    for w in weeks:
        try:
            dates.append(date.fromisoformat(w.monday))
        except ValueError:
            continue
    if not dates:
        return []
    dmin, dmax = min(dates), max(dates)
    start = date(dmin.year, dmin.month, 1)
    end_inclusive = _add_one_month(date(dmax.year, dmax.month, 1))
    out: list[date] = []
    cur = start
    while cur <= end_inclusive:
        out.append(cur)
        cur = _add_one_month(cur)
    return out


def _weeks_in_month(weeks: list, month_start: date) -> list:
    """Weeks whose **Monday** is in this calendar month (first Monday = start of that month’s report).

    Does **not** include the prior month’s straddle week (e.g. 2026-04-27 when building May).
    """
    y, m = month_start.year, month_start.month
    return weeks_with_monday_in_calendar_month(weeks, y, m)


def _regenerate_binge_for_month(
    cfg_path: Path, schedule_anchor: Any
) -> tuple[bool, list[str], Optional[tuple[Path, Path, Path]]]:
    """After a grid swap, rebuild BINGE + GRIDS for the calendar month that contains the anchor date."""
    parsed = parse_schedule_anchor(schedule_anchor)
    if not parsed:
        return False, [], None
    d, _ = parsed
    month_start = date(d.year, d.month, 1)
    cfg = load_build_config(cfg_path)
    weeks = _weeks_in_month(cfg.weeks, month_start)
    if not weeks:
        return (
            False,
            [
                f"No **weeks:** for **{month_start.strftime('%B %Y')}** — automatic BINGE export was skipped."
            ],
            None,
        )
    nikki_path = resolved_nikki_workbook_path(cfg)
    if not nikki_path.is_file():
        return (
            False,
            [f"Nikki workbook not found — automatic export skipped: `{nikki_path}`"],
            None,
        )
    missing_grids: list[str] = []
    for w in weeks:
        if not Path(w.grids_file).is_file():
            missing_grids.append(str(Path(w.grids_file)))
    if missing_grids:
        return (
            False,
            ["Grids file missing — automatic export skipped:"]
            + [f"- `{p}`" for p in missing_grids],
            None,
        )
    try:
        created = ensure_grids_workbooks_for_weeks(weeks)
    except (OSError, ValueError) as e:
        return False, [f"Could not prepare grids: {e}"], None

    out_dir = Path(tempfile.mkdtemp(prefix="binge_after_swap_"))
    station_kw: Optional[List[str]] = None
    if cfg.export_stations:
        station_kw = list(cfg.export_stations)
    try:
        binge_path, grids_path, ovw, seeded = export_both(
            cfg, out_dir, weeks=weeks, export_stations=station_kw
        )
    except Exception as e:
        return False, [f"**Create Schedule** failed after swap: {e}"], None

    msgs: list[str] = [
        f"Automatically regenerated **BINGE.xlsx** and **BINGE GRIDS.xlsx** for **{month_start.strftime('%B %Y')}** "
        f"({len(weeks)} week tab(s)) — use the downloads below."
    ]
    if created:
        msgs.append("Created grids shell(s): " + ", ".join(f"`{p}`" for p in created))
    for w in ovw:
        msgs.append(str(w))
    for s in seeded:
        if is_verbose_seed_noise(s):
            continue
        msgs.append(s)
    return True, msgs, (binge_path, grids_path, out_dir)


@lru_cache(maxsize=1)
def _streamlit_container_supports_border() -> bool:
    return "border" in inspect.signature(st.container).parameters


@lru_cache(maxsize=1)
def _dataframe_row_selection_supported() -> bool:
    sig = inspect.signature(st.dataframe)
    return "on_select" in sig.parameters and "selection_mode" in sig.parameters


def _archive_detail_panel():
    if _streamlit_container_supports_border():
        return st.container(border=True)
    return st.container()


def _nikki_mtime(path: Path) -> float:
    try:
        return path.stat().st_mtime
    except OSError:
        return -1.0


def _archive_wkey(sel: str) -> str:
    return hashlib.sha256(sel.encode("utf-8")).hexdigest()[:26]


def _mobile_styles() -> None:
    """Inject once: top-nav layout, stack-friendly columns, tap targets."""
    st.markdown(
        """
        <style>
        /* Top navigation: hide sidebar — nav is the segmented control below the title */
        [data-testid="stSidebar"] {
            display: none !important;
        }
        [data-testid="collapsedControl"] {
            display: none !important;
        }
        section[data-testid="stMain"] > div {
            margin-left: 0 !important;
        }
        .main .block-container {
            padding-top: 0.75rem !important;
            max-width: 42rem !important;
        }
        @media (max-width: 768px) {
            .block-container {
                padding-left: max(12px, env(safe-area-inset-left)) !important;
                padding-right: max(12px, env(safe-area-inset-right)) !important;
            }
            [data-testid="stHorizontalBlock"] {
                flex-wrap: wrap !important;
                gap: 0.5rem !important;
            }
            [data-testid="stHorizontalBlock"] > [data-testid="column"] {
                flex: 1 1 100% !important;
                min-width: unset !important;
                width: 100% !important;
            }
        }
        button[kind="primary"], button[kind="secondary"], .stDownloadButton button {
            min-height: 2.75rem;
        }
        /* Primary actions + downloads: green (was default Streamlit red primary) */
        button[kind="primary"],
        .stDownloadButton button,
        div[data-testid="stDownloadButton"] button {
            background-color: #16a34e !important;
            background-image: none !important;
            border-color: #15803d !important;
            color: #ffffff !important;
        }
        button[kind="primary"]:hover,
        .stDownloadButton button:hover,
        div[data-testid="stDownloadButton"] button:hover {
            background-color: #15803d !important;
            border-color: #166534 !important;
            color: #ffffff !important;
        }
        button[kind="primary"]:focus-visible,
        .stDownloadButton button:focus-visible,
        div[data-testid="stDownloadButton"] button:focus-visible {
            box-shadow: 0 0 0 2px rgba(34, 197, 94, 0.5) !important;
        }
        /* Segmented nav: Streamlit uses stButtonGroup (radio + aria-checked), not stSegmentedControl */
        div[data-testid="stButtonGroup"],
        div[data-testid="stSegmentedControl"] {
            width: 100%;
        }
        div[data-testid="stButtonGroup"] button,
        div[data-testid="stSegmentedControl"] button {
            border-style: solid !important;
            border-radius: 12px !important;
            transition: border-width 0.12s ease, border-color 0.12s ease, background-color 0.12s ease !important;
        }
        div[data-testid="stButtonGroup"] button[aria-checked="false"],
        div[data-testid="stButtonGroup"] button[aria-pressed="false"],
        div[data-testid="stSegmentedControl"] button[aria-pressed="false"] {
            border-width: 1px !important;
            border-color: rgba(255, 255, 255, 0.22) !important;
            background-color: transparent !important;
            background-image: none !important;
        }
        /* Selected segment: green (default “primary” / checked state is red) */
        div[data-testid="stButtonGroup"] button[aria-checked="true"],
        div[data-testid="stButtonGroup"] button[aria-pressed="true"],
        div[data-testid="stButtonGroup"] button[aria-selected="true"],
        div[data-testid="stSegmentedControl"] button[aria-pressed="true"],
        div[data-testid="stSegmentedControl"] button[aria-checked="true"],
        div[data-testid="stSegmentedControl"] button[aria-selected="true"],
        div[data-testid="stSegmentedControl"] [role="option"][aria-selected="true"] {
            border-width: 3px !important;
            border-color: #22c55e !important;
            background-color: #15803d !important;
            background-image: none !important;
            color: #ffffff !important;
        }
        div[data-testid="stButtonGroup"] button[aria-checked="true"] *,
        div[data-testid="stButtonGroup"] button[aria-pressed="true"] *,
        div[data-testid="stSegmentedControl"] button[aria-pressed="true"] *,
        div[data-testid="stSegmentedControl"] button[aria-checked="true"] * {
            color: #ffffff !important;
        }
        div[data-testid="stButtonGroup"] button[aria-checked="true"]:hover,
        div[data-testid="stButtonGroup"] button[aria-pressed="true"]:hover,
        div[data-testid="stSegmentedControl"] button[aria-pressed="true"]:hover,
        div[data-testid="stSegmentedControl"] button[aria-checked="true"]:hover {
            background-color: #166534 !important;
            border-color: #4ade80 !important;
            color: #ffffff !important;
        }
        /* Horizontal radio fallback: thick border on chosen option */
        div[data-testid="stRadio"] > div {
            gap: 0.5rem !important;
            flex-wrap: wrap !important;
        }
        div[data-testid="stRadio"] label {
            border-radius: 12px !important;
            padding: 0.45rem 0.75rem !important;
            border: 1px solid rgba(255, 255, 255, 0.22) !important;
        }
        div[data-testid="stRadio"] label:has(input:checked) {
            border-width: 3px !important;
            border-color: #22c55e !important;
            background-color: rgba(34, 197, 94, 0.18) !important;
        }
        /* Hide Streamlit Cloud / dev “Deploy” link in the top-right toolbar */
        div[data-testid="stToolbarDeployButton"],
        .stDeployButton,
        .stAppDeployButton {
            display: none !important;
            visibility: hidden !important;
            width: 0 !important;
            height: 0 !important;
            overflow: hidden !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


_NAV_BUILD = "Create Schedule"
_NAV_ARCHIVE = "Available Content"
_NAV_EDIT_SCHEDULE = "Edit schedules"
_MAIN_NAV_OPTIONS = (_NAV_BUILD, _NAV_ARCHIVE, _NAV_EDIT_SCHEDULE)
_LEGACY_MAIN_NAV_TAB: dict[str, str] = {
    "Build": _NAV_BUILD,
    "Build playlist": _NAV_BUILD,
    "Content archive": _NAV_ARCHIVE,
    "View content archive": _NAV_ARCHIVE,
    "Available Content": _NAV_ARCHIVE,
    "Playlist": _NAV_EDIT_SCHEDULE,
    "Edit playlist": _NAV_EDIT_SCHEDULE,
    "Edit playlists": _NAV_EDIT_SCHEDULE,
}
# Must not assign to ``main_nav_tabs`` after the segmented control renders — use ``main_nav_pending`` + rerun instead.
_MAIN_NAV_PENDING_KEY = "main_nav_pending"


def _coerce_main_nav_value(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val)
    if s in _MAIN_NAV_OPTIONS:
        return s
    return _LEGACY_MAIN_NAV_TAB.get(s)


def _render_top_nav() -> str:
    """Primary section switcher — top bar (replaces sidebar nav). Returns selected page name."""
    cur = st.session_state.get("main_nav_tabs")
    coerced = _coerce_main_nav_value(cur)
    if cur is not None and coerced is not None:
        st.session_state["main_nav_tabs"] = coerced
    elif cur is not None and coerced is None:
        st.session_state["main_nav_tabs"] = _NAV_BUILD

    pending_raw = st.session_state.pop(_MAIN_NAV_PENDING_KEY, None)
    pending = _coerce_main_nav_value(pending_raw)
    if pending in _MAIN_NAV_OPTIONS:
        st.session_state["main_nav_tabs"] = pending

    if st.session_state.get("main_nav_tabs") not in _MAIN_NAV_OPTIONS:
        st.session_state["main_nav_tabs"] = _NAV_BUILD

    nav_col, setup_col = st.columns([5, 1], vertical_alignment="center")
    with nav_col:
        if hasattr(st, "segmented_control"):
            page = st.segmented_control(
                "Section",
                options=_MAIN_NAV_OPTIONS,
                default=_NAV_BUILD,
                key="main_nav_tabs",
                label_visibility="collapsed",
                width="stretch",
            )
        else:
            page = st.radio(
                "Section",
                _MAIN_NAV_OPTIONS,
                horizontal=True,
                key="main_nav_tabs",
                label_visibility="collapsed",
            )
    with setup_col:
        if "main_setup_yaml" not in st.session_state:
            st.session_state["main_setup_yaml"] = _default_config_display()
        base_records = _base_schedule_records()
        base_choices = [str(r["path"]) for r in base_records]
        record_by_path = {str(r["path"]): r for r in base_records}
        custom_opt = "__custom__"

        def _render_base_schedule_picker() -> None:
            current = str(st.session_state.get("main_setup_yaml", "")).strip()
            opts = list(base_choices) + [custom_opt]
            default_idx = opts.index(current) if current in opts else len(opts) - 1
            pick = st.selectbox(
                "Base schedule version",
                opts,
                index=default_idx,
                key="main_base_schedule_pick",
                format_func=lambda v: (
                    str(record_by_path.get(str(v), {}).get("label"))
                    if str(v) in record_by_path
                    else "Other (advanced)"
                ),
            )
            if "main_setup_yaml_custom" not in st.session_state:
                st.session_state["main_setup_yaml_custom"] = current if current not in base_choices else ""
            rec = record_by_path.get(str(pick))
            if rec is not None:
                st.caption(str(rec.get("detail") or ""))
            if pick == custom_opt:
                custom_val = st.text_input(
                    "Custom base schedule file",
                    key="main_setup_yaml_custom",
                    placeholder="config/may_test.yaml",
                ).strip()
                if custom_val:
                    st.session_state["main_setup_yaml"] = custom_val
            else:
                st.session_state["main_setup_yaml"] = str(pick)

        if hasattr(st, "popover"):
            with st.popover("Base schedule", use_container_width=True):
                _render_base_schedule_picker()
        else:
            _render_base_schedule_picker()
    if page is None:
        return _NAV_BUILD
    return str(page)


@st.cache_data(show_spinner=False)
def _nikki_workbook_sheet_names(workbook_resolved: str, _mtime: float) -> tuple[str, ...]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_resolved, read_only=True)
    try:
        return tuple(wb.sheetnames)
    finally:
        wb.close()


def _nikki_headers_from_json(blob: str) -> NikkiColumnHeaders:
    raw = json.loads(blob)
    if not raw:
        return NikkiColumnHeaders.standard_series()
    return NikkiColumnHeaders(
        episode=raw.get("episode") or "Episode",
        season_episode=raw.get("season_episode"),
        year=raw.get("year"),
        stars=raw.get("stars"),
        synopsis=raw.get("synopsis"),
    )


@st.cache_data(show_spinner="Loading episodes from workbook…")
def _archive_sheet_episodes(
    workbook: str,
    _workbook_mtime: float,
    sheet: str,
    style: str,
    prefix: str,
    row_filter: Optional[str],
    headers_json: str,
) -> list[dict[str, Any]]:
    columns = _nikki_headers_from_json(headers_json)
    eps = nikki.load_sheet(
        workbook,
        sheet,
        style=style,
        prefix=prefix,
        columns=columns,
        row_filter=row_filter,
    )
    return normalize_episodes_for_archive(eps, style)


def _nikki_movies_sheet_name(nikki_path: Path) -> Optional[str]:
    if not nikki_path.is_file():
        return None
    tabs = _nikki_workbook_sheet_names(str(nikki_path.resolve()), _nikki_mtime(nikki_path))
    for t in tabs:
        if str(t).strip().casefold() == "movies":
            return str(t)
    return None


@st.cache_data(show_spinner=False)
def _nikki_movie_catalog_titles(workbook: str, _workbook_mtime: float, movies_sheet: str) -> list[str]:
    hdrs = NikkiColumnHeaders.movies_tab()
    rows = _archive_sheet_episodes(
        workbook,
        _workbook_mtime,
        movies_sheet,
        "movies",
        "MOV",
        None,
        json.dumps(asdict(hdrs), sort_keys=True),
    )
    out: list[str] = []
    seen: set[str] = set()
    for r in rows:
        title = str(r.get("title") or "").strip()
        if not title or title in seen:
            continue
        seen.add(title)
        out.append(title)
    out.sort(key=str.casefold)
    return out


def _nikki_movie_catalog_options(nikki_path: Path) -> list[str]:
    movies_sheet = _nikki_movies_sheet_name(nikki_path)
    if not movies_sheet:
        return []
    titles = _nikki_movie_catalog_titles(
        str(nikki_path.resolve()),
        _nikki_mtime(nikki_path),
        movies_sheet,
    )
    return [_literal_text_option(t) for t in titles]


def _render_archive_episode_browser(
    sel: str,
    sd: ShowDef,
    nikki_path: Path,
    *,
    style: str,
    hdrs: NikkiColumnHeaders,
    sheet_ok: bool | None,
    browse_only: bool = False,
) -> None:
    st.markdown("### Episodes")
    if browse_only:
        st.caption(
            "Browse only — not on the schedule until you add this tab under **`nikki_sheet`** in your base schedule. "
            "**Create Schedule** skips it until then."
        )
    if sd.nikki_row_filter == nikki.ROW_FILTER_GREEN_EPISODE_CELL:
        st.caption(
            "Only **green** Episode cells count (same rule as **Create Schedule**); other rows on this sheet are ignored."
        )
    elif sd.nikki_row_filter:
        st.caption(
            f"Row filter `{sd.nikki_row_filter}` — table matches what **Create Schedule** would load."
        )
    if not nikki_path.is_file():
        st.warning("Spreadsheet file not found — check **nikki_workbook** in your base schedule.")
        return
    if not sd.nikki_sheet:
        st.info("This show has no **nikki_sheet**; there is no Excel tab to list.")
        return
    if sheet_ok is False:
        st.error(
            "Your workbook has no tab with this exact name. Fix **nikki_sheet** in the base schedule or rename the tab in Excel."
        )
        return
    if sheet_ok is None:
        st.warning("Could not verify the tab name against the workbook file.")
    wb_abs = str(nikki_path.resolve())
    mtime = _nikki_mtime(nikki_path)
    headers_json = json.dumps(asdict(hdrs), sort_keys=True)
    try:
        rows = _archive_sheet_episodes(
            wb_abs,
            mtime,
            sd.nikki_sheet,
            style,
            sd.prefix or "",
            sd.nikki_row_filter,
            headers_json,
        )
    except Exception as e:
        st.error(f"Could not read this tab: {e}")
        return
    if not rows:
        if sd.nikki_row_filter == nikki.ROW_FILTER_GREEN_EPISODE_CELL:
            st.warning(
                "**No green Episode cells matched.** **Create Schedule** would also get an empty list for this "
                "show—confirm playable rows use the expected green fill on the Episode column, and that "
                "**nikki_workbook** points at the real file (not a placeholder)."
            )
        else:
            st.info("No episode rows were parsed (check the header row and column titles on this tab).")
        return

    st.caption(
        f"**{len(rows)}** rows — schedule **#** column matches **Create Schedule**"
        + (" (when on the schedule)." if not browse_only else " (browse only until added to base schedule).")
        + " Click a row for detail."
    )

    seasons_found = sorted({r["season"] for r in rows if r["season"] is not None})
    has_unnumbered = any(r["season"] is None for r in rows)
    season_opts = ["All seasons"]
    season_opts.extend(f"Season {s}" for s in seasons_found)
    if has_unnumbered and seasons_found:
        season_opts.append("Unnumbered / list order")

    pick_season = st.selectbox(
        "Season filter",
        season_opts,
        key=f"archive_season_{_archive_wkey(sel)}",
    )
    q = st.text_input(
        "Search",
        "",
        key=f"archive_search_{_archive_wkey(sel)}",
        placeholder="Title, code, or cell text…",
    )

    def season_match(r: dict[str, Any]) -> bool:
        if pick_season == "All seasons":
            return True
        if pick_season.startswith("Season "):
            sn = int(pick_season.replace("Season ", ""))
            return r["season"] == sn
        if pick_season == "Unnumbered / list order":
            return r["season"] is None
        return True

    qn = q.strip().casefold()

    def search_match(r: dict[str, Any]) -> bool:
        if not qn:
            return True
        for fld in ("title", "code", "raw_cell", "sheet_se"):
            if qn in str(r.get(fld, "")).casefold():
                return True
        return False

    filtered = [r for r in rows if season_match(r) and search_match(r)]
    if not filtered:
        st.warning("No episodes match the current season filter and search.")
        return

    disp = pd.DataFrame(
        {
            "#": [r["schedule_num"] for r in filtered],
            "S×E": [r["se_compact"] for r in filtered],
            "Season": [("—" if r["season"] is None else str(r["season"])) for r in filtered],
            "Ep": [("—" if r["ep_in_season"] is None else str(r["ep_in_season"])) for r in filtered],
            "Sheet S/E": [r["sheet_se"] or "—" for r in filtered],
            "Code": [r["code"] for r in filtered],
            "Title": [r["title"] for r in filtered],
        }
    )
    df_key = f"archive_df_{_archive_wkey(sel)}"
    sel_supported = _dataframe_row_selection_supported()

    if sel_supported:
        event = st.dataframe(
            disp,
            column_config={
                "Title": st.column_config.TextColumn("Title", width="large"),
                "#": st.column_config.NumberColumn("#", format="%d", width="small"),
            },
            use_container_width=True,
            height=360,
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row-required",
            key=df_key,
        )
        rows_sel: list[int] = []
        try:
            rows_sel = list(event["selection"]["rows"])  # type: ignore[index]
        except (KeyError, TypeError, AttributeError):
            pass
        picked_idx: Optional[int] = None
        if rows_sel:
            try:
                picked_idx = int(rows_sel[0])
            except (TypeError, ValueError, IndexError):
                picked_idx = None
        st.markdown("**Selected row**")
        if picked_idx is not None and 0 <= picked_idx < len(filtered):
            one = filtered[picked_idx]
            sr1, sr2 = st.columns(2)
            with sr1:
                st.metric("Schedule #", str(one["schedule_num"]))
                st.metric("Code", one["code"] or "—")
            with sr2:
                st.metric("S×E (normalized)", one["se_compact"])
                st.metric("0-based index", str(one["idx0"]))
            st.caption("Normalized **Episode** cell (whitespace collapsed)")
            raw = str(one["raw_cell"])
            st.code(raw if len(raw) <= 800 else raw[:800] + "…", language=None)
        else:
            st.info("Click a row in the table above to see full detail here.")
    else:
        st.dataframe(disp, use_container_width=True, height=360, hide_index=True)
        st.warning(
            "Your Streamlit is too old for **clickable table rows**. Upgrade: "
            "`pip install -U \"streamlit>=1.35\"`, then restart the app."
        )
        ix = st.selectbox(
            "Pick a row (fallback)",
            list(range(len(filtered))),
            format_func=lambda i: (
                f"#{filtered[i]['schedule_num']}  {filtered[i]['se_compact']}  {filtered[i]['code']}  —  "
                f"{str(filtered[i]['title'])[:160]}"
            ),
            key=f"archive_jump_{_archive_wkey(sel)}",
        )
        one = filtered[int(ix)]
        st.metric("Schedule #", str(one["schedule_num"]))
        st.metric("S×E", one["se_compact"])
        raw = str(one["raw_cell"])
        st.caption("Normalized **Episode** cell")
        st.code(raw if len(raw) <= 800 else raw[:800] + "…", language=None)


def _month_key(m: date) -> str:
    return f"{m.year:04d}-{m.month:02d}"


def _build_state_path(cfg_path: Path) -> Path:
    return cfg_path.resolve().parent / "schedule_build_state.json"


def _legacy_build_state_path(cfg_path: Path) -> Path:
    return cfg_path.resolve().parent / "playlist_build_state.json"


def _completed_months_from_file(path: Path, config_resolved: str) -> set[str]:
    if not path.is_file():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return set()
    if data.get("config_resolved") != config_resolved:
        return set()
    cm = data.get("completed_months")
    if not isinstance(cm, list):
        return set()
    return {str(x) for x in cm if x}


def _load_completed_months(cfg_path: Path) -> set[str]:
    resolved = str(cfg_path.resolve())
    return _completed_months_from_file(_build_state_path(cfg_path), resolved) | _completed_months_from_file(
        _legacy_build_state_path(cfg_path), resolved
    )


def _record_completed_month(cfg_path: Path, month_start: date) -> None:
    p = _build_state_path(cfg_path)
    resolved = str(cfg_path.resolve())
    prev = _load_completed_months(cfg_path)
    prev.add(_month_key(month_start))
    out = {
        "version": 1,
        "config_resolved": resolved,
        "completed_months": sorted(prev),
    }
    p.write_text(json.dumps(out, indent=2), encoding="utf-8")


def _parse_sequence_start(raw: Optional[str]) -> Optional[date]:
    if not raw or not str(raw).strip():
        return None
    s = str(raw).strip()
    try:
        d = date.fromisoformat(s[:10])
    except ValueError:
        return None
    return date(d.year, d.month, 1)


def _pipeline_months(months_all: list[date], build_sequence_start: Optional[str]) -> list[date]:
    """Months in the in-app unlock chain (subset of ``months_all``), in calendar order."""
    if not months_all:
        return []
    start = _parse_sequence_start(build_sequence_start)
    if start is None:
        return list(months_all)
    sm = (start.year, start.month)
    return [m for m in months_all if (m.year, m.month) >= sm]


def _unlocked_months(pipeline: list[date], completed: set[str]) -> list[date]:
    """Sequential unlock: first month always; each next month appears after the previous is marked complete."""
    if not pipeline:
        return []
    out: list[date] = []
    for i, m in enumerate(pipeline):
        if i == 0:
            out.append(m)
            continue
        prev = pipeline[i - 1]
        if _month_key(prev) in completed:
            out.append(m)
        else:
            break
    return out


def _sorted_weeks(weeks: list) -> list:
    return sorted(weeks, key=lambda w: parse_monday(w.monday))


def _monday_on_or_after(day: date) -> date:
    """First calendar Monday on or after ``day`` (weeks anchor on Monday rows)."""
    wd = day.weekday()
    if wd == 0:
        return day
    return day + timedelta(days=7 - wd)


def _weeks_for_unlocked_months(weeks: list, unlocked_months: list[date]) -> list:
    mm = {(d.year, d.month) for d in unlocked_months}
    return [w for w in _sorted_weeks(weeks) if (parse_monday(w.monday).year, parse_monday(w.monday).month) in mm]


def _effective_weeks_from_start(all_weeks: list, start_on: date, count: int) -> list:
    if not all_weeks:
        return []
    sorted_weeks = _sorted_weeks(all_weeks)
    anchor = _monday_on_or_after(start_on)
    start_idx: Optional[int] = None
    for i, w in enumerate(sorted_weeks):
        mon = parse_monday(w.monday)
        if mon >= anchor:
            start_idx = i
            break
    if start_idx is None:
        for i, w in enumerate(sorted_weeks):
            mon = parse_monday(w.monday)
            if mon <= start_on < mon + timedelta(days=7):
                start_idx = i
                break
    if start_idx is None:
        return []
    tail = sorted_weeks[start_idx:]
    return tail[: max(1, int(count))]


def _next_week_start_after_selection(selected_weeks: list, buildable_weeks: list) -> Optional[date]:
    """Next buildable Monday after the current selected window."""
    if not selected_weeks or not buildable_weeks:
        return None
    last_selected = max(parse_monday(w.monday) for w in selected_weeks)
    for w in _sorted_weeks(buildable_weeks):
        mon = parse_monday(w.monday)
        if mon > last_selected:
            return mon
    return None


def _week_floor_from_reference_cutoff(cfg) -> Optional[date]:
    raw = str(getattr(cfg, "reference_binge_literal_copy_before", "") or "").strip()
    if not raw:
        return None
    try:
        d = date.fromisoformat(raw[:10])
    except ValueError:
        return None
    return d


def _format_duration_minutes(minutes: int) -> str:
    h, m = divmod(int(minutes), 60)
    if h and m:
        return f"{h}h {m}m"
    if h:
        return f"{h}h"
    return f"{m}m"


def _clock_label_from_minutes(total_minutes: float) -> str:
    mins = int(round(float(total_minutes))) % (24 * 60)
    hh = mins // 60
    mm = mins % 60
    ampm = "AM" if hh < 12 else "PM"
    hh12 = hh % 12
    if hh12 == 0:
        hh12 = 12
    return f"{hh12}:{mm:02d} {ampm}"


def _runtime_timing_notes_for_day(
    *,
    cfg,
    day_iso: str,
    rows_for_day: list[dict[str, Any]],
    assigned_by_slot: dict[str, str],
    runtime_map: dict[str, int],
    fallback_runtime: int,
    commercials_pct: float,
) -> list[str]:
    if not rows_for_day:
        return []
    rows = sorted(rows_for_day, key=lambda r: int(r["start_slot"]))
    for i in range(1, len(rows)):
        if int(rows[i]["start_slot"]) != int(rows[i - 1]["end_slot"]):
            return []
    picks: list[tuple[dict[str, Any], str]] = []
    for r in rows:
        sid = str(r["slot_id"])
        opt = assigned_by_slot.get(sid)
        if not opt:
            return []
        picks.append((r, opt))
    runs: list[tuple[str, int, int]] = []
    for r, opt in picks:
        st_slot = int(r["start_slot"])
        end_slot = int(r["end_slot"])
        if not runs or runs[-1][0] != opt:
            runs.append((opt, st_slot, end_slot))
        else:
            prev_opt, prev_start, _prev_end = runs[-1]
            runs[-1] = (prev_opt, prev_start, end_slot)
    d = date.fromisoformat(day_iso)
    day_label = f"{d.strftime('%A')} {d.month}/{d.day}"
    cur_mins = int(rows[0]["start_slot"]) * 30
    window_end_mins = int(rows[-1]["end_slot"]) * 30
    notes: list[str] = []
    for idx, (opt, run_start_slot, _run_end_slot) in enumerate(runs):
        title = _display_name_for_archive_pick(cfg, opt)
        runtime = _runtime_for_archive_option(cfg, opt, runtime_map)
        if runtime is None:
            runtime = int(fallback_runtime)
        sched_start_mins = run_start_slot * 30
        if idx > 0 and abs(cur_mins - sched_start_mins) >= 1:
            notes.append(
                f"Please note that on {day_label} {title} will start at {_clock_label_from_minutes(cur_mins)}."
            )
        airtime = float(runtime) * (1.0 + float(commercials_pct) / 100.0)
        cur_mins += airtime
    drift = int(round(cur_mins - window_end_mins))
    if drift > 0:
        notes.append(
            f"Runtime fit on {day_label}: estimated over by {drift} minute(s) "
            "(the sequence may end later within the selected window)."
        )
    elif drift < 0:
        notes.append(
            f"Runtime fit on {day_label}: estimated under by {abs(drift)} minute(s) "
            "(extra filler/commercial time may be needed)."
        )
    return notes


def _schedule_template_slots(weeks: list) -> tuple[list[dict[str, Any]], list[str]]:
    slots: list[dict[str, Any]] = []
    warnings: list[str] = []
    for w in _sorted_weeks(weeks):
        mon = parse_monday(w.monday)
        try:
            grid = load_grid_sheet(w.grids_file, w.sheet_name)
        except Exception as e:  # noqa: BLE001
            warnings.append(f"Could not load grid `{w.sheet_name}` in `{Path(w.grids_file).name}`: {e}")
            continue
        dates = day_dates(mon)
        for day_idx in range(7):
            col = [grid[r][day_idx] for r in range(48)]
            try:
                segs = segments_for_day(col)
            except ValueError as e:
                warnings.append(f"Bad grid column in `{w.sheet_name}` (day {day_idx}): {e}")
                continue
            for seg in segs:
                minutes = int(seg.end_slot - seg.start_slot) * 30
                d = dates[day_idx]
                st = slot_label(seg.start_slot)
                fin = slot_label(seg.end_slot % 48)
                show_text = str(seg.cell_text).strip()
                slot_id = f"{d.isoformat()}|{st}|{day_idx}|{seg.start_slot}|{w.monday}"
                slots.append(
                    {
                        "slot_id": slot_id,
                        "date": d,
                        "date_iso": d.isoformat(),
                        "week_monday": w.monday,
                        "day_index": day_idx,
                        "start_slot": int(seg.start_slot),
                        "end_slot": int(seg.end_slot),
                        "start": st,
                        "finish": fin,
                        "duration_minutes": minutes,
                        "duration_label": _format_duration_minutes(minutes),
                        "show": show_text,
                    }
                )
    slots.sort(key=lambda r: (r["date_iso"], r["start_slot"], r["show"].casefold()))
    return slots, warnings


def _slot_picker_label(row: dict[str, Any]) -> str:
    return (
        f"{row['date_iso']} {row['start']}-{row['finish']} ({row['duration_label']}) · "
        f"{row['show']} · week {row['week_monday']}"
    )


def _monday_for_calendar_date(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _map_output_grid_tabs_by_monday(grids_path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(grids_path, read_only=True)
        try:
            for sn in wb.sheetnames:
                md = parse_sheet_tab_monday(sn)
                if md is not None:
                    out[md.isoformat()] = sn
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return out
    return out


def _write_output_grid_merged_run(ws: Any, col: int, row_start: int, row_end: int, new_display: str) -> None:
    """Write one replacement run and only adjust merges that overlap this run."""
    if row_end < row_start:
        return
    for mr in list(ws.merged_cells.ranges):
        if mr.max_col != mr.min_col or mr.min_col != col:
            continue
        if mr.max_row < row_start or mr.min_row > row_end:
            continue
        top_val = ws.cell(row=mr.min_row, column=col).value
        old_start, old_end = int(mr.min_row), int(mr.max_row)
        ws.unmerge_cells(str(mr))
        # Preserve upper untouched section of the old merged block.
        if old_start < row_start:
            ws.cell(row=old_start, column=col, value=top_val)
            for rr in range(old_start + 1, row_start):
                ws.cell(row=rr, column=col, value=None)
            if row_start - 1 > old_start:
                ws.merge_cells(start_row=old_start, start_column=col, end_row=row_start - 1, end_column=col)
        # Preserve lower untouched section of the old merged block.
        if old_end > row_end:
            low_start = row_end + 1
            ws.cell(row=low_start, column=col, value=top_val)
            for rr in range(low_start + 1, old_end + 1):
                ws.cell(row=rr, column=col, value=None)
            if old_end > low_start:
                ws.merge_cells(start_row=low_start, start_column=col, end_row=old_end, end_column=col)

    ws.cell(row=row_start, column=col, value=new_display)
    for rr in range(row_start + 1, row_end + 1):
        ws.cell(row=rr, column=col, value=None)
    if row_end > row_start:
        ws.merge_cells(start_row=row_start, start_column=col, end_row=row_end, end_column=col)


def _apply_output_grid_slot_replacements(
    grids_path: Path,
    slot_rows: list[dict[str, Any]],
    new_display: str,
) -> list[str]:
    msgs: list[str] = []
    if not grids_path.is_file():
        return [f"OTO grids update skipped: output file missing `{grids_path}`."]
    tab_by_monday = _map_output_grid_tabs_by_monday(grids_path)
    if not tab_by_monday:
        return [f"OTO grids update skipped: could not map week tabs in `{grids_path.name}`."]
    try:
        import openpyxl

        wb = openpyxl.load_workbook(grids_path, read_only=False, data_only=False)
    except OSError as e:
        return [f"OTO grids update skipped: could not open `{grids_path}` ({e})."]
    changed = 0
    runs: list[tuple[str, int, int, int, str]] = []
    try:
        for r in slot_rows:
            d = r["date"]
            mon_key = _monday_for_calendar_date(d).isoformat()
            tab = tab_by_monday.get(mon_key)
            if not tab or tab not in wb.sheetnames:
                continue
            col = 2 + int(r["day_index"])
            row_start = 5 + int(r["start_slot"])
            row_end = 5 + int(r["end_slot"]) - 1
            runs.append((tab, col, row_start, row_end, new_display))
            changed += max(0, row_end - row_start + 1)
        runs.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
        merged_runs: list[tuple[str, int, int, int, str]] = []
        for run in runs:
            if not merged_runs:
                merged_runs.append(run)
                continue
            p_tab, p_col, p_start, p_end, p_disp = merged_runs[-1]
            c_tab, c_col, c_start, c_end, c_disp = run
            if c_tab == p_tab and c_col == p_col and c_disp == p_disp and c_start <= p_end + 1:
                merged_runs[-1] = (p_tab, p_col, p_start, max(p_end, c_end), p_disp)
            else:
                merged_runs.append(run)
        for tab, col, row_start, row_end, disp in merged_runs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            _write_output_grid_merged_run(ws, col, row_start, row_end, disp)
        wb.save(grids_path)
    finally:
        wb.close()
    msgs.append(f"OTO grids update: replaced {changed} output cell(s) with `{new_display}` in `{grids_path.name}`.")
    return msgs


def _apply_output_grid_slot_replacements_multi(
    grids_path: Path,
    slot_rows: list[dict[str, Any]],
    new_displays: list[str],
) -> list[str]:
    msgs: list[str] = []
    if not grids_path.is_file():
        return [f"OTO grids update skipped: output file missing `{grids_path}`."]
    tab_by_monday = _map_output_grid_tabs_by_monday(grids_path)
    if not tab_by_monday:
        return [f"OTO grids update skipped: could not map week tabs in `{grids_path.name}`."]
    try:
        import openpyxl

        wb = openpyxl.load_workbook(grids_path, read_only=False, data_only=False)
    except OSError as e:
        return [f"OTO grids update skipped: could not open `{grids_path}` ({e})."]
    changed = 0
    runs: list[tuple[str, int, int, int, str]] = []
    try:
        for i, r in enumerate(slot_rows):
            if i >= len(new_displays):
                break
            new_display = str(new_displays[i]).strip()
            if not new_display:
                continue
            d = r["date"]
            mon_key = _monday_for_calendar_date(d).isoformat()
            tab = tab_by_monday.get(mon_key)
            if not tab or tab not in wb.sheetnames:
                continue
            col = 2 + int(r["day_index"])
            row_start = 5 + int(r["start_slot"])
            row_end = 5 + int(r["end_slot"]) - 1
            runs.append((tab, col, row_start, row_end, new_display))
            changed += max(0, row_end - row_start + 1)
        runs.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
        merged_runs: list[tuple[str, int, int, int, str]] = []
        for run in runs:
            if not merged_runs:
                merged_runs.append(run)
                continue
            p_tab, p_col, p_start, p_end, p_disp = merged_runs[-1]
            c_tab, c_col, c_start, c_end, c_disp = run
            if c_tab == p_tab and c_col == p_col and c_disp == p_disp and c_start <= p_end + 1:
                merged_runs[-1] = (p_tab, p_col, p_start, max(p_end, c_end), p_disp)
            else:
                merged_runs.append(run)
        for tab, col, row_start, row_end, disp in merged_runs:
            if tab not in wb.sheetnames:
                continue
            ws = wb[tab]
            _write_output_grid_merged_run(ws, col, row_start, row_end, disp)
        wb.save(grids_path)
    finally:
        wb.close()
    msgs.append(f"OTO grids update: replaced {changed} output cell(s) in `{grids_path.name}`.")
    return msgs


def _list_xlsx_sheet_names(path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _list_binge_data_sheets(path: Path) -> list[str]:
    return [s for s in _list_xlsx_sheet_names(path) if s != "BINGE notes"]


def _list_grids_data_sheets(path: Path) -> list[str]:
    return [s for s in _list_xlsx_sheet_names(path) if s != "BINGE notes"]


def _find_binge_column_ci(df: pd.DataFrame, name: str) -> Optional[str]:
    """Return actual column name whose stripped upper case equals ``name``."""
    nu = name.strip().upper()
    for c in df.columns:
        if str(c).strip().upper() == nu:
            return str(c)
    return None


def _schedule_anchor_dict_from_binge_row(df: pd.DataFrame, idx: int) -> Optional[dict[str, Any]]:
    """DATE + START TIME from a BINGE row for single-slot grid edits."""
    dc = _find_binge_column_ci(df, "DATE")
    sc = _find_binge_column_ci(df, "START TIME")
    if not dc or not sc:
        return None
    return {"date": df.iloc[int(idx)][dc], "start": df.iloc[int(idx)][sc]}


def _binge_row_swap_summary(df: pd.DataFrame, idx: int) -> str:
    """One-line label for fallback row picker."""
    r = df.iloc[int(idx)]

    def _cell(col: Optional[str]) -> str:
        if not col or col not in r.index:
            return "?"
        v = r[col]
        if pd.isna(v):
            return "?"
        s = str(v).strip()
        return s if s else "?"

    sc = _find_binge_column_ci(df, "SHOW")
    stc = _find_binge_column_ci(df, "START TIME")
    ec = _find_binge_column_ci(df, "EPISODE")
    return f"Row {idx + 1}: {_cell(sc)} · {_cell(ec)} · {_cell(stc)}"


def _display_name_for_archive_pick(cfg, sel: str) -> str:
    imported = _parse_imported_content_option(sel)
    if imported is not None:
        return imported
    raw_literal = _parse_literal_text_option(sel)
    if raw_literal is not None:
        return raw_literal
    tab = parse_workbook_tab_option(sel)
    if tab is not None:
        return synthetic_series_for_tab(tab).display_name
    return cfg.shows[sel].display_name


def _showdef_for_archive_pick(cfg, sel: str) -> Optional[ShowDef]:
    imported = _parse_imported_content_option(sel)
    if imported is not None:
        return ShowDef(key="literal", display_name=imported, kind="literal")
    raw_literal = _parse_literal_text_option(sel)
    if raw_literal is not None:
        return ShowDef(key="literal", display_name=raw_literal, kind="literal")
    tab = parse_workbook_tab_option(sel)
    if tab is not None:
        return synthetic_series_for_tab(tab)
    return cfg.shows.get(sel)


def _literal_text_option(text: str) -> str:
    return f"{_RAW_LITERAL_PREFIX}{str(text).strip()}"


def _parse_literal_text_option(opt: str) -> Optional[str]:
    if not isinstance(opt, str):
        return None
    if opt.startswith(_RAW_LITERAL_PREFIX):
        raw = opt[len(_RAW_LITERAL_PREFIX) :].strip()
        return raw or None
    return None


def _imported_content_option(text: str) -> str:
    return f"{_IMPORTED_CONTENT_PREFIX}{str(text).strip()}"


def _parse_imported_content_option(opt: str) -> Optional[str]:
    if not isinstance(opt, str):
        return None
    if opt.startswith(_IMPORTED_CONTENT_PREFIX):
        raw = opt[len(_IMPORTED_CONTENT_PREFIX) :].strip()
        return raw or None
    return None


def _imported_catalog_path(cfg_path: Path) -> Path:
    cfg_dir = cfg_path.resolve().parent
    repo_root = cfg_dir.parent if cfg_dir.name.casefold() == "config" else cfg_dir
    return repo_root / "config" / "imported_content_catalog.json"


def _load_imported_catalog_rows(cfg_path: Path) -> list[dict[str, Any]]:
    p = _imported_catalog_path(cfg_path)
    if not p.is_file():
        return []
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []
    if isinstance(raw, dict):
        rows = raw.get("rows", [])
        if isinstance(rows, list):
            return _dedupe_import_rows([r for r in rows if isinstance(r, dict)])
    if isinstance(raw, list):
        return _dedupe_import_rows([r for r in raw if isinstance(r, dict)])
    return []


def _save_imported_catalog_rows(cfg_path: Path, rows: list[dict[str, Any]]) -> None:
    p = _imported_catalog_path(cfg_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    payload = {"rows": rows}
    p.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _normalize_key(text: Any) -> str:
    v = " ".join(str(text or "").strip().lower().split())
    if v in {"nan", "none", "null", "nat"}:
        return ""
    return v


def _normalize_episode_number(value: Any) -> str:
    raw = _normalize_key(value)
    if not raw:
        return ""
    if re.fullmatch(r"\d+", raw):
        try:
            return str(int(raw))
        except Exception:
            return raw
    return raw


def _import_row_identity_key(r: dict[str, Any]) -> str:
    kind = _normalize_key(r.get("content_type", ""))
    series_title = _normalize_key(r.get("series_title", ""))
    display_name = _normalize_key(r.get("display_name", ""))
    ep_num = _normalize_episode_number(r.get("episode_number", ""))
    ep_title = _normalize_key(r.get("episode_title", ""))
    if kind == "series":
        base = series_title or display_name
        episode_token = ep_num or ep_title
        return f"{kind}|{base}|{episode_token}"
    return f"{kind or 'movie'}|{display_name or series_title}"


def _dedupe_import_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # Keep latest row for each identity key so corrected re-imports replace older rows.
    out: list[dict[str, Any]] = []
    seen_idx: dict[str, int] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = _import_row_identity_key(row)
        if key in seen_idx:
            out[seen_idx[key]] = row
        else:
            seen_idx[key] = len(out)
            out.append(row)
    return out


def _import_aliases() -> dict[str, set[str]]:
    return {
        "series_title": {
            "series title",
            "series",
            "artist/series",
            "show",
            "program",
        },
        "title": {
            "title",
            "episode",
            "episode title",
            "title (internal)",
            "sort title",
        },
        "episode_number": {
            "episode number",
            "season/episode",
            "season_episode",
            "ep #",
            "episode #",
        },
        "synopsis_short": {
            "episode short synopsis",
            "short description",
            "series short synopsis (150 characters)",
            "synopsis short",
        },
        "synopsis_long": {
            "episode long synopsis",
            "description",
            "synopsis",
            "series long synopsis (250 characters)",
            "synopsis long",
        },
        "original_airdate": {
            "original airdate",
            "year/original airdate",
            "air date",
            "production year",
            "year",
        },
        "runtime": {
            "runtime",
            "trt",
            "rt",
            "duration",
        },
        "production_company": {
            "production company",
            "studios",
            "studio",
            "producer(s)",
        },
        "genre": {
            "genre",
            "amazon channels genre",
            "roku genre tags",
        },
        "copyright": {
            "copyright",
        },
    }


def _normalize_import_dataframe(df_raw: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    aliases = _import_aliases()
    alias_flat = {a for vals in aliases.values() for a in vals}
    best_row = 0
    best_score = -1
    for hdr in (0, 1, 2):
        if hdr >= len(df_raw):
            continue
        vals = [_normalize_key(v) for v in list(df_raw.iloc[hdr].values)]
        score = sum(1 for v in vals if v in alias_flat)
        if score > best_score:
            best_score = score
            best_row = hdr
    header_vals = [str(v).strip() if pd.notna(v) else "" for v in list(df_raw.iloc[best_row].values)]
    header_vals = [h if h else f"col_{i}" for i, h in enumerate(header_vals)]
    data = df_raw.iloc[best_row + 1 :].copy()
    data.columns = header_vals
    data = data.dropna(how="all")
    return data


def _runtime_minutes_from_cell(v: Any) -> Optional[int]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        fv = float(v)
        # Excel duration cells can come through as day fractions.
        if 0 < fv < 1:
            return max(1, int(round(fv * 24 * 60)))
        return max(1, int(round(fv)))
    if hasattr(v, "total_seconds"):
        try:
            return max(1, int(round(float(v.total_seconds()) / 60.0)))
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            nums = [int(float(x)) for x in parts]
        except Exception:
            return None
        if len(nums) == 3:
            return max(1, int(round(nums[0] * 60 + nums[1] + nums[2] / 60)))
        if len(nums) == 2:
            a, b = nums
            # Heuristic:
            # - 10:00+ is almost always mm:ss in metadata forms.
            # - 0:00..5:59 is often hh:mm for long-form content.
            if a >= 10:
                return max(1, int(round(a + b / 60)))
            return max(1, int(round(a * 60 + b)))
    try:
        return max(1, int(round(float(s))))
    except Exception:
        return None


def _import_rows_from_dataframe(df: pd.DataFrame, sheet_name: str, source_name: str) -> list[dict[str, Any]]:
    def _clean_text(v: Any) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        t = str(v).strip()
        return "" if t.lower() == "nan" else t

    aliases = _import_aliases()
    col_map: dict[str, str] = {}
    for c in df.columns:
        nk = _normalize_key(c)
        for canon, vals in aliases.items():
            if nk in vals and canon not in col_map:
                col_map[canon] = str(c)
    out: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        series_title = _clean_text(r.get(col_map.get("series_title", ""), ""))
        title = _clean_text(r.get(col_map.get("title", ""), ""))
        ep_num = _clean_text(r.get(col_map.get("episode_number", ""), ""))
        if not series_title and ep_num:
            series_title = sheet_name.strip()
        is_series = bool(series_title and (ep_num or title))
        display = title if not is_series else (series_title or title)
        if not display:
            continue
        rt = _runtime_minutes_from_cell(r.get(col_map.get("runtime", ""), None))
        air_raw = r.get(col_map.get("original_airdate", ""), None)
        air_iso = ""
        try:
            if pd.notna(air_raw):
                air_iso = pd.to_datetime(air_raw).date().isoformat()
        except Exception:
            air_iso = str(air_raw or "").strip()
        row = {
            "content_type": "series" if is_series else "movie",
            "display_name": display,
            "series_title": series_title,
            "episode_number": ep_num,
            "episode_title": title if is_series else "",
            "genre": _clean_text(r.get(col_map.get("genre", ""), "")).split(",")[0].strip().lower(),
            "runtime_minutes": int(rt) if rt is not None else None,
            "original_airdate": air_iso,
            "production_company": _clean_text(r.get(col_map.get("production_company", ""), "")),
            "copyright": _clean_text(r.get(col_map.get("copyright", ""), "")),
            "synopsis_short": _clean_text(r.get(col_map.get("synopsis_short", ""), "")),
            "synopsis_long": _clean_text(r.get(col_map.get("synopsis_long", ""), "")),
            "source_sheet": sheet_name,
            "source_file": source_name,
        }
        out.append(row)
    return out


def _parse_uploaded_content_file(name: str, payload: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    notes: list[str] = []
    lname = str(name or "").lower()
    if lname.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(payload))
        rows.extend(_import_rows_from_dataframe(df, "CSV", name))
        return rows, notes
    xls = pd.ExcelFile(io.BytesIO(payload))
    for sn in xls.sheet_names:
        try:
            raw = pd.read_excel(io.BytesIO(payload), sheet_name=sn, header=None)
            norm = _normalize_import_dataframe(raw, sn)
            got = _import_rows_from_dataframe(norm, sn, name)
            rows.extend(got)
            notes.append(f"{sn}: {len(got)} row(s) parsed")
        except Exception as e:
            notes.append(f"{sn}: skipped ({e})")
    return rows, notes


def _merge_import_rows(existing: list[dict[str, Any]], incoming: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = _dedupe_import_rows(list(existing))
    seen_idx: dict[str, int] = {}
    for i, r in enumerate(out):
        key = _import_row_identity_key(r)
        seen_idx[key] = i
    for r in incoming:
        key = _import_row_identity_key(r)
        if key in seen_idx:
            out[seen_idx[key]] = r
        else:
            seen_idx[key] = len(out)
            out.append(r)
    return _dedupe_import_rows(out)


def _looks_like_movie_program_name(name: str) -> bool:
    s = str(name or "").strip()
    if not s:
        return False
    if re.search(r"\((19|20)\d{2}\)", s):
        return True
    if re.search(r"\b(19|20)\d{2}\b", s):
        return True
    return False


def _episode_rows_for_archive_pick(cfg, sel: str, nikki_path: Path) -> list[dict[str, Any]]:
    sd = _showdef_for_archive_pick(cfg, sel)
    if sd is None or sd.kind != "series" or not sd.nikki_sheet or not nikki_path.is_file():
        return []
    style = sd.nikki_style or nikki.default_style_for_sheet(sd.nikki_sheet)
    hdrs = nikki.effective_column_headers(sd, style=style)
    try:
        return _archive_sheet_episodes(
            str(nikki_path.resolve()),
            _nikki_mtime(nikki_path),
            sd.nikki_sheet,
            style,
            sd.prefix or "",
            sd.nikki_row_filter,
            json.dumps(asdict(hdrs), sort_keys=True),
        )
    except Exception:
        return []


def _episode_num_text(ep: dict[str, Any]) -> str:
    for k in ("sheet_se", "se_compact", "ep_in_season", "code"):
        v = ep.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s and s != "—":
            return s
    return ""


def _slot_source_show_key(cfg, slot_show_text: str) -> Optional[str]:
    key, sd = resolve_show(str(slot_show_text or ""), cfg.shows)
    if sd is None or key == "literal":
        return None
    return key


def _semantic_group_for_show(cfg, show_key: str) -> str:
    sd = cfg.shows.get(show_key)
    if sd is None:
        return ""
    return str(getattr(sd, "semantic_group", None) or "").strip().lower()


def _title_key_variants(text: str) -> set[str]:
    s = " ".join(str(text or "").strip().lower().split())
    if not s:
        return set()
    out = {s}
    no_year = re.sub(r"\s*\((19|20)\d{2}\)\s*$", "", s).strip()
    if no_year:
        out.add(no_year)
    for cur in list(out):
        m = re.match(r"^(.*?),\s*(the|a|an)$", cur)
        if m:
            out.add(f"{m.group(2)} {m.group(1)}".strip())
    return {x for x in out if x}


@st.cache_data(show_spinner=False)
def _movie_semantic_groups(cfg_path_str: str) -> dict[str, str]:
    """Optional movie semantic-group map from config/movie_semantic_groups.json."""
    cfg_path = Path(cfg_path_str)
    cfg_dir = cfg_path.resolve().parent
    repo_root = cfg_dir.parent if cfg_dir.name.casefold() == "config" else cfg_dir
    p = repo_root / "config" / "movie_semantic_groups.json"
    if not p.is_file():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(raw, dict):
        return {}
    out: dict[str, str] = {}
    for k, v in raw.items():
        g = str(v or "").strip().lower()
        if not g:
            continue
        for kk in _title_key_variants(str(k)):
            out[kk] = g
    return out


@st.cache_data(show_spinner=False)
def _nikki_movie_semantic_groups(nikki_path_str: str, _workbook_mtime: float) -> dict[str, str]:
    """Infer movie semantic groups from Nikki movies tab Genre column."""
    p = Path(nikki_path_str)
    if not p.is_file():
        return {}
    try:
        df = pd.read_excel(p, sheet_name="movies", header=1)
    except Exception:
        return {}
    out: dict[str, str] = {}
    for _, r in df.iterrows():
        title = str(r.get("Title") or "").strip()
        if not title:
            continue
        genre_raw = str(r.get("Genre") or "").strip()
        if not genre_raw:
            continue
        # Use primary genre to keep filter options manageable.
        primary = genre_raw.split(",")[0].split("/")[0].strip().lower()
        if not primary:
            continue
        year_val = r.get("Year")
        year_num: Optional[int] = None
        try:
            if pd.notna(year_val):
                year_num = int(float(year_val))
        except Exception:
            year_num = None
        keys = [title]
        if year_num is not None:
            keys.append(f"{title} ({year_num})")
        for key in keys:
            for kk in _title_key_variants(key):
                out[kk] = primary
    return out


def _semantic_group_for_archive_option(
    cfg,
    opt: str,
    movie_groups: dict[str, str],
) -> str:
    imported = _parse_imported_content_option(opt)
    if imported is not None:
        for kk in _title_key_variants(imported):
            g = movie_groups.get(kk)
            if g:
                return g
        return ""
    if opt in cfg.shows:
        return str(getattr(cfg.shows[opt], "semantic_group", None) or "").strip().lower()
    raw_literal = _parse_literal_text_option(opt)
    title = raw_literal if raw_literal is not None else _display_name_for_archive_pick(cfg, opt)
    for kk in _title_key_variants(title):
        g = movie_groups.get(kk)
        if g:
            return g
    return ""


@st.cache_data(show_spinner=False)
def _movie_runtime_minutes(cfg_path_str: str) -> dict[str, int]:
    """Optional movie runtime map from config/movie_runtime_minutes.json."""
    cfg_path = Path(cfg_path_str)
    cfg_dir = cfg_path.resolve().parent
    repo_root = cfg_dir.parent if cfg_dir.name.casefold() == "config" else cfg_dir
    p = repo_root / "config" / "movie_runtime_minutes.json"
    if not p.is_file():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(raw, dict):
        return {}
    out: dict[str, int] = {}
    for k, v in raw.items():
        try:
            mins = int(v)
        except (TypeError, ValueError):
            continue
        if mins <= 0:
            continue
        for kk in _title_key_variants(str(k)):
            out[kk] = mins
    return out


def _runtime_for_archive_option(cfg, opt: str, runtime_map: dict[str, int]) -> Optional[int]:
    title = _display_name_for_archive_pick(cfg, opt)
    for kk in _title_key_variants(title):
        if kk in runtime_map:
            return int(runtime_map[kk])
    return None


def _semantic_candidates(cfg, *, group: str, kind: str, exclude_keys: set[str]) -> list[str]:
    out: list[str] = []
    for k, sd in cfg.shows.items():
        if sd.kind != kind:
            continue
        if k in exclude_keys:
            continue
        g = str(getattr(sd, "semantic_group", None) or "").strip().lower()
        if group and g != group:
            continue
        out.append(k)
    out.sort(key=lambda k: cfg.shows[k].display_name.casefold())
    return out


def _movie_program_picker_options(
    cfg,
    extra_tab_names: list[str],
    template_slots: list[dict[str, Any]],
    nikki_path: Path,
) -> list[str]:
    """Movie/program options from config, movie-like archive tabs, and literal slot labels."""
    opts: list[str] = [k for k, sd in cfg.shows.items() if sd.kind == "literal"]
    opts.extend(workbook_tab_option(t) for t in extra_tab_names if _looks_like_movie_program_name(t))
    opts.extend(_nikki_movie_catalog_options(nikki_path))
    for row in template_slots:
        show_text = str(row.get("show", "")).strip()
        if not show_text:
            continue
        _, sd = resolve_show(show_text, cfg.shows)
        if sd is not None and sd.kind == "series":
            continue
        if sd is not None and sd.kind == "literal" and show_text == sd.display_name.strip():
            continue
        opts.append(_literal_text_option(show_text))
    uniq: list[str] = []
    seen: set[str] = set()
    for opt in opts:
        if opt in seen:
            continue
        seen.add(opt)
        uniq.append(opt)
    uniq.sort(key=lambda opt: _display_name_for_archive_pick(cfg, opt).casefold())
    return uniq


def _auto_movie_candidates(
    cfg,
    cfg_path: Path,
    extra_tab_names: list[str],
    template_slots: list[dict[str, Any]],
    nikki_path: Path,
    source_group: str,
) -> list[str]:
    """Movie/program candidates for auto-swap, preferring semantic and movie-like matches."""
    movie_opts = _movie_program_picker_options(cfg, extra_tab_names, template_slots, nikki_path)
    if not movie_opts:
        return []
    movie_groups = _nikki_movie_semantic_groups(str(nikki_path.resolve()), _nikki_mtime(nikki_path))
    movie_groups.update(_movie_semantic_groups(str(cfg_path.resolve())))
    candidates: list[str] = []
    if source_group:
        candidates = [
            opt
            for opt in movie_opts
            if (_semantic_group_for_archive_option(cfg, opt, movie_groups) or "") == source_group
        ]
    if not candidates:
        # Fallback to movie-like titles (typically include a year) before generic literals.
        movie_like = [
            opt
            for opt in movie_opts
            if re.search(r"\((19|20)\d{2}\)", _display_name_for_archive_pick(cfg, opt))
        ]
        candidates = movie_like if movie_like else list(movie_opts)
    uniq = sorted(set(candidates), key=lambda opt: _display_name_for_archive_pick(cfg, opt).casefold())
    return uniq


def _literal_options_from_slots(cfg, slots: list[dict[str, Any]]) -> list[str]:
    """Distinct literal show labels discovered in schedule slots."""
    opts: list[str] = []
    seen: set[str] = set()
    for row in slots:
        show_text = str(row.get("show", "")).strip()
        if not show_text:
            continue
        _, sd = resolve_show(show_text, cfg.shows)
        if sd is not None and sd.kind == "series":
            continue
        if sd is not None and sd.kind == "literal" and show_text == sd.display_name.strip():
            continue
        token = _literal_text_option(show_text)
        if token in seen:
            continue
        seen.add(token)
        opts.append(token)
    opts.sort(key=lambda opt: _display_name_for_archive_pick(cfg, opt).casefold())
    return opts


def _slot_rows_from_grids_workbook(path: Path) -> list[dict[str, Any]]:
    """Extract slot-like rows from every week sheet in a GRIDS workbook."""
    out: list[dict[str, Any]] = []
    if not path.is_file():
        return out
    for sheet in _list_grids_data_sheets(path):
        mon = parse_sheet_tab_monday(sheet)
        if mon is None:
            continue
        try:
            grid = load_grid_sheet(str(path), sheet)
        except Exception:
            continue
        dates = day_dates(mon)
        for day_idx in range(7):
            col = [grid[r][day_idx] for r in range(48)]
            try:
                segs = segments_for_day(col)
            except ValueError:
                continue
            for seg in segs:
                out.append(
                    {
                        "show": str(seg.cell_text).strip(),
                        "date_iso": dates[day_idx].isoformat(),
                        "start_slot": int(seg.start_slot),
                    }
                )
    return out


def _grids_preview_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Show blank cells in the GRIDS preview instead of ``None`` / ``nan`` text."""

    def _cell(v: Any) -> Any:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        if isinstance(v, str) and v.strip().lower() in ("none", "nan"):
            return ""
        return v

    dfc = df.copy()
    if hasattr(dfc, "map"):
        return dfc.map(_cell)
    return dfc.applymap(_cell)  # type: ignore[attr-defined]


def _render_binge_grids_preview(
    *,
    key_prefix: str,
    show_swap: bool,
    cfg=None,
    cfg_path: Optional[Path] = None,
    nikki_path: Optional[Path] = None,
) -> None:
    """In-page tables from the last generated BINGE / BINGE GRIDS in session (optional swap → archive)."""
    if "binge_path" not in st.session_state or "grids_path" not in st.session_state:
        return
    bp = Path(st.session_state["binge_path"])
    gp = Path(st.session_state["grids_path"])
    if not bp.is_file() or not gp.is_file():
        st.warning("BINGE or BINGE GRIDS file is missing on disk for this session.")
        return

    binge_sheets = _list_binge_data_sheets(bp)
    grid_sheets = _list_grids_data_sheets(gp)
    if not binge_sheets:
        st.warning("No data sheets found in the BINGE workbook (excluding notes).")
        return

    st.markdown("##### Preview in app")
    st.caption("Same files as the downloads — pick a week tab for each workbook.")

    c1, c2 = st.columns(2)
    with c1:
        bs = st.selectbox(
            "BINGE week tab",
            binge_sheets,
            key=f"{key_prefix}_preview_binge_sheet",
        )
    with c2:
        gs = st.selectbox(
            "BINGE GRIDS week tab",
            grid_sheets if grid_sheets else ["(no sheets)"],
            key=f"{key_prefix}_preview_grids_sheet",
        )

    try:
        binge_df = pd.read_excel(bp, sheet_name=bs)
        binge_df = normalize_binge_df_columns(binge_df)
    except Exception as e:
        st.error(f"Could not read BINGE sheet `{bs}`: {e}")
        binge_df = None

    if binge_df is not None:
        st.markdown("###### BINGE")
        show_col = _find_binge_column_ci(binge_df, "SHOW")
        picked_row_idx: Optional[int] = None
        picked_row_indices: list[int] = []

        if show_swap and show_col and len(binge_df) > 0:
            sel_supported = _dataframe_row_selection_supported()
            df_key = f"{key_prefix}_binge_swap_df"
            if sel_supported:
                event = st.dataframe(
                    binge_df,
                    use_container_width=True,
                    height=340,
                    hide_index=True,
                    on_select="rerun",
                    selection_mode="multi-row",
                    key=df_key,
                )
                rows_sel: list[int] = []
                try:
                    rows_sel = list(event["selection"]["rows"])  # type: ignore[index]
                except (KeyError, TypeError, AttributeError):
                    pass
                picked_row_indices = [int(i) for i in rows_sel if isinstance(i, int) and 0 <= int(i) < len(binge_df)]
                if rows_sel:
                    try:
                        picked_row_idx = int(rows_sel[0])
                    except (TypeError, ValueError, IndexError):
                        picked_row_idx = None
                if picked_row_idx is not None and not (0 <= picked_row_idx < len(binge_df)):
                    picked_row_idx = None
            else:
                st.dataframe(binge_df, use_container_width=True, height=340, hide_index=True)
                st.warning(
                    "Upgrade Streamlit for row selection in the table: "
                    "`pip install -U \"streamlit>=1.35\"`. Using row list instead."
                )
                picked_row_idx = int(
                    st.selectbox(
                        "Row to swap",
                        list(range(len(binge_df))),
                        format_func=lambda i: _binge_row_swap_summary(binge_df, int(i)),
                        key=f"{key_prefix}_binge_row_fallback",
                    )
                )
                picked_row_indices = [picked_row_idx]

            st.markdown("###### Change a show")
            st.info(
                "**After the BINGE build:** select the **row** you’re replacing, then **Swap for…** and pick **whatever show** "
                "you want from the archive. **Time and day stay the same** — only the program in that slot changes in your **grids** "
                "(and the base schedule file if the show is new). Run **Create Schedule** again so the spreadsheet matches."
            )
            st.caption("One row → **Swap for… → Available Content** → confirm.")
            if picked_row_idx is not None:
                sv = binge_df.iloc[picked_row_idx][show_col]
                show_val = str(sv).strip() if pd.notna(sv) else ""
                st.caption(f"**Selected row {picked_row_idx + 1} · SHOW:** {show_val or '—'}")
            else:
                st.caption("**No row selected yet** — click a row in the table above.")

            if cfg is not None and cfg_path is not None and nikki_path is not None:
                st.markdown("###### Block swap")
                if picked_row_indices:
                    selected_rows = sorted(set(int(i) for i in picked_row_indices))
                    row_meta: list[dict[str, Any]] = []
                    for idx in selected_rows:
                        sv = binge_df.iloc[idx][show_col]
                        show_val = str(sv).strip() if pd.notna(sv) else ""
                        row_meta.append({"idx": idx, "show": show_val})
                    st.caption(
                        f"Assign replacements for **{len(selected_rows)}** selected row(s), then apply all at once."
                    )
                    yaml_keys = sorted(cfg.shows.keys(), key=lambda k: cfg.shows[k].display_name.lower())
                    extra_tab_names: list[str] = []
                    if nikki_path.is_file():
                        tabs = _nikki_workbook_sheet_names(str(nikki_path.resolve()), _nikki_mtime(nikki_path))
                        extra_tab_names = workbook_tabs_not_in_yaml(cfg, tabs)
                    option_keys = yaml_keys + [workbook_tab_option(t) for t in extra_tab_names]
                    source_keys = {
                        k
                        for k in (_slot_source_show_key(cfg, str(m.get("show", ""))) for m in row_meta)
                        if k is not None
                    }
                    source_groups = [g for g in (_semantic_group_for_show(cfg, k) for k in source_keys) if g]
                    source_group = sorted(source_groups)[0] if source_groups else ""
                    mode = st.radio(
                        "Swap mode",
                        ("Quick manual", "Advanced (Create Schedule logic)"),
                        horizontal=True,
                        key=f"{key_prefix}_block_swap_mode",
                    )
                    adv_fill_choice = "manual_replace"
                    adv_manual_kind = "show"
                    if mode == "Advanced (Create Schedule logic)":
                        if source_group:
                            st.caption(f"Semantic source group: `{source_group}`")
                        adv_fill_choice = st.radio(
                            "Replacement style",
                            (
                                "auto_show",
                                "auto_movie",
                                "manual_replace",
                            ),
                            format_func=lambda v: {
                                "auto_show": "Auto-swap with shows",
                                "auto_movie": "Auto-swap with movies",
                                "manual_replace": "Manual replace",
                            }[str(v)],
                            horizontal=True,
                            key=f"{key_prefix}_block_adv_fill_choice",
                        )
                        if adv_fill_choice == "manual_replace":
                            adv_manual_kind = st.radio(
                                "Manual replace type",
                                ("show", "movie"),
                                format_func=lambda v: "Show" if v == "show" else "Movie/program",
                                horizontal=True,
                                key=f"{key_prefix}_block_adv_manual_kind",
                            )

                    def _opt_label(opt: str) -> str:
                        tab = parse_workbook_tab_option(opt)
                        if tab is not None:
                            return f"{tab} _(not on schedule)_"
                        return cfg.shows[opt].display_name

                    adv_movie_opts = _movie_program_picker_options(cfg, extra_tab_names, [], nikki_path)
                    adv_auto_show_opts = _semantic_candidates(
                        cfg,
                        group=source_group if mode == "Advanced (Create Schedule logic)" else "",
                        kind="series",
                        exclude_keys=source_keys,
                    )
                    if not adv_auto_show_opts:
                        adv_auto_show_opts = _semantic_candidates(cfg, group="", kind="series", exclude_keys=source_keys)
                    adv_auto_movie_opts = _auto_movie_candidates(
                        cfg,
                        cfg_path,
                        extra_tab_names,
                        [],
                        nikki_path,
                        source_group,
                    )

                    for r_idx, meta in enumerate(row_meta):
                        idx = int(meta["idx"])
                        show_val = str(meta["show"])
                        select_opts = option_keys
                        select_fmt = _opt_label
                        if mode == "Advanced (Create Schedule logic)":
                            if adv_fill_choice == "auto_show":
                                select_opts = adv_auto_show_opts or option_keys
                                select_fmt = lambda opt: cfg.shows.get(opt, ShowDef(key="literal", display_name=str(opt), kind="literal")).display_name if opt in cfg.shows else _display_name_for_archive_pick(cfg, str(opt))
                            elif adv_fill_choice == "auto_movie":
                                select_opts = adv_auto_movie_opts or adv_movie_opts or option_keys
                                select_fmt = lambda opt: _display_name_for_archive_pick(cfg, str(opt))
                            elif adv_manual_kind == "movie":
                                select_opts = adv_movie_opts or option_keys
                                select_fmt = lambda opt: _display_name_for_archive_pick(cfg, str(opt))
                            else:
                                select_opts = option_keys
                                select_fmt = _opt_label
                        if not select_opts:
                            st.warning("No replacement options available for the selected mode.")
                            continue
                        default_idx = r_idx % len(select_opts) if mode == "Advanced (Create Schedule logic)" else 0
                        st.selectbox(
                            f"Row {idx + 1}: {show_val or '—'}",
                            select_opts,
                            index=default_idx,
                            format_func=select_fmt,
                            key=f"{key_prefix}_block_pick_{idx}",
                        )

                    if st.button(
                        "Apply block swap to selected rows",
                        type="primary",
                        use_container_width=True,
                        key=f"{key_prefix}_apply_block_swap",
                    ):
                        applied = 0
                        all_msgs: list[str] = []
                        first_anchor: Optional[dict[str, Any]] = None
                        for idx in selected_rows:
                            pick = st.session_state.get(f"{key_prefix}_block_pick_{idx}")
                            if not pick:
                                all_msgs.append(f"Row {idx + 1}: no replacement selected.")
                                continue
                            sv = binge_df.iloc[idx][show_col]
                            show_val = str(sv).strip() if pd.notna(sv) else ""
                            if not show_val:
                                all_msgs.append(f"Row {idx + 1}: no SHOW value.")
                                continue
                            anchor_dict = _schedule_anchor_dict_from_binge_row(binge_df, idx)
                            if anchor_dict is None:
                                all_msgs.append(f"Row {idx + 1}: missing DATE/START TIME anchor.")
                                continue
                            ok, swap_msgs = apply_show_swap(
                                cfg_path,
                                [show_val],
                                str(pick),
                                schedule_anchor=anchor_dict,
                            )
                            all_msgs.extend([f"Row {idx + 1}: {m}" for m in swap_msgs])
                            if ok:
                                applied += 1
                                if first_anchor is None:
                                    first_anchor = anchor_dict
                        auto_ok = False
                        if applied > 0 and first_anchor is not None:
                            r_ok, r_msgs, paths = _regenerate_binge_for_month(cfg_path, first_anchor)
                            all_msgs.extend(r_msgs)
                            if r_ok and paths:
                                bp, gp, od = paths
                                st.session_state["binge_path"] = bp
                                st.session_state["grids_path"] = gp
                                st.session_state["out_dir"] = od
                                auto_ok = True
                        st.session_state["swap_result"] = {
                            "summary": f"Applied block swap to **{applied}** row(s).",
                            "messages": all_msgs,
                            "auto_export_ok": auto_ok,
                        }
                        st.rerun()
                else:
                    st.caption("Select one or more BINGE rows above to enable block swap.")

            if st.button(
                "Swap for… → Available Content",
                type="secondary",
                use_container_width=True,
                key=f"{key_prefix}_swap_open_archive",
            ):
                if picked_row_idx is None:
                    st.warning("Select exactly one row in the BINGE table first.")
                else:
                    sv = binge_df.iloc[picked_row_idx][show_col]
                    show_val = str(sv).strip() if pd.notna(sv) else ""
                    if not show_val:
                        st.warning("That row has no SHOW value.")
                    else:
                        anchor_dict = _schedule_anchor_dict_from_binge_row(binge_df, picked_row_idx)
                        if anchor_dict is None:
                            st.warning(
                                "This BINGE sheet needs **DATE** and **START TIME** columns so we only edit "
                                "**that** clock slot in the grids (not every week)."
                            )
                        else:
                            st.session_state["swap_context"] = {
                                "old_show_labels": [show_val],
                                "binge_sheet": bs,
                                "binge_row": picked_row_idx + 1,
                                "schedule_anchor": anchor_dict,
                            }
                            st.session_state[_MAIN_NAV_PENDING_KEY] = _NAV_ARCHIVE
                            st.rerun()
        else:
            st.dataframe(binge_df, use_container_width=True, height=340, hide_index=True)
            if show_swap and not show_col:
                st.warning("This BINGE sheet has no **SHOW** column — use another tab or rebuild.")
            elif show_swap and len(binge_df) == 0:
                st.info("This tab has no rows.")

    if grid_sheets and gs != "(no sheets)":
        try:
            grids_df = pd.read_excel(gp, sheet_name=gs, header=None)
        except Exception as e:
            st.error(f"Could not read GRIDS sheet `{gs}`: {e}")
        else:
            st.markdown("###### BINGE GRIDS")
            st.caption("Full sheet layout; program cells are typically rows 5–52, columns B–H (Mon–Sun).")
            max_r = min(len(grids_df), 52)
            st.dataframe(
                _grids_preview_dataframe(grids_df.iloc[:max_r]),
                use_container_width=True,
                height=340,
                hide_index=True,
            )


def _render_content_archive(cfg, cfg_path: Path, nikki_path: Path) -> None:
    title_col, action_col = st.columns([8, 2])
    with title_col:
        st.markdown(f"## {_NAV_ARCHIVE}")
    with action_col:
        if "archive_show_upload_picker" not in st.session_state:
            st.session_state["archive_show_upload_picker"] = False
        if st.button("Upload new content", key="archive_upload_new_content_btn", type="primary"):
            st.session_state["archive_show_upload_picker"] = not bool(
                st.session_state.get("archive_show_upload_picker", False)
            )

    swap_ctx = st.session_state.get("swap_context")
    if swap_ctx:
        olds = swap_ctx.get("old_show_labels") or []
        tab_hint = swap_ctx.get("binge_sheet")
        row_hint = swap_ctx.get("binge_row")
        ctx_bits: list[str] = []
        if tab_hint:
            ctx_bits.append(f"tab `{tab_hint}`")
        if row_hint:
            ctx_bits.append(f"row **{row_hint}**")
        ctx_suffix = f" ({', '.join(ctx_bits)})" if ctx_bits else ""
        st.info(
            f"**Swap:** Under **Pick content**, choose the program you want in that **same time slot**, then "
            f"**Use selected content as replacement**. "
            f"Current label: **{', '.join(olds)}**{ctx_suffix}."
        )

    if st.session_state.get("archive_show_upload_picker"):
        uploaded = st.file_uploader(
            "Upload content metadata file",
            type=["xlsx", "csv"],
            key="archive_import_upload_compact",
            label_visibility="collapsed",
        )
        if uploaded is not None:
            try:
                payload = uploaded.getvalue()
                sig = f"{uploaded.name}:{hashlib.sha1(payload).hexdigest()}"
                if st.session_state.get("archive_import_last_sig") != sig:
                    imp_rows, _imp_notes = _parse_uploaded_content_file(uploaded.name, payload)
                    existing = _load_imported_catalog_rows(cfg_path)
                    merged = _merge_import_rows(existing, imp_rows)
                    _save_imported_catalog_rows(cfg_path, merged)
                    st.session_state["archive_import_last_sig"] = sig
                    st.session_state["archive_show_upload_picker"] = False
                    st.toast(f"Imported {len(merged) - len(existing)} new row(s).")
                    st.rerun()
            except Exception as e:
                st.error(f"Import failed: {e}")

    yaml_keys = sorted(cfg.shows.keys(), key=lambda k: cfg.shows[k].display_name.lower())
    extra_tab_names: list[str] = []
    if nikki_path.is_file():
        tabs = _nikki_workbook_sheet_names(str(nikki_path.resolve()), _nikki_mtime(nikki_path))
        extra_tab_names = workbook_tabs_not_in_yaml(cfg, tabs)
    extra_opts = [workbook_tab_option(t) for t in extra_tab_names]
    nikki_movie_opts = _nikki_movie_catalog_options(nikki_path)
    slot_rows, _slot_warn = _schedule_template_slots(cfg.weeks)
    # Also include titles from known GRIDS outputs (e.g., recently generated files in out/).
    grid_paths: list[Path] = []
    seen_grids: set[str] = set()

    def _add_grid_path(p: Path) -> None:
        rp = str(p.resolve())
        if rp in seen_grids:
            return
        seen_grids.add(rp)
        grid_paths.append(p)

    for w in cfg.weeks:
        gp = Path(w.grids_file)
        if not gp.is_absolute():
            gp = (cfg_path.resolve().parent / gp).resolve()
        if gp.is_file():
            _add_grid_path(gp)
    sess_gp = st.session_state.get("grids_path")
    if sess_gp:
        p = Path(str(sess_gp))
        if p.is_file():
            _add_grid_path(p)
    for p in _schedule_workbook_candidates(cfg_path):
        if "GRIDS" in p.name.upper():
            _add_grid_path(p)
    for gp in grid_paths:
        slot_rows.extend(_slot_rows_from_grids_workbook(gp))

    slot_literal_opts = _literal_options_from_slots(cfg, slot_rows)
    imported_rows = _load_imported_catalog_rows(cfg_path)
    imported_movie_names = sorted(
        {
            str(r.get("display_name") or "").strip()
            for r in imported_rows
            if str(r.get("display_name") or "").strip() and str(r.get("content_type") or "").strip().lower() != "series"
        },
        key=str.casefold,
    )
    imported_series_names = sorted(
        {
            str(r.get("series_title") or r.get("display_name") or "").strip()
            for r in imported_rows
            if str(r.get("series_title") or r.get("display_name") or "").strip()
            and str(r.get("content_type") or "").strip().lower() == "series"
        },
        key=str.casefold,
    )
    imported_opts = [_imported_content_option(x) for x in (imported_series_names + imported_movie_names)]
    imported_rows_by_name: dict[str, list[dict[str, Any]]] = {}
    for r in imported_rows:
        name = (
            str(r.get("series_title") or r.get("display_name") or "").strip()
            if str(r.get("content_type") or "").strip().lower() == "series"
            else str(r.get("display_name") or "").strip()
        )
        if not name:
            continue
        imported_rows_by_name.setdefault(name, []).append(r)

    all_option_keys = yaml_keys + extra_opts + slot_literal_opts + nikki_movie_opts + imported_opts
    archive_view = st.radio(
        "Archive view",
        ("All", "Shows (series)", "Movies/programs"),
        horizontal=True,
        key="archive_view_filter",
    )

    def _is_movie_option(opt: str) -> bool:
        raw_literal = _parse_literal_text_option(opt)
        if raw_literal is not None:
            return True
        tab = parse_workbook_tab_option(opt)
        if tab is not None:
            return _looks_like_movie_program_name(tab)
        sd = cfg.shows.get(opt)
        return bool(sd and sd.kind == "literal")

    if archive_view == "Shows (series)":
        option_keys = [opt for opt in all_option_keys if not _is_movie_option(opt)]
    elif archive_view == "Movies/programs":
        option_keys = [opt for opt in all_option_keys if _is_movie_option(opt)]
    else:
        option_keys = list(all_option_keys)

    movie_groups = _nikki_movie_semantic_groups(str(nikki_path.resolve()), _nikki_mtime(nikki_path))
    movie_groups.update(_movie_semantic_groups(str(cfg_path.resolve())))
    for name, rows in imported_rows_by_name.items():
        genre = ""
        for r in rows:
            g = str(r.get("genre") or "").strip().lower()
            if g:
                genre = g
                break
        if genre:
            for kk in _title_key_variants(name):
                movie_groups[kk] = genre
    genre_vals = sorted(
        {
            _semantic_group_for_archive_option(cfg, opt, movie_groups) or "unlabeled"
            for opt in option_keys
        }
    )
    if st.session_state.get("archive_genre_filter") not in (["All genres"] + genre_vals):
        st.session_state["archive_genre_filter"] = "All genres"
    genre_pick = st.selectbox(
        "Genre filter",
        ["All genres"] + genre_vals,
        key="archive_genre_filter",
    )
    if genre_pick != "All genres":
        option_keys = [
            opt
            for opt in option_keys
            if (_semantic_group_for_archive_option(cfg, opt, movie_groups) or "unlabeled") == genre_pick
        ]

    if not option_keys:
        st.info("No content matches the current archive/genre filters.")
        return

    def _archive_option_label(opt: str) -> str:
        imported = _parse_imported_content_option(opt)
        if imported is not None:
            rows = imported_rows_by_name.get(imported, [])
            kind = "series import" if any(str(r.get("content_type") or "").lower() == "series" for r in rows) else "movie import"
            return f"{imported} _({kind})_"
        raw_literal = _parse_literal_text_option(opt)
        if raw_literal is not None:
            return f"{raw_literal} _(from schedule)_"
        tab = parse_workbook_tab_option(opt)
        if tab is not None:
            suffix = "movie/program tab" if _looks_like_movie_program_name(tab) else "not on schedule"
            return f"{tab} _({suffix})_"
        sd = cfg.shows[opt]
        kind = "movie/program" if sd.kind == "literal" else "series"
        return f"{sd.display_name} _({kind})_"

    none_opt = "__none__"
    select_opts = [none_opt] + option_keys
    prior_pick = st.session_state.get("archive_show_pick")
    if prior_pick not in select_opts:
        st.session_state["archive_show_pick"] = none_opt

    sel = st.selectbox(
        "Pick content",
        select_opts,
        format_func=lambda opt: "— Select content —" if opt == none_opt else _archive_option_label(opt),
        key="archive_show_pick",
    )
    if swap_ctx:
        if st.button(
            "Use selected content as replacement",
            type="primary",
            use_container_width=True,
            key="archive_swap_confirm",
        ):
            pick = st.session_state.get("archive_show_pick")
            if not pick or pick == none_opt:
                st.warning("Pick content in the list first.")
            else:
                ok, swap_msgs = apply_show_swap(
                    cfg_path,
                    list(swap_ctx.get("old_show_labels") or []),
                    pick,
                    schedule_anchor=swap_ctx.get("schedule_anchor"),
                )
                if ok:
                    combined_msgs = list(swap_msgs)
                    auto_ok = False
                    anchor = swap_ctx.get("schedule_anchor")
                    noop_same_show = "no grid change" in " ".join(swap_msgs).casefold()
                    if anchor and not noop_same_show:
                        r_ok, r_msgs, paths = _regenerate_binge_for_month(cfg_path, anchor)
                        combined_msgs.extend(r_msgs)
                        if r_ok and paths:
                            bp, gp, od = paths
                            st.session_state["binge_path"] = bp
                            st.session_state["grids_path"] = gp
                            st.session_state["out_dir"] = od
                            auto_ok = True
                    st.session_state["swap_result"] = {
                        "old_show_labels": list(swap_ctx.get("old_show_labels") or []),
                        "archive_pick": pick,
                        "new_display": _display_name_for_archive_pick(cfg, pick),
                        "messages": combined_msgs,
                        "auto_export_ok": auto_ok,
                    }
                    st.session_state.pop("swap_context", None)
                    st.session_state[_MAIN_NAV_PENDING_KEY] = _NAV_EDIT_SCHEDULE
                    st.rerun()
                else:
                    for m in swap_msgs:
                        st.error(m)
        if st.button(
            "Cancel swap",
            use_container_width=True,
            key="archive_swap_cancel",
        ):
            st.session_state.pop("swap_context", None)
            st.rerun()

    if sel == none_opt:
        st.caption("Choose a show/movie/program to view details.")
        return

    imported_pick = _parse_imported_content_option(sel)
    if imported_pick is not None:
        rows = imported_rows_by_name.get(imported_pick, [])
        st.caption("Imported catalog entry")
        with _archive_detail_panel():
            st.markdown(f"## {imported_pick}")
            st.caption(f"Imported rows: **{len(rows)}**")
            if rows:
                genre_vals = sorted({str(r.get("genre") or "").strip() for r in rows if str(r.get("genre") or "").strip()})
                if genre_vals:
                    st.caption("Genre: " + ", ".join(genre_vals))
                runtime_vals = sorted({str(r.get("runtime_minutes") or "").strip() for r in rows if r.get("runtime_minutes")})
                if runtime_vals:
                    st.caption("Runtime(s): " + ", ".join(f"{v}m" for v in runtime_vals))
                meta_cols = [
                    "content_type",
                    "series_title",
                    "episode_number",
                    "episode_title",
                    "genre",
                    "runtime_minutes",
                    "original_airdate",
                    "production_company",
                    "copyright",
                ]
                st.dataframe(pd.DataFrame(rows)[meta_cols].head(60), use_container_width=True, hide_index=True)
        return

    raw_literal = _parse_literal_text_option(sel)
    tab_only = parse_workbook_tab_option(sel)
    if raw_literal is not None:
        st.caption("Literal title pulled from schedule content.")
    elif tab_only is not None:
        if _looks_like_movie_program_name(tab_only):
            st.caption(f"Excel tab `{tab_only}` — movie/program tab (not on schedule)")
        else:
            st.caption(f"Excel tab `{tab_only}` — **not on schedule**")
    else:
        st.caption(f"Schedule entry `{sel}`")

    browse_only = tab_only is not None and raw_literal is None
    if raw_literal is not None:
        sd = ShowDef(key="literal", display_name=raw_literal, kind="literal")
    elif browse_only and _looks_like_movie_program_name(tab_only or ""):
        sd = ShowDef(key="literal", display_name=str(tab_only), kind="literal")
    else:
        sd = synthetic_series_for_tab(tab_only) if browse_only else cfg.shows[sel]
    with _archive_detail_panel():
        st.markdown(f"## {sd.display_name}")
        if browse_only:
            st.caption(
                "Browse only — add this show to your **base schedule file** (with the same **`nikki_sheet`** "
                "name as this tab) so **Create Schedule** can use it."
            )
        else:
            st.caption(f"Base schedule key `{sel}`")

        if sd.kind == "series":
            style = sd.nikki_style or (
                nikki.default_style_for_sheet(sd.nikki_sheet) if sd.nikki_sheet else "generic"
            )
            hdrs = nikki.effective_column_headers(sd, style=style)
            sheet_ok = _nikki_sheet_exists(nikki_path, sd.nikki_sheet)

            _render_archive_episode_browser(
                sel,
                sd,
                nikki_path,
                style=style,
                hdrs=hdrs,
                sheet_ok=sheet_ok,
                browse_only=browse_only,
            )

            with st.expander("Source & technical details", expanded=False):
                st.caption(
                    f"Parser **`{style}`** · start index **{sd.start_episode_index}** · prefix **{sd.prefix or '—'}**"
                )
                if sd.nikki_row_filter:
                    st.caption(f"Row filter: `{sd.nikki_row_filter}`")
                st.markdown("**Spreadsheet tab**")
                st.code(sd.nikki_sheet or "(none)", language=None)
                if sheet_ok is True:
                    st.success("Tab name matches your spreadsheet file.")
                elif sheet_ok is False:
                    st.error("Tab not found — compare with the exact name in Excel.")
                elif nikki_path.is_file():
                    st.warning("Could not read the spreadsheet file.")
                else:
                    st.warning("Spreadsheet path missing in base schedule.")
                if nikki_path.is_file():
                    if st.button(
                        "Open folder",
                        key=f"archive_open_{_archive_wkey(sel)}",
                        type="secondary",
                        use_container_width=True,
                    ):
                        err = _open_folder(nikki_path.parent)
                        if err:
                            if "download buttons" in err:
                                st.info(err)
                            else:
                                st.error(err)
                with st.expander("Column headers (advanced)", expanded=False):
                    if sd.nikki_columns is not None:
                        st.json({k: v for k, v in asdict(sd.nikki_columns).items() if v is not None})
                    else:
                        st.json({k: v for k, v in asdict(hdrs).items() if v is not None})
                with st.expander(
                    "Row rules (advanced)",
                    expanded=sd.nikki_row_filter == nikki.ROW_FILTER_GREEN_EPISODE_CELL,
                ):
                    if sd.nikki_row_filter:
                        st.code(sd.nikki_row_filter, language=None)
                        if sd.nikki_row_filter == nikki.ROW_FILTER_GREEN_EPISODE_CELL:
                            st.caption(
                                "Only green-filled **Episode** cells count for the schedule; **Create Schedule** "
                                "uses the same rule, and this table matches it."
                            )
                    else:
                        st.caption("Standard rows — no extra filter.")
        else:
            st.markdown("This show is filled from the **weekly grid** only (no episode list).")
            st.metric("Kind", "Literal")
            st.caption(
                "To swap a literal slot, edit the grid Excel for that week or change how the cell text "
                "maps to **display_name** in your base schedule—use **Create Schedule** to confirm names match."
            )


def _render_last_build_outputs(cfg, cfg_path: Path) -> None:
    """Download buttons and details when a build has been run this session."""
    if "binge_path" not in st.session_state or "grids_path" not in st.session_state:
        return
    bp = st.session_state["binge_path"]
    gp = st.session_state["grids_path"]
    od: Path = st.session_state["out_dir"]
    with open(bp, "rb") as f:
        binge_bytes = f.read()
    with open(gp, "rb") as f:
        grids_bytes = f.read()
    st.download_button(
        "BINGE.xlsx",
        binge_bytes,
        file_name="BINGE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_binge_shared",
    )
    st.download_button(
        "BINGE GRIDS.xlsx",
        grids_bytes,
        file_name="BINGE GRIDS.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_grids_shared",
    )
    if st.button("Open output folder", use_container_width=True, key="open_out_shared"):
        err = _open_folder(od)
        if err:
            if "download buttons" in err:
                st.info(err)
            else:
                st.error(err)
        else:
            st.toast(f"Opened: {od}")

    with st.expander("Details", expanded=False):
        st.caption(
            "**BINGE.xlsx** episode code / # / name columns come from your **Nikki** content workbook and the "
            "saved cursor file — not from the grids file. Grids only say *what show* airs *when*. The downloaded "
            "**BINGE GRIDS.xlsx** keeps the same program text as your grids source (titles), not episode lines."
        )
        cur = resolved_cursor_state_path(cfg)
        if cur:
            st.caption(f"Episode order save file: `{cur}`")
        st.caption(f"Wrap when a show runs out: **{cfg.wrap_episodes}**")


def _schedule_workbook_candidates(cfg_path: Path) -> list[Path]:
    """Candidate BINGE/GRIDS workbooks users can load in Edit schedules."""
    cfg_dir = cfg_path.resolve().parent
    repo_root = cfg_dir.parent if cfg_dir.name.casefold() == "config" else cfg_dir
    roots = [repo_root / "data", repo_root / "out"]
    out: list[Path] = []
    seen: set[str] = set()
    for r in roots:
        if not r.is_dir():
            continue
        for p in sorted(r.glob("*.xlsx"), key=lambda x: x.name.casefold()):
            nm = p.name
            if nm.startswith("~$") or nm.endswith(".xlsx~"):
                continue
            rp = str(p.resolve())
            if rp in seen:
                continue
            seen.add(rp)
            out.append(Path(rp))
    return out


def _render_schedule_tab(cfg, cfg_path: Path, nikki_path: Path) -> None:
    sr = st.session_state.get("swap_result")
    if sr:
        if sr.get("summary"):
            st.success(str(sr.get("summary")))
        elif sr.get("auto_export_ok"):
            st.success(
                f"**Grids updated** and **BINGE files regenerated** for that month: **{', '.join(sr['old_show_labels'])}** → "
                f"**{sr['new_display']}** (`{sr['archive_pick']}`). Downloads below are the new **BINGE.xlsx** / **BINGE GRIDS.xlsx**."
            )
        elif sr.get("auto_export_ok") is False:
            st.success(
                f"**Grids updated** for that slot: **{', '.join(sr['old_show_labels'])}** → **{sr['new_display']}** "
                f"(`{sr['archive_pick']}`). **BINGE export** did not run automatically — use **Create Schedule**, "
                "or see *What changed* for details."
            )
        else:
            st.success(
                f"**Grids updated** for that slot: **{', '.join(sr['old_show_labels'])}** → **{sr['new_display']}** "
                f"(`{sr['archive_pick']}`). Run **Create Schedule** to refresh **BINGE.xlsx**."
            )
        msgs = sr.get("messages") or []
        if msgs:
            with st.expander("What changed", expanded=True):
                for m in msgs:
                    st.markdown(f"- {m}")
        if st.button("Dismiss note", key="schedule_dismiss_swap"):
            st.session_state.pop("swap_result", None)
            st.rerun()
        st.divider()

    st.markdown(
        "Your latest export is here and on **Create Schedule**. Below: pick the **BINGE row** to replace, then choose the **archive** show — "
        "**clock times stay put**; grids update for the next build."
    )
    candidates = _schedule_workbook_candidates(cfg_path)
    if candidates:
        st.markdown("##### Schedule source")
        src_mode = st.radio(
            "Preview/edit source",
            ("Latest session export", "Existing files on disk"),
            horizontal=True,
            key="schedule_source_mode",
        )
        if src_mode == "Existing files on disk":
            binge_opts = [p for p in candidates if "GRIDS" not in p.name.upper()]
            grids_opts = [p for p in candidates if "GRIDS" in p.name.upper()]
            if not binge_opts or not grids_opts:
                st.warning("Could not find both BINGE and BINGE GRIDS files under `data/` or `out/`.")
            else:
                def _fmt_path(p: Path) -> str:
                    return str(p).replace("\\", "/")

                default_binge = next((p for p in binge_opts if "BINGE" in p.name.upper()), binge_opts[0])
                default_grids = next((p for p in grids_opts if "BINGE" in p.name.upper()), grids_opts[0])
                pick_binge = st.selectbox(
                    "BINGE workbook",
                    binge_opts,
                    index=binge_opts.index(default_binge),
                    format_func=_fmt_path,
                    key="schedule_pick_binge",
                )
                pick_grids = st.selectbox(
                    "BINGE GRIDS workbook",
                    grids_opts,
                    index=grids_opts.index(default_grids),
                    format_func=_fmt_path,
                    key="schedule_pick_grids",
                )
                if st.button("Load selected schedule", use_container_width=True, key="schedule_load_existing"):
                    st.session_state["binge_path"] = Path(pick_binge)
                    st.session_state["grids_path"] = Path(pick_grids)
                    st.session_state["out_dir"] = Path(pick_binge).parent
                    st.rerun()
    else:
        st.caption("No existing workbook files found under `data/` or `out/`.")

    completed = _load_completed_months(cfg_path)
    if completed:
        st.caption(
            f"Months marked built in-app: **{', '.join(sorted(completed))}** "
            f"(see `schedule_build_state.json` or legacy `playlist_build_state.json`)."
        )

    if "binge_path" not in st.session_state:
        st.info("Nothing generated yet — go to **Create Schedule** and run it.")
    else:
        st.markdown("##### Latest files")
        _render_last_build_outputs(cfg, cfg_path)
        _render_binge_grids_preview(
            key_prefix="schedule",
            show_swap=True,
            cfg=cfg,
            cfg_path=cfg_path,
            nikki_path=nikki_path,
        )

    st.divider()
    st.markdown("##### Make changes")
    st.caption(
        "**Edit schedules** in your sources: episodes, order, and show keys live in the base schedule YAML and Nikki spreadsheet — "
        "not only inside the export files. Edit those, then run **Create Schedule** again."
    )
    setup_abs = cfg_path.resolve()
    st.markdown(f"- **Base schedule (YAML):** `{setup_abs}`")
    st.markdown(f"- **Content workbook:** `{nikki_path.resolve()}`")
    cur = resolved_cursor_state_path(cfg)
    if cur:
        st.markdown(f"- **Episode cursors:** `{cur}`")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Open base schedule folder", use_container_width=True, key="pl_open_cfg"):
            err = _open_folder(setup_abs.parent)
            if err and "download buttons" not in err:
                st.error(err)
            elif err:
                st.info(err)
    with c2:
        if nikki_path.is_file() and st.button("Open Nikki folder", use_container_width=True, key="pl_open_nikki"):
            err = _open_folder(nikki_path.parent)
            if err and "download buttons" not in err:
                st.error(err)
            elif err:
                st.info(err)
        elif not nikki_path.is_file():
            st.caption("Nikki path missing — fix **nikki_workbook** in the base schedule file.")


def _render_build_schedule(cfg, cfg_path: Path, nikki: Path) -> None:
    if not nikki.is_file():
        st.error(
            f"Spreadsheet file not found:\n`{nikki}`\n\n"
            f"Edit **nikki_workbook** in `{cfg_path.name}`."
        )
        return

    if not cfg.weeks:
        st.error("No **weeks** in your base schedule file — add week lines or use another base schedule file.")
        return

    months_all = _months_for_build_selector(cfg.weeks)
    if not months_all:
        st.error("No weeks with valid dates in your base schedule file.")
        return

    pipeline = _pipeline_months(months_all, cfg.build_sequence_start)
    if not pipeline:
        st.error(
            "No months left in the build sequence — check **weeks** dates and **build_sequence_start** in your base schedule."
        )
        return

    completed = _load_completed_months(cfg_path)
    unlocked = _unlocked_months(pipeline, completed)
    if not unlocked:
        st.error("Could not determine which month to build — check **weeks** in your base schedule.")
        return

    buildable_weeks = _weeks_for_unlocked_months(cfg.weeks, unlocked)
    floor_d = _week_floor_from_reference_cutoff(cfg)
    if floor_d is not None:
        buildable_weeks = [w for w in buildable_weeks if parse_monday(w.monday) >= floor_d]
    if not buildable_weeks:
        st.error("No weeks are currently unlocked to build from your reference cutoff.")
        return

    base_window_start, base_anchor = _baseline_window_for_cfg(cfg)
    st.caption(
        f"Baseline last airdate: **{_friendly_date(base_anchor)}** · "
        f"rolling window: **{_friendly_date(base_window_start)} -> {_friendly_date(base_anchor)}**"
    )

    prev_m = st.session_state.get("_build_month_iso")
    cur_m = parse_monday(buildable_weeks[0].monday).isoformat()
    if prev_m is not None and prev_m != cur_m:
        for k in ("binge_path", "grids_path", "out_dir"):
            st.session_state.pop(k, None)
    st.session_state["_build_month_iso"] = cur_m

    min_day = parse_monday(buildable_weeks[0].monday)
    max_day = parse_monday(buildable_weeks[-1].monday) + timedelta(days=6)
    default_start = parse_monday(buildable_weeks[0].monday)
    pending_start = st.session_state.pop("_pending_schedule_start_date", None)
    if isinstance(pending_start, date):
        if pending_start < min_day:
            pending_start = min_day
        elif pending_start > max_day:
            pending_start = max_day
        st.session_state["schedule_start_date"] = pending_start
    pending_weeks = st.session_state.pop("_pending_schedule_week_count", None)
    if isinstance(pending_weeks, int) and pending_weeks >= 1:
        st.session_state["schedule_week_count"] = pending_weeks
    auto_advance_msg = st.session_state.pop("_pending_auto_advance_msg", "")
    start_date = st.date_input(
        "Start date",
        value=default_start,
        min_value=min_day,
        max_value=max_day,
        key="schedule_start_date",
        help="First calendar week to build: schedules start at the Monday on or after this date.",
    )
    all_from_start = _effective_weeks_from_start(buildable_weeks, start_date, len(buildable_weeks))
    if not all_from_start:
        st.warning("No weeks available from that start date.")
        return
    default_weeks = 1
    week_count = int(
        st.number_input(
            "How many weeks",
            min_value=1,
            max_value=len(all_from_start),
            value=default_weeks,
            step=1,
            key="schedule_week_count",
        )
    )
    if auto_advance_msg:
        st.info(auto_advance_msg)
    selected_weeks = all_from_start[:week_count]
    selected_mondays = [w.monday for w in selected_weeks]
    st.caption(
        f"Selected weeks: **{len(selected_weeks)}** · "
        + ", ".join(f"`{m}`" for m in selected_mondays)
    )
    # Reset change pickers when build window changes, so OTO/mass can never carry stale weeks.
    scope_key = f"{start_date.isoformat()}|{week_count}|{'|'.join(selected_mondays)}"
    if st.session_state.get("_build_scope_key") != scope_key:
        for k in (
            "build_oto_slot_ids",
            "build_mass_seed_ids",
            "build_oto_preview_week",
            "build_oto_preview_day",
        ):
            st.session_state.pop(k, None)
        # Clear per-day editor/multiselect widget state tied to the old build scope.
        for k in list(st.session_state.keys()):
            if (
                k.startswith("build_oto_slot_editor_")
                or k.startswith("build_oto_day_multi_")
                or k.startswith("build_oto_pick_slot_")
                or k.startswith("build_oto_movie_for_")
                or k.startswith("build_oto_movie_seq_")
                or k.startswith("build_oto_movie_seq_len_")
                or k.startswith("build_oto_movie_same_")
                or k.startswith("build_oto_runtime_")
                or k.startswith("build_oto_auto_movie_seq_")
                or k.startswith("build_oto_auto_show_seq_")
                or k.startswith("build_oto_auto_show_seq_len_")
            ):
                st.session_state.pop(k, None)
        st.session_state["_build_scope_key"] = scope_key

    template_slots, template_warnings = _schedule_template_slots(selected_weeks)
    for wmsg in template_warnings:
        st.warning(wmsg)
    slot_by_id = {r["slot_id"]: r for r in template_slots}
    slot_ids = [r["slot_id"] for r in template_slots]

    yaml_keys = sorted(cfg.shows.keys(), key=lambda k: cfg.shows[k].display_name.lower())
    extra_tab_names: list[str] = []
    if nikki.is_file():
        tabs = _nikki_workbook_sheet_names(str(nikki.resolve()), _nikki_mtime(nikki))
        extra_tab_names = workbook_tabs_not_in_yaml(cfg, tabs)
    extra_opts = [workbook_tab_option(t) for t in extra_tab_names]
    archive_options = yaml_keys + extra_opts
    current_show_options = sorted({str(r["show"]).strip() for r in template_slots if str(r["show"]).strip()})

    def _archive_pick_label(opt: str) -> str:
        raw_literal = _parse_literal_text_option(opt)
        if raw_literal is not None:
            return f"{raw_literal} _(from schedule)_"
        tab = parse_workbook_tab_option(opt)
        if tab is not None:
            return f"{tab} _(not on schedule)_"
        return cfg.shows[opt].display_name

    st.markdown("##### Optional schedule changes")
    st.caption("OTO/mass block pickers are limited to the selected weeks above.")
    use_oto = st.checkbox(
        "Apply OTO (one-time-only) changes for this schedule",
        key="build_use_oto_changes",
    )
    oto_rows: list[dict[str, Any]] = []
    oto_pick: Optional[str] = None
    oto_source_group = ""
    oto_source_keys: set[str] = set()
    oto_fill_mode = "Auto-populate matching genre show"
    oto_fill_choice = "auto_show"
    oto_manual_replace_kind = "show"
    oto_episode_rows: list[dict[str, Any]] = []
    oto_manual_pool: list[dict[str, Any]] = []
    oto_manual_start_idx: Optional[int] = None
    oto_manual_advance = True
    oto_auto_show_by_slot: dict[str, str] = {}
    oto_auto_show_plan_texts: list[str] = []
    oto_movie_by_slot: dict[str, str] = {}
    oto_auto_movie_by_slot: dict[str, str] = {}
    oto_auto_movie_plan_texts: list[str] = []
    oto_movie_fill_rule = "Sequential by slot"
    oto_timing_notes: list[str] = []
    if use_oto:
        if not slot_ids:
            st.warning("No editable schedule blocks found in the selected weeks.")
        else:
            oto_mode = st.radio(
                "OTO mode",
                ("Choose time blocks", "Replace all blocks for one current show"),
                horizontal=True,
                key="build_oto_mode",
            )
            if oto_mode == "Replace all blocks for one current show":
                if not current_show_options:
                    st.warning("No current shows were found in the selected weeks.")
                else:
                    oto_source_show = st.selectbox(
                        "OTO current show to replace",
                        current_show_options,
                        key="build_oto_source_show",
                    )
                    oto_rows = [r for r in template_slots if r["show"] == oto_source_show]
                    st.caption(f"Selected **{len(oto_rows)}** block(s) for `{oto_source_show}`.")
            else:
                picker_ui = st.radio(
                    "OTO block picker",
                    ("Clickable schedule preview", "Quick list picker"),
                    horizontal=True,
                    key="build_oto_picker_ui",
                )
                if picker_ui == "Clickable schedule preview":
                    week_opts = sorted({str(r["week_monday"]) for r in template_slots})
                    preview_week = st.selectbox(
                        "Preview week",
                        week_opts,
                        key="build_oto_preview_week",
                    )
                    prior_selected = {
                        sid for sid in st.session_state.get("build_oto_slot_ids", []) if sid in slot_by_id
                    }
                    week_rows = [r for r in template_slots if str(r["week_monday"]) == str(preview_week)]
                    week_rows.sort(key=lambda r: (r["date_iso"], int(r["start_slot"])))
                    if week_rows:
                        day_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
                        week_days = sorted({int(r["day_index"]) for r in week_rows})
                        day_opts = [day_labels[i] for i in week_days if 0 <= i < len(day_labels)]
                        day_pick = st.radio(
                            "Preview day",
                            day_opts,
                            horizontal=True,
                            key="build_oto_preview_day",
                        )
                        day_index = day_labels.index(day_pick)
                        day_rows = [r for r in week_rows if int(r["day_index"]) == day_index]
                        st.caption(
                            "Click rows to select blocks. Selections persist across day/week changes."
                        )
                        day_slot_ids = [str(r["slot_id"]) for r in day_rows]
                        day_selected_ids: set[str] = set()
                        for r in day_rows:
                            sid = str(r["slot_id"])
                            cb_key = f"build_oto_pick_slot_{sid}"
                            if cb_key not in st.session_state:
                                st.session_state[cb_key] = sid in prior_selected
                            label = (
                                f"{r['start']}-{r['finish']} ({r['duration_label']}) · "
                                f"{str(r['show'])} · {r['date_iso']}"
                            )
                            if st.checkbox(label, key=cb_key):
                                day_selected_ids.add(sid)
                        merged_selected = (prior_selected - set(day_slot_ids)) | day_selected_ids
                        st.session_state["build_oto_slot_ids"] = [sid for sid in slot_ids if sid in merged_selected]
                        oto_ids = st.session_state["build_oto_slot_ids"]
                    else:
                        oto_ids = []
                else:
                    oto_ids = st.multiselect(
                        "OTO: choose time blocks",
                        slot_ids,
                        format_func=lambda sid: _slot_picker_label(slot_by_id[sid]),
                        key="build_oto_slot_ids",
                    )
                oto_ids = [sid for sid in oto_ids if sid in slot_by_id]
                oto_rows = [slot_by_id[sid] for sid in oto_ids]
                st.caption(f"Selected **{len(oto_rows)}** block(s) for OTO.")
            if oto_rows:
                oto_source_keys = {
                    k for k in (_slot_source_show_key(cfg, r["show"]) for r in oto_rows) if k is not None
                }
                src_groups = [g for g in (_semantic_group_for_show(cfg, k) for k in oto_source_keys) if g]
                if src_groups:
                    oto_source_group = sorted(src_groups)[0]
                    st.caption(f"Semantic source group: `{oto_source_group}`")
            oto_fill_choice = st.radio(
                "OTO swap choice",
                ("auto_show", "auto_movie", "manual_replace"),
                format_func=lambda v: {
                    "auto_show": "Auto-swap with shows",
                    "auto_movie": "Auto-swap with movies",
                    "manual_replace": "Manual replace",
                }.get(str(v), str(v)),
                horizontal=True,
                key="build_oto_fill_mode",
            )
            if oto_fill_choice == "manual_replace":
                oto_manual_replace_kind = st.radio(
                    "Manual replace with",
                    ("show", "movie"),
                    format_func=lambda v: "Show (show > season > episode)" if v == "show" else "Movie/program",
                    horizontal=True,
                    key="build_oto_manual_replace_kind",
                )
                oto_fill_mode = (
                    "Manual: show > season > episode"
                    if oto_manual_replace_kind == "show"
                    else "Replace time window with movie list"
                )
            elif oto_fill_choice == "auto_show":
                oto_fill_mode = "Auto-populate matching genre show"
            else:
                oto_fill_mode = "Auto-populate matching genre movie/program"

            if oto_fill_mode == "Manual: show > season > episode":
                if archive_options:
                    oto_pick = st.selectbox(
                        "Manual show",
                        archive_options,
                        format_func=_archive_pick_label,
                        key="build_oto_pick",
                    )
                if oto_pick:
                    oto_episode_rows = _episode_rows_for_archive_pick(cfg, oto_pick, nikki)
                    seasons_found = sorted({r["season"] for r in oto_episode_rows if r.get("season") is not None})
                    season_opts: list[str] = ["All seasons"] + [f"Season {s}" for s in seasons_found]
                    if any(r.get("season") is None for r in oto_episode_rows) and seasons_found:
                        season_opts.append("Unnumbered / list order")
                    pick_season = st.selectbox("Season", season_opts, key="build_oto_pick_season")

                    def _manual_season_match(r: dict[str, Any]) -> bool:
                        if pick_season == "All seasons":
                            return True
                        if pick_season.startswith("Season "):
                            return r.get("season") == int(pick_season.replace("Season ", ""))
                        if pick_season == "Unnumbered / list order":
                            return r.get("season") is None
                        return True

                    oto_manual_pool = [r for r in oto_episode_rows if _manual_season_match(r)]
                    if not oto_manual_pool:
                        st.warning("No episodes match the chosen season filter.")
                    else:
                        idx_opts = list(range(len(oto_manual_pool)))
                        oto_manual_start_idx = st.selectbox(
                            "Episode",
                            idx_opts,
                            format_func=lambda i: (
                                f"{oto_manual_pool[i].get('se_compact', '—')}  {oto_manual_pool[i].get('code', '—')}  —  "
                                f"{str(oto_manual_pool[i].get('title', ''))[:140]}"
                            ),
                            key="build_oto_manual_episode_idx",
                        )
                        oto_manual_advance = st.checkbox(
                            "Advance episodes across selected OTO blocks",
                            value=True,
                            key="build_oto_manual_advance",
                        )
            elif oto_fill_mode == "Auto-populate matching genre show":
                auto_show_opts = _semantic_candidates(
                    cfg,
                    group=oto_source_group,
                    kind="series",
                    exclude_keys=oto_source_keys,
                )
                if not auto_show_opts:
                    auto_show_opts = _semantic_candidates(cfg, group="", kind="series", exclude_keys=oto_source_keys)
                if auto_show_opts:
                    series_archive_options = [k for k in yaml_keys if cfg.shows[k].kind == "series"] + extra_opts
                    if not oto_rows:
                        oto_pick = auto_show_opts[0]
                        pick_group = getattr(cfg.shows[oto_pick], "semantic_group", None) or "unlabeled"
                        st.caption(
                            f"Auto-picked matching genre show: **{cfg.shows[oto_pick].display_name}** (`{pick_group}`)."
                        )
                        oto_episode_rows = _episode_rows_for_archive_pick(cfg, oto_pick, nikki)
                    else:
                        ordered_rows = sorted(oto_rows, key=lambda r: (r["date_iso"], int(r["start_slot"])))
                        by_date: dict[str, list[dict[str, Any]]] = {}
                        for r in ordered_rows:
                            by_date.setdefault(str(r["date_iso"]), []).append(r)
                        tab_dates = list(by_date.keys())
                        date_tabs = st.tabs(tab_dates)
                        for i, d in enumerate(tab_dates):
                            with date_tabs[i]:
                                rows_for_day = by_date[d]
                                st.caption(f"Auto show plan for `{d}` (editable):")
                                max_seq = max(1, min(8, len(rows_for_day), len(series_archive_options)))
                                default_seq = max(1, min(max_seq, max(1, len(rows_for_day) // 4)))
                                seq_len = int(
                                    st.number_input(
                                        "How many shows in sequence?",
                                        min_value=1,
                                        max_value=max_seq,
                                        value=default_seq,
                                        step=1,
                                        key=f"build_oto_auto_show_seq_len_{d}",
                                    )
                                )
                                seq_opts: list[str] = []
                                for j in range(seq_len):
                                    default_opt = auto_show_opts[j % len(auto_show_opts)]
                                    try:
                                        default_idx = series_archive_options.index(default_opt)
                                    except ValueError:
                                        default_idx = 0
                                    edited_opt = st.selectbox(
                                        f"{d} selected show {j + 1}",
                                        series_archive_options,
                                        index=default_idx,
                                        format_func=_archive_pick_label,
                                        key=f"build_oto_auto_show_seq_{d}_{j}",
                                    )
                                    seq_opts.append(edited_opt)
                                for r_idx, r in enumerate(rows_for_day):
                                    sid = str(r["slot_id"])
                                    oto_auto_show_by_slot[sid] = seq_opts[r_idx % len(seq_opts)]
                                oto_auto_show_plan_texts.append(
                                    f"{d}: " + " -> ".join(_display_name_for_archive_pick(cfg, opt) for opt in seq_opts)
                                )
                        if oto_auto_show_plan_texts:
                            st.caption("Auto show plan: " + " | ".join(oto_auto_show_plan_texts))
                        oto_pick = auto_show_opts[0]
                        oto_episode_rows = _episode_rows_for_archive_pick(cfg, oto_pick, nikki)
                else:
                    st.warning("No related series candidates were found for auto-populate.")
            elif oto_fill_mode == "Replace time window with movie list":
                movie_opts = _movie_program_picker_options(cfg, extra_tab_names, template_slots, nikki)
                st.caption("Swap by day: pick an ordered movie/program sequence once, then auto-fill selected blocks.")
                if movie_opts:
                    if not oto_rows:
                        st.info("Select one or more OTO blocks first, then assign a movie/program to each block.")
                    else:
                        runtime_map = _movie_runtime_minutes(str(cfg_path.resolve()))
                        commercials_pct = float(
                            st.number_input(
                                "Estimated commercial %",
                                min_value=0.0,
                                max_value=90.0,
                                value=30.0,
                                step=5.0,
                                key="build_oto_runtime_commercial_pct",
                            )
                        )
                        fallback_runtime = int(
                            st.number_input(
                                "Fallback runtime for unknown movies (minutes)",
                                min_value=30,
                                max_value=240,
                                value=95,
                                step=5,
                                key="build_oto_runtime_fallback",
                            )
                        )
                        ordered_rows = sorted(oto_rows, key=lambda r: (r["date_iso"], int(r["start_slot"])))
                        by_date: dict[str, list[dict[str, Any]]] = {}
                        for r in ordered_rows:
                            by_date.setdefault(str(r["date_iso"]), []).append(r)
                        tab_dates = list(by_date.keys())
                        date_tabs = st.tabs(tab_dates)
                        oto_movie_fill_rule = st.radio(
                            "Fill rule",
                            ("Sequential by slot", "Equal chunks"),
                            horizontal=True,
                            key="build_oto_movie_fill_rule",
                        )
                        assign_mode = st.radio(
                            "Assignment mode",
                            ("Program sequence (recommended)", "Per-slot (advanced)"),
                            horizontal=True,
                            key="build_oto_movie_assign_mode",
                        )
                        none_opt = "__none__"
                        slot_opts = [none_opt] + movie_opts

                        for i, d in enumerate(tab_dates):
                            with date_tabs[i]:
                                rows_for_day = by_date[d]
                                st.caption(f"{len(rows_for_day)} selected block(s) on `{d}`.")
                                if assign_mode == "Program sequence (recommended)":
                                    max_seq = max(1, min(12, len(rows_for_day)))
                                    seq_len = int(
                                        st.number_input(
                                            "How many items in the sequence?",
                                            min_value=1,
                                            max_value=max_seq,
                                            value=1,
                                            step=1,
                                            key=f"build_oto_movie_seq_len_{d}",
                                        )
                                    )
                                    seq_opts: list[str] = []
                                    for j in range(seq_len):
                                        picked = st.selectbox(
                                            f"Sequence item {j + 1}",
                                            slot_opts,
                                            format_func=lambda opt: "— Select movie/program —"
                                            if opt == none_opt
                                            else _archive_pick_label(opt),
                                            key=f"build_oto_movie_seq_{d}_{j}",
                                        )
                                        if picked != none_opt:
                                            seq_opts.append(picked)
                                    if seq_opts:
                                        runtime_mode = st.checkbox(
                                            "Use runtime-aware fit",
                                            value=True,
                                            key=f"build_oto_runtime_mode_{d}",
                                        )
                                        if runtime_mode:
                                            content_per_slot = 30.0 * max(0.05, (1.0 - commercials_pct / 100.0))
                                            raw_weights: list[float] = []
                                            runtime_notes: list[str] = []
                                            for pick in seq_opts:
                                                rt = _runtime_for_archive_option(cfg, pick, runtime_map)
                                                if rt is None:
                                                    rt = fallback_runtime
                                                    runtime_notes.append(
                                                        f"{_display_name_for_archive_pick(cfg, pick)}: using fallback {fallback_runtime}m"
                                                    )
                                                raw_weights.append(max(1.0, rt / content_per_slot))
                                            total_weight = sum(raw_weights) or float(len(seq_opts))
                                            scaled = [w * len(rows_for_day) / total_weight for w in raw_weights]
                                            counts = [max(1, int(v)) for v in scaled]
                                            while sum(counts) > len(rows_for_day):
                                                i_max = max(range(len(counts)), key=lambda i: counts[i])
                                                if counts[i_max] > 1:
                                                    counts[i_max] -= 1
                                                else:
                                                    break
                                            while sum(counts) < len(rows_for_day):
                                                frac = [scaled[i] - int(scaled[i]) for i in range(len(scaled))]
                                                i_best = max(range(len(counts)), key=lambda i: frac[i])
                                                counts[i_best] += 1
                                            idx = 0
                                            for i_pick, pick in enumerate(seq_opts):
                                                for _ in range(counts[i_pick]):
                                                    if idx >= len(rows_for_day):
                                                        break
                                                    sid = str(rows_for_day[idx]["slot_id"])
                                                    oto_movie_by_slot[sid] = pick
                                                    idx += 1
                                            if runtime_notes:
                                                st.caption("Runtime notes: " + " | ".join(runtime_notes))
                                        else:
                                            if oto_movie_fill_rule == "Equal chunks":
                                                per = len(rows_for_day) // len(seq_opts)
                                                rem = len(rows_for_day) % len(seq_opts)
                                                idx = 0
                                                for m_idx, pick in enumerate(seq_opts):
                                                    repeat = per + (1 if m_idx < rem else 0)
                                                    for _ in range(repeat):
                                                        if idx >= len(rows_for_day):
                                                            break
                                                        sid = str(rows_for_day[idx]["slot_id"])
                                                        oto_movie_by_slot[sid] = pick
                                                        idx += 1
                                            else:
                                                for r_idx, r in enumerate(rows_for_day):
                                                    sid = str(r["slot_id"])
                                                    oto_movie_by_slot[sid] = seq_opts[r_idx % len(seq_opts)]
                                else:
                                    for r in rows_for_day:
                                        sid = str(r["slot_id"])
                                        pick_key = f"build_oto_movie_for_{sid}"
                                        label = (
                                            f"{r['start']}-{r['finish']} ({r['duration_label']}) · "
                                            f"{str(r['show'])}"
                                        )
                                        picked = st.selectbox(
                                            label,
                                            slot_opts,
                                            format_func=lambda opt: "— Select movie/program —"
                                            if opt == none_opt
                                            else _archive_pick_label(opt),
                                            key=pick_key,
                                        )
                                        if picked != none_opt:
                                            oto_movie_by_slot[sid] = picked
                                assigned_day = sum(
                                    1 for r in rows_for_day if oto_movie_by_slot.get(str(r["slot_id"]))
                                )
                                timing_notes = _runtime_timing_notes_for_day(
                                    cfg=cfg,
                                    day_iso=d,
                                    rows_for_day=rows_for_day,
                                    assigned_by_slot=oto_movie_by_slot,
                                    runtime_map=runtime_map,
                                    fallback_runtime=fallback_runtime,
                                    commercials_pct=commercials_pct,
                                )
                                oto_timing_notes.extend(timing_notes)
                                st.caption(
                                    f"Assigned **{assigned_day}/{len(rows_for_day)}** block(s) for `{d}`."
                                )
                else:
                    st.warning("No movie/program entries were found in the archive list.")
            else:
                full_movie_opts = _movie_program_picker_options(cfg, extra_tab_names, template_slots, nikki)
                auto_movie_opts = _auto_movie_candidates(
                    cfg,
                    cfg_path,
                    extra_tab_names,
                    template_slots,
                    nikki,
                    oto_source_group,
                )
                if auto_movie_opts:
                    runtime_map = _movie_runtime_minutes(str(cfg_path.resolve()))
                    commercials_pct = 30.0
                    fallback_runtime = 95
                    edit_movie_opts = full_movie_opts if full_movie_opts else auto_movie_opts
                    by_date: dict[str, list[dict[str, Any]]] = {}
                    for r in sorted(oto_rows, key=lambda rr: (rr["date_iso"], int(rr["start_slot"]))):
                        by_date.setdefault(str(r["date_iso"]), []).append(r)

                    def _build_auto_plan(rows_for_day: list[dict[str, Any]]) -> list[str]:
                        if not rows_for_day:
                            return []
                        content_per_slot = 30.0 * max(0.05, (1.0 - commercials_pct / 100.0))
                        target_content = len(rows_for_day) * content_per_slot
                        selected: list[str] = []
                        total_rt = 0
                        for opt in auto_movie_opts:
                            rt = _runtime_for_archive_option(cfg, opt, runtime_map) or fallback_runtime
                            selected.append(opt)
                            total_rt += int(rt)
                            # Keep adding titles until content estimate fills the selected window.
                            if total_rt >= target_content and len(selected) >= 1:
                                break
                        if len(rows_for_day) >= 6 and len(selected) == 1 and len(auto_movie_opts) > 1:
                            selected.append(auto_movie_opts[1])
                        return selected

                    for d, rows_for_day in by_date.items():
                        seq_opts = _build_auto_plan(rows_for_day)
                        if not seq_opts:
                            continue
                        st.caption(f"Auto plan for `{d}` (editable):")
                        editable_seq: list[str] = []
                        for i_seq, default_opt in enumerate(seq_opts):
                            try:
                                default_idx = edit_movie_opts.index(default_opt)
                            except ValueError:
                                default_idx = 0
                            edited_opt = st.selectbox(
                                f"{d} selected movie {i_seq + 1}",
                                edit_movie_opts,
                                index=default_idx,
                                format_func=lambda opt: _archive_pick_label(opt),
                                key=f"build_oto_auto_movie_seq_{d}_{i_seq}",
                            )
                            editable_seq.append(edited_opt)
                        seq_opts = editable_seq
                        content_per_slot = 30.0 * max(0.05, (1.0 - commercials_pct / 100.0))
                        raw_weights = [
                            max(1.0, float(_runtime_for_archive_option(cfg, opt, runtime_map) or fallback_runtime) / content_per_slot)
                            for opt in seq_opts
                        ]
                        total_weight = sum(raw_weights) or float(len(seq_opts))
                        scaled = [w * len(rows_for_day) / total_weight for w in raw_weights]
                        counts = [max(1, int(v)) for v in scaled]
                        while sum(counts) > len(rows_for_day):
                            i_max = max(range(len(counts)), key=lambda i: counts[i])
                            if counts[i_max] > 1:
                                counts[i_max] -= 1
                            else:
                                break
                        while sum(counts) < len(rows_for_day):
                            frac = [scaled[i] - int(scaled[i]) for i in range(len(scaled))]
                            i_best = max(range(len(counts)), key=lambda i: frac[i])
                            counts[i_best] += 1

                        idx = 0
                        for i_pick, opt in enumerate(seq_opts):
                            for _ in range(counts[i_pick]):
                                if idx >= len(rows_for_day):
                                    break
                                sid = str(rows_for_day[idx]["slot_id"])
                                oto_auto_movie_by_slot[sid] = opt
                                idx += 1
                        oto_timing_notes.extend(
                            _runtime_timing_notes_for_day(
                                cfg=cfg,
                                day_iso=d,
                                rows_for_day=rows_for_day,
                                assigned_by_slot=oto_auto_movie_by_slot,
                                runtime_map=runtime_map,
                                fallback_runtime=fallback_runtime,
                                commercials_pct=commercials_pct,
                            )
                        )
                        oto_auto_movie_plan_texts.append(
                            f"{d}: " + " -> ".join(_display_name_for_archive_pick(cfg, opt) for opt in seq_opts)
                        )
                    if oto_auto_movie_plan_texts:
                        st.caption("Auto movie plan: " + " | ".join(oto_auto_movie_plan_texts))
                    if auto_movie_opts:
                        oto_pick = auto_movie_opts[0]
                else:
                    st.warning("No related movie/program candidates were found for auto-populate.")

    use_mass = st.checkbox(
        "Apply mass changes and persist to source schedule files",
        key="build_use_mass_changes",
    )
    mass_rows: list[dict[str, Any]] = []
    mass_pick: Optional[str] = None
    mass_fill_mode = "Auto-populate matching genre show"
    mass_source_group = ""
    mass_source_keys: set[str] = set()
    if use_mass:
        if not slot_ids:
            st.warning("No editable schedule blocks found in the selected weeks.")
        else:
            mass_mode = st.radio(
                "Mass mode",
                (
                    "Replace all blocks for one current show (selected weeks)",
                    "Choose seed time blocks + expand pattern",
                ),
                horizontal=False,
                key="build_mass_mode",
            )
            if mass_mode == "Replace all blocks for one current show (selected weeks)":
                if not current_show_options:
                    st.warning("No current shows were found in the selected weeks.")
                else:
                    mass_source_show = st.selectbox(
                        "Mass current show to replace",
                        current_show_options,
                        key="build_mass_source_show",
                    )
                    mass_rows = [r for r in template_slots if r["show"] == mass_source_show]
                    st.caption(f"Selected **{len(mass_rows)}** block(s) for `{mass_source_show}`.")
            else:
                base_ids = st.multiselect(
                    "Mass: choose seed time blocks",
                    slot_ids,
                    format_func=lambda sid: _slot_picker_label(slot_by_id[sid]),
                    key="build_mass_seed_ids",
                )
                base_ids = [sid for sid in base_ids if sid in slot_by_id]
                expand_pattern = st.checkbox(
                    "Expand by pattern (same weekday + start slot + current show)",
                    value=True,
                    key="build_mass_expand_pattern",
                )
                if expand_pattern and base_ids:
                    expanded: set[str] = set()
                    for sid in base_ids:
                        src = slot_by_id.get(sid)
                        if not src:
                            continue
                        for row in template_slots:
                            if (
                                row["day_index"] == src["day_index"]
                                and row["start_slot"] == src["start_slot"]
                                and row["show"].casefold() == src["show"].casefold()
                            ):
                                expanded.add(row["slot_id"])
                    mass_rows = [slot_by_id[sid] for sid in sorted(expanded)]
                else:
                    mass_rows = [slot_by_id[sid] for sid in base_ids if sid in slot_by_id]
            if mass_rows:
                mass_source_keys = {
                    k for k in (_slot_source_show_key(cfg, r["show"]) for r in mass_rows) if k is not None
                }
                src_groups = [g for g in (_semantic_group_for_show(cfg, k) for k in mass_source_keys) if g]
                if src_groups:
                    mass_source_group = sorted(src_groups)[0]
                    st.caption(f"Mass semantic source group: `{mass_source_group}`")

            mass_fill_mode = st.radio(
                "Mass fill mode",
                (
                    "Auto-populate matching genre show",
                    "Auto-populate matching genre movie/program",
                    "Manual pick replacement show",
                ),
                horizontal=True,
                key="build_mass_fill_mode",
            )
            if mass_fill_mode == "Manual pick replacement show":
                if archive_options:
                    mass_pick = st.selectbox(
                        "Manual replacement show",
                        archive_options,
                        format_func=_archive_pick_label,
                        key="build_mass_pick",
                    )
            elif mass_fill_mode == "Auto-populate matching genre show":
                auto_mass_show_opts = _semantic_candidates(
                    cfg,
                    group=mass_source_group,
                    kind="series",
                    exclude_keys=mass_source_keys,
                )
                if not auto_mass_show_opts:
                    auto_mass_show_opts = _semantic_candidates(
                        cfg, group="", kind="series", exclude_keys=mass_source_keys
                    )
                if auto_mass_show_opts:
                    mass_pick = auto_mass_show_opts[0]
                    pick_group = getattr(cfg.shows[mass_pick], "semantic_group", None) or "unlabeled"
                    st.caption(
                        f"Auto-picked matching genre show: **{cfg.shows[mass_pick].display_name}** (`{pick_group}`)."
                    )
                else:
                    st.warning("No related series candidates were found for mass auto-populate.")
            else:
                auto_mass_movie_opts = _semantic_candidates(
                    cfg,
                    group=mass_source_group,
                    kind="literal",
                    exclude_keys=set(),
                )
                if not auto_mass_movie_opts:
                    auto_mass_movie_opts = _semantic_candidates(cfg, group="", kind="literal", exclude_keys=set())
                if auto_mass_movie_opts:
                    mass_pick = auto_mass_movie_opts[0]
                    pick_group = getattr(cfg.shows[mass_pick], "semantic_group", None) or "unlabeled"
                    st.caption(
                        f"Auto-picked matching genre movie/program: **{cfg.shows[mass_pick].display_name}** (`{pick_group}`)."
                    )
                else:
                    st.warning("No related movie/program candidates were found for mass auto-populate.")
            st.checkbox(
                "I understand mass changes persist to source files.",
                key="build_mass_confirm",
            )

    st.markdown("##### Preview schedule")
    st.caption("This run uses the selected start date/week count, plus optional OTO/mass changes below.")
    st.markdown(
        "\n".join(
            [
                f"- Start date: **{start_date.isoformat()}**",
                f"- Duration: **{len(selected_weeks)}** week(s)",
                f"- Week starts: {', '.join(f'`{w.monday}`' for w in selected_weeks)}",
            ]
        )
    )
    if use_oto:
        oto_dur = _format_duration_minutes(sum(int(r["duration_minutes"]) for r in oto_rows)) if oto_rows else "0m"
        oto_mode_label = {
            "Auto-populate matching genre show": "Auto-swap with shows",
            "Auto-populate matching genre movie/program": "Auto-swap with movies",
            "Manual: show > season > episode": "Manual replace (show)",
            "Replace time window with movie list": "Manual replace (movie/program)",
        }.get(oto_fill_mode, oto_fill_mode)
        if oto_fill_mode == "Replace time window with movie list":
            assigned = [oto_movie_by_slot.get(str(r["slot_id"])) for r in sorted(oto_rows, key=lambda r: (r["date_iso"], int(r["start_slot"])))]
            assigned = [a for a in assigned if a]
            oto_name = f"{len(assigned)} assigned block(s)" if assigned else "—"
        elif oto_fill_mode == "Auto-populate matching genre show" and oto_auto_show_plan_texts:
            oto_name = f"{len(oto_auto_show_plan_texts)} day plan(s)"
        elif oto_fill_mode == "Auto-populate matching genre movie/program":
            oto_name = (
                f"{len(oto_auto_movie_plan_texts)} day plan(s)"
                if oto_auto_movie_plan_texts
                else (_display_name_for_archive_pick(cfg, oto_pick) if oto_pick else "—")
            )
        else:
            oto_name = _display_name_for_archive_pick(cfg, oto_pick) if oto_pick else "—"
        st.markdown(
            f"- Schedule: **{len(oto_rows)}** block(s), duration **{oto_dur}**, replacement **{oto_name}**, mode **{oto_mode_label}**"
        )
        if oto_timing_notes:
            st.markdown(f"- Schedule timing notes: **{len(oto_timing_notes)}**")
    else:
        st.markdown("- Schedule: none")
    if use_mass:
        mass_dur = _format_duration_minutes(sum(int(r["duration_minutes"]) for r in mass_rows)) if mass_rows else "0m"
        mass_name = _display_name_for_archive_pick(cfg, mass_pick) if mass_pick else "—"
        st.markdown(
            f"- Mass: **{len(mass_rows)}** block(s), duration **{mass_dur}**, replacement **{mass_name}** (persists)"
        )
    else:
        st.markdown("- Mass: none")

    preflight_issues: list[str] = []
    if use_oto and not oto_rows:
        preflight_issues.append("OTO is enabled but no blocks were selected.")
    if use_oto and oto_fill_mode != "Replace time window with movie list" and not oto_pick:
        preflight_issues.append("OTO is enabled but replacement show is missing.")
    if use_oto and oto_fill_mode == "Manual: show > season > episode":
        if not oto_manual_pool or oto_manual_start_idx is None:
            preflight_issues.append("OTO manual mode requires a season and episode selection.")
    if use_oto and oto_fill_mode == "Auto-populate matching genre show":
        if oto_rows:
            missing_auto_show_slots = [r for r in oto_rows if not oto_auto_show_by_slot.get(str(r["slot_id"]))]
            if missing_auto_show_slots and not oto_episode_rows:
                preflight_issues.append(
                    f"OTO auto-show mode could not plan replacements for {len(missing_auto_show_slots)} selected block(s)."
                )
        elif not oto_episode_rows:
            preflight_issues.append("OTO auto-related-show requires a series with parsed episode rows.")
    if use_oto and oto_fill_mode == "Auto-populate matching genre movie/program":
        missing_auto_movie_slots = [r for r in oto_rows if not oto_auto_movie_by_slot.get(str(r["slot_id"]))]
        if missing_auto_movie_slots:
            preflight_issues.append(
                f"OTO auto-movie mode could not plan replacements for {len(missing_auto_movie_slots)} selected block(s)."
            )
    if use_oto and oto_fill_mode == "Replace time window with movie list":
        if not oto_rows:
            preflight_issues.append("OTO movie-list mode requires selected time blocks.")
        missing_movie_slots = [r for r in oto_rows if not oto_movie_by_slot.get(str(r["slot_id"]))]
        if missing_movie_slots:
            preflight_issues.append(
                f"OTO movie-list mode requires a movie/program selection for each block ({len(missing_movie_slots)} missing)."
            )
    if use_mass and (not mass_rows or not mass_pick):
        preflight_issues.append("Mass is enabled but block selection and/or replacement show is missing.")
    if use_mass and not st.session_state.get("build_mass_confirm"):
        preflight_issues.append("Mass persistence confirmation is not checked.")
    for issue in preflight_issues:
        st.warning(issue)

    stations_input = st.text_input(
        "Stations (optional)",
        value="",
        placeholder="Comma-separated call letters, e.g. WXYZ, KABC — copies into subfolders under the output",
        key="export_stations_input",
    )

    run = st.button(
        "Create Schedule",
        type="primary",
        use_container_width=True,
        disabled=bool(preflight_issues),
        help=f"{len(selected_weeks)} selected week tab(s).",
    )

    if run:
        can_run = True
        oto_overrides: list[BingeRowOverride] = []
        oto_display = ""
        oto_grid_displays: list[str] = []
        if use_oto:
            if not oto_rows:
                st.error("OTO changes are enabled, but no schedule blocks are selected.")
                can_run = False
            elif oto_fill_mode != "Replace time window with movie list" and not oto_pick:
                st.error("OTO changes are enabled, but no replacement show was selected.")
                can_run = False
            else:
                ordered_oto_rows = sorted(oto_rows, key=lambda r: (r["date_iso"], int(r["start_slot"])))
                if oto_fill_mode == "Replace time window with movie list":
                    missing_rows = [r for r in ordered_oto_rows if not oto_movie_by_slot.get(str(r["slot_id"]))]
                    if missing_rows:
                        st.error(
                            f"OTO movie-list mode is missing replacements for {len(missing_rows)} selected block(s)."
                        )
                        can_run = False
                    else:
                        for row in ordered_oto_rows:
                            show_key = str(oto_movie_by_slot[str(row["slot_id"])])
                            oto_grid_displays.append(_display_name_for_archive_pick(cfg, show_key))
                elif oto_fill_mode == "Auto-populate matching genre show" and oto_auto_show_by_slot:
                    missing_rows = [r for r in ordered_oto_rows if not oto_auto_show_by_slot.get(str(r["slot_id"]))]
                    if missing_rows:
                        st.error(
                            f"OTO auto-show mode is missing replacements for {len(missing_rows)} selected block(s)."
                        )
                        can_run = False
                    else:
                        for row in ordered_oto_rows:
                            show_key = str(oto_auto_show_by_slot[str(row["slot_id"])])
                            oto_grid_displays.append(_display_name_for_archive_pick(cfg, show_key))
                elif oto_fill_mode == "Auto-populate matching genre movie/program":
                    missing_rows = [r for r in ordered_oto_rows if not oto_auto_movie_by_slot.get(str(r["slot_id"]))]
                    if missing_rows:
                        st.error(
                            f"OTO auto-movie mode is missing replacements for {len(missing_rows)} selected block(s)."
                        )
                        can_run = False
                    else:
                        for row in ordered_oto_rows:
                            show_key = str(oto_auto_movie_by_slot[str(row["slot_id"])])
                            oto_grid_displays.append(_display_name_for_archive_pick(cfg, show_key))
                else:
                    oto_display = _display_name_for_archive_pick(cfg, oto_pick)
                    oto_grid_displays = [oto_display] * len(ordered_oto_rows)
                # Build one replacement payload per selected slot using manual or auto fill mode.
                episode_plan: list[tuple[str, str, str, str]] = []
                if oto_fill_mode == "Auto-populate matching genre movie/program":
                    for title in oto_grid_displays:
                        episode_plan.append(("MOVIE", "MOVIE", title, title))
                elif oto_fill_mode == "Auto-populate matching genre show" and oto_auto_show_by_slot:
                    show_pool_cache: dict[str, list[dict[str, Any]]] = {}
                    show_cursor_by_key: dict[str, int] = {}
                    for row in ordered_oto_rows:
                        show_key = str(oto_auto_show_by_slot[str(row["slot_id"])])
                        show_display = _display_name_for_archive_pick(cfg, show_key)
                        if show_key not in show_pool_cache:
                            pool = _episode_rows_for_archive_pick(cfg, show_key, nikki)
                            pool = sorted(pool, key=lambda r: int(r.get("idx0", 0)))
                            show_pool_cache[show_key] = pool
                            sd_auto = _showdef_for_archive_pick(cfg, show_key)
                            start_cursor = int(getattr(sd_auto, "start_episode_index", 0) or 0) if sd_auto else 0
                            start_idx = 0
                            for i_pool, ep in enumerate(pool):
                                if int(ep.get("idx0", i_pool)) >= start_cursor:
                                    start_idx = i_pool
                                    break
                            show_cursor_by_key[show_key] = start_idx
                        pool = show_pool_cache.get(show_key) or []
                        if pool:
                            cur_idx = int(show_cursor_by_key.get(show_key, 0))
                            ep = pool[cur_idx % len(pool)]
                            show_cursor_by_key[show_key] = cur_idx + 1
                            episode_plan.append(
                                (
                                    str(ep.get("code") or show_display),
                                    _episode_num_text(ep) or str(ep.get("code") or show_display),
                                    str(ep.get("title") or show_display),
                                    show_display,
                                )
                            )
                        else:
                            episode_plan.append((show_display, show_display, show_display, show_display))
                elif oto_fill_mode == "Replace time window with movie list":
                    for title in oto_grid_displays:
                        episode_plan.append(("MOVIE", "MOVIE", title, title))
                else:
                    pool = list(oto_episode_rows)
                    if oto_fill_mode == "Manual: show > season > episode":
                        pool = list(oto_manual_pool)
                    if not pool:
                        st.error("OTO episode pool is empty for this mode.")
                        can_run = False
                    else:
                        pool.sort(key=lambda r: int(r.get("idx0", 0)))
                        if oto_fill_mode == "Manual: show > season > episode" and oto_manual_start_idx is not None:
                            start_idx = int(oto_manual_start_idx)
                            if oto_manual_advance:
                                for i, _ in enumerate(ordered_oto_rows):
                                    ep = pool[(start_idx + i) % len(pool)]
                                    episode_plan.append(
                                        (
                                            str(ep.get("code") or ""),
                                            _episode_num_text(ep),
                                            str(ep.get("title") or ""),
                                            oto_display,
                                        )
                                    )
                            else:
                                ep = pool[start_idx]
                                episode_plan = [
                                    (
                                        str(ep.get("code") or ""),
                                        _episode_num_text(ep),
                                        str(ep.get("title") or ""),
                                        oto_display,
                                    )
                                ] * len(ordered_oto_rows)
                        else:
                            sd_auto = _showdef_for_archive_pick(cfg, oto_pick)
                            start_cursor = int(getattr(sd_auto, "start_episode_index", 0) or 0) if sd_auto else 0
                            start_idx = 0
                            for i, ep in enumerate(pool):
                                if int(ep.get("idx0", i)) >= start_cursor:
                                    start_idx = i
                                    break
                            for i, _ in enumerate(ordered_oto_rows):
                                ep = pool[(start_idx + i) % len(pool)]
                                episode_plan.append(
                                    (
                                        str(ep.get("code") or ""),
                                        _episode_num_text(ep),
                                        str(ep.get("title") or ""),
                                        oto_display,
                                    )
                                )

                for i, row in enumerate(ordered_oto_rows):
                    st_norm = parse_flexible_time(str(row["start"]))
                    fin_norm = parse_flexible_time(str(row["finish"]))
                    ep_code, ep_num, ep_title, row_show = (
                        episode_plan[i] if i < len(episode_plan) else ("", "", "", oto_display)
                    )
                    if not ep_code and oto_fill_mode != "Auto-populate matching genre movie/program":
                        ep_code = oto_display
                    if not ep_num and oto_fill_mode != "Auto-populate matching genre movie/program":
                        ep_num = ep_code
                    if not ep_title:
                        ep_title = row_show or oto_display
                    oto_overrides.append(
                        BingeRowOverride(
                            match_date=row["date"],
                            match_start=st_norm,
                            new_date=row["date"],
                            new_start=st_norm,
                            new_finish=fin_norm,
                            new_episode=ep_code,
                            new_show=row_show or oto_display,
                            new_episode_num=ep_num,
                            new_episode_name=ep_title,
                        )
                    )

        if use_mass:
            if not st.session_state.get("build_mass_confirm"):
                st.error("Confirm that mass changes persist to source files before running.")
                can_run = False
            elif not mass_rows:
                st.error("Mass changes are enabled, but no schedule blocks are selected.")
                can_run = False
            elif not mass_pick:
                st.error("Mass changes are enabled, but no replacement show was selected.")
                can_run = False

        if use_mass and can_run and mass_pick:
            mass_total_changed = 0
            mass_messages: list[str] = []
            for row in mass_rows:
                ok, msgs = apply_show_swap(
                    cfg_path,
                    [str(row["show"])],
                    mass_pick,
                    schedule_anchor={"date": row["date"], "start": row["start"]},
                )
                mass_messages.extend(msgs)
                if ok:
                    mass_total_changed += 1
                else:
                    can_run = False
            if mass_total_changed:
                st.success(f"Mass source changes applied to **{mass_total_changed}** selected block(s).")
            for m in mass_messages:
                if "no grid change" in str(m).casefold():
                    st.info(m)
                elif "could not" in str(m).casefold() or "missing" in str(m).casefold():
                    st.warning(m)
                else:
                    st.caption(m)
            if can_run:
                cfg = load_build_config(cfg_path)

        if not can_run:
            return

        try:
            created_grids = ensure_grids_workbooks_for_weeks(selected_weeks)
        except (OSError, ValueError) as e:
            st.error(f"Could not create missing grids workbook(s): {e}")
            created_grids = []
        if created_grids:
            st.success(
                "Created **grids** workbook shell(s). On export, the app copies the **previous month's** "
                "Mon-Sun program into blank weeks when that month is in your base schedule (e.g. April to May):\n"
                + "\n".join(f"- `{p}`" for p in created_grids)
            )
        missing_grids: list[str] = []
        for w in selected_weeks:
            if not Path(w.grids_file).is_file():
                missing_grids.append(str(Path(w.grids_file)))
        if missing_grids:
            st.error("Grids file not found:\n" + "\n".join(f"- `{p}`" for p in missing_grids))
        else:
            out_dir = Path(tempfile.mkdtemp(prefix="binge_out_"))
            station_kw: Optional[List[str]] = None
            if stations_input.strip():
                station_kw = [x.strip() for x in stations_input.split(",") if x.strip()]
            bootstrap_prev_week_df: Optional[pd.DataFrame] = None
            prev_binge_path = st.session_state.get("binge_path")
            if prev_binge_path and selected_weeks:
                prev_path = Path(str(prev_binge_path))
                if prev_path.is_file():
                    first_mon = parse_monday(selected_weeks[0].monday)
                    prior_mon = first_mon - timedelta(days=7)
                    try:
                        prior_sheets = read_binge_workbook_sheets(prev_path)
                        for _name, sdf in prior_sheets.items():
                            by_mon = split_binge_df_by_monday(sdf)
                            if prior_mon in by_mon:
                                bootstrap_prev_week_df = by_mon[prior_mon]
                                break
                    except Exception:
                        bootstrap_prev_week_df = None
            try:
                with st.spinner("Working…"):
                    binge_path, grids_path, ovw, seeded = export_both(
                        cfg,
                        out_dir,
                        weeks=selected_weeks,
                        binge_row_overrides=oto_overrides or None,
                        binge_ui_notes={
                            "Schedule window": (
                                f"start={start_date.isoformat()} · weeks={len(selected_weeks)}"
                            ),
                            "Schedule changes": (
                                f"{len(oto_overrides)} row override(s)"
                                if oto_overrides
                                else "none"
                            ),
                            "Mass changes": (
                                f"{len(mass_rows)} slot swap(s) persisted to source"
                                if use_mass and mass_rows
                                else "none"
                            ),
                            "Schedule timing notes": (
                                " | ".join(oto_timing_notes[:6]) if oto_timing_notes else "none"
                            ),
                        },
                        export_stations=station_kw,
                        bootstrap_prev_week_df=bootstrap_prev_week_df,
                    )
            except Exception as e:
                st.error(str(e))
                st.exception(e)
            else:
                if oto_rows and oto_grid_displays:
                    if len(set(oto_grid_displays)) == 1:
                        for msg in _apply_output_grid_slot_replacements(grids_path, oto_rows, oto_grid_displays[0]):
                            st.info(msg)
                    else:
                        for msg in _apply_output_grid_slot_replacements_multi(grids_path, sorted(oto_rows, key=lambda r: (r["date_iso"], int(r["start_slot"]))), oto_grid_displays):
                            st.info(msg)
                for note in oto_timing_notes:
                    st.warning(note)
                st.session_state["binge_path"] = binge_path
                st.session_state["grids_path"] = grids_path
                st.session_state["out_dir"] = out_dir
                for s in seeded:
                    if is_verbose_seed_noise(s):
                        continue
                    if s.startswith("Copied"):
                        st.success(s)
                    elif s.startswith("Archived BINGE reference copy") or s.startswith("Station copy ["):
                        st.success(s)
                    elif any(
                        x in s.lower()
                        for x in ("could not", "cannot load", "missing", "skipping", "no program", "no ``weeks")
                    ):
                        st.warning(s)
                for w in ovw:
                    st.warning(w)
                built_months = sorted(
                    {
                        (parse_monday(w.monday).year, parse_monday(w.monday).month)
                        for w in selected_weeks
                    }
                )
                for y, m in built_months:
                    _record_completed_month(cfg_path, date(y, m, 1))
                next_start = _next_week_start_after_selection(selected_weeks, buildable_weeks)
                if next_start is not None:
                    st.session_state["_pending_schedule_start_date"] = next_start
                    st.session_state["_pending_schedule_week_count"] = 1
                    st.session_state["_pending_auto_advance_msg"] = (
                        f"Next week auto-selected: `{next_start.isoformat()}`"
                    )
                    st.rerun()

    if "binge_path" in st.session_state and "grids_path" in st.session_state:
        st.markdown("##### Latest files")
        _render_last_build_outputs(cfg, cfg_path)
        _render_binge_grids_preview(key_prefix="build", show_swap=False)


def main() -> None:
    st.set_page_config(
        page_title="Schedule Builder",
        layout="centered",
        initial_sidebar_state="collapsed",
    )
    _mobile_styles()

    page = _render_top_nav()

    st.divider()
    _render_desktop_download_cta()

    cfg_path = Path(st.session_state["main_setup_yaml"])
    if not cfg_path.is_file():
        st.error(f"Base schedule file not found: `{cfg_path.resolve()}`")
        st.stop()

    cfg = load_build_config(cfg_path)
    nikki_path = resolved_nikki_workbook_path(cfg)

    if page == _NAV_ARCHIVE:
        _render_content_archive(cfg, cfg_path, nikki_path)
    elif page == _NAV_EDIT_SCHEDULE:
        st.header(_NAV_EDIT_SCHEDULE)
        _render_schedule_tab(cfg, cfg_path, nikki_path)
    else:
        st.header(_NAV_BUILD)
        _render_build_schedule(cfg, cfg_path, nikki_path)


if __name__ == "__main__":
    main()
